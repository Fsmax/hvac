# -*- coding: utf-8 -*-
"""Тесты экспорта в HLGC Design Table."""

import os
import tempfile
import pytest

from hvac.project import HVACProject
from hvac.io_hlgc import (
    export_to_hlgc, _get_value_for_field, HLGC_COLUMN_MAP,
)


def _make_template_xlsx(tmp_path):
    """Создаёт минимальную xlsx-имитацию HLGC-таблицы для тестов."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "HLGC"
    # Минимум заголовков (наши строки 2-11 пусты — не проверяются)
    # Данные начинаются с DATA_START_ROW = 12
    # Колонка 3 = номер комнаты
    # Колонка 6 = площадь
    rows = [
        # row, room_num, name_en, name_ru, area
        (13, "B01-001", "Office 1", "Офис 1", 30.0),
        (14, "B01-002", "Office 2", "Офис 2", 45.0),
        (15, "B02-001", "Storage", "Склад", 50.0),
    ]
    for r, num, en, ru, area in rows:
        ws.cell(r, 3).value = num
        ws.cell(r, 4).value = en
        ws.cell(r, 5).value = ru
        ws.cell(r, 6).value = area
        # В col 26 ставим формулу — проверим что не перезаписывается
        ws.cell(r, 26).value = f"=Y{r}/F{r}"
    path = os.path.join(tmp_path, "template.xlsx")
    wb.save(path)
    return path


class TestColumnMap:

    def test_all_keys_are_ints(self):
        for col in HLGC_COLUMN_MAP:
            assert isinstance(col, int) and 1 <= col <= 100

    def test_no_duplicate_columns(self):
        cols = list(HLGC_COLUMN_MAP.keys())
        assert len(cols) == len(set(cols))


class TestGetValueForField:

    def _sp(self, **kwargs):
        from hvac.models import Space
        # Дефолты, которые могут быть переопределены через kwargs
        defaults = dict(area_m2=20, volume_m3=60, height_m=3)
        defaults.update(kwargs)
        return Space(space_id="1", number="x", name="x", level="L1",
                     **defaults)

    def test_direct_field(self):
        sp = self._sp(occupancy_people=5)
        assert _get_value_for_field(sp, "occupancy_people", 14) == 5

    def test_q_per_m2(self):
        sp = self._sp()
        sp.heat_gain_w = 4000
        # 4000/20 = 200
        assert _get_value_for_field(sp, "__q_cool_per_m2", 24) == 200.0

    def test_q_per_m2_zero_area(self):
        sp = self._sp(area_m2=0)
        sp.heat_gain_w = 4000
        assert _get_value_for_field(sp, "__q_cool_per_m2", 24) == 0

    def test_min_fresh_air(self):
        sp = self._sp(occupancy_people=4)
        sp.ventilation_breakdown = {"fresh_air_per_person": 25}
        # 25 × 4 = 100
        assert _get_value_for_field(sp, "__min_fresh_air", 30) == 100

    def test_rounding(self):
        sp = self._sp()
        sp.heat_gain_w = 100
        # 100/20 = 5.0 → округление до 1 знака
        v = _get_value_for_field(sp, "__q_cool_per_m2", 24)
        assert v == 5.0


class TestExport:

    def test_matches_by_room_number(self, tmp_path):
        template = _make_template_xlsx(tmp_path)
        out = os.path.join(tmp_path, "out.xlsx")

        p = HVACProject()
        p.new_empty_project("test", "Tashkent")
        # B01-001 и B01-002 — есть в template
        sp1 = p.add_space("B01-001", "Office", "B01", 30.0)
        sp1.heat_loss_w = 2000
        sp1.heat_gain_w = 3000
        sp1.heat_gain_sensible_w = 2500
        sp1.heat_gain_latent_w = 500
        sp1.supply_m3h = 200

        sp2 = p.add_space("B01-002", "Office 2", "B01", 45.0)
        sp2.heat_loss_w = 3000

        # XXX-999 — нет в template
        sp3 = p.add_space("XXX-999", "Ghost", "X", 10.0)

        stats = export_to_hlgc(p, template, out, engine="openpyxl")
        assert stats["rows_matched"] == 2
        assert stats["rows_total"] == 3
        # B02-001 в template, но не в проекте → unmatched
        assert "B02-001" in stats["rows_unmatched"]
        # Проверим что ячейки записались
        from openpyxl import load_workbook
        wb = load_workbook(out, data_only=False)
        ws = wb["HLGC"]
        # B01-001 в row 13
        assert ws.cell(13, 23).value == 3000  # Q_охл итого
        assert ws.cell(13, 25).value == 2000  # Q_отоп
        assert ws.cell(13, 31).value == 200   # supply
        # B01-002 в row 14
        assert ws.cell(14, 25).value == 3000  # Q_отоп

    def test_preserve_formulas(self, tmp_path):
        """Формулы в исходной таблице (col 26) не перезаписываются."""
        template = _make_template_xlsx(tmp_path)
        out = os.path.join(tmp_path, "out.xlsx")

        p = HVACProject()
        p.new_empty_project("test", "Tashkent")
        sp = p.add_space("B01-001", "Office", "B01", 30.0)
        sp.heat_loss_w = 2000

        export_to_hlgc(p, template, out, preserve_formulas=True,
                        engine="openpyxl")
        from openpyxl import load_workbook
        wb = load_workbook(out, data_only=False)
        ws = wb["HLGC"]
        # col 26 имела формулу =Y13/F13 — должна сохраниться
        assert ws.cell(13, 26).value == "=Y13/F13"

    def test_overwrite_only_empty(self, tmp_path):
        """Режим: записываем только в пустые ячейки."""
        template = _make_template_xlsx(tmp_path)
        # Заполним col 25 для B01-001 вручную
        from openpyxl import load_workbook
        wb = load_workbook(template)
        wb["HLGC"].cell(13, 25).value = 9999    # ручная правка
        wb.save(template)

        out = os.path.join(tmp_path, "out.xlsx")
        p = HVACProject()
        p.new_empty_project("test", "Tashkent")
        sp = p.add_space("B01-001", "Office", "B01", 30.0)
        sp.heat_loss_w = 2000

        export_to_hlgc(p, template, out, overwrite_only_empty=True,
                        engine="openpyxl")
        wb = load_workbook(out, data_only=False)
        ws = wb["HLGC"]
        # col 25 должна остаться 9999 (была не пустой)
        assert ws.cell(13, 25).value == 9999
        # col 31 (supply, была пустой) должна стать значением sp.supply_m3h=0
        # → но 0 пропускается фильтром val==None or val==""
        # → пустая ячейка остаётся пустой
        # Положим supply явно
        sp.supply_m3h = 100
        export_to_hlgc(p, template, out, overwrite_only_empty=True,
                        engine="openpyxl")
        wb = load_workbook(out, data_only=False)
        ws = wb["HLGC"]
        assert ws.cell(13, 31).value == 100

    def test_normalized_matching_case_insensitive(self, tmp_path):
        """Сопоставление номеров без учёта регистра и пробелов."""
        template = _make_template_xlsx(tmp_path)
        out = os.path.join(tmp_path, "out.xlsx")
        p = HVACProject()
        p.new_empty_project("test", "Tashkent")
        # В шаблоне B01-001, в проекте b01-001 с пробелом
        sp = p.add_space(" b01-001 ", "Office", "B01", 30.0)
        sp.heat_loss_w = 1234
        stats = export_to_hlgc(p, template, out, engine="openpyxl")
        assert stats["rows_matched"] == 1
        from openpyxl import load_workbook
        ws = load_workbook(out)["HLGC"]
        assert ws.cell(13, 25).value == 1234
