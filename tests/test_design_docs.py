# -*- coding: utf-8 -*-
"""Acceptance tests for design documents built from Calc response fixtures."""

from __future__ import annotations

from collections import defaultdict
from decimal import Decimal, ROUND_CEILING
import json
from pathlib import Path

import pytest

from hvac.design_docs import (
    DOCX_OUTPUT_NAME,
    ENGINEER_REVIEW_NOTE,
    JSON_OUTPUT_NAME,
    XLSX_OUTPUT_NAME,
    DesignDocs,
    DesignDocsInputError,
    main,
)


FIXTURES = Path(__file__).resolve().parent / "fixtures"
TERMINAL_RESPONSE = FIXTURES / "hvac-terminal-layout.response.json"
ROUTE_RESPONSE = FIXTURES / "hvac-route-network.response.json"


@pytest.fixture
def documents() -> DesignDocs:
    return DesignDocs.from_files(TERMINAL_RESPONSE, ROUTE_RESPONSE)


def _load_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _write_json(path: Path, payload: dict) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _ceil_tenth_m(length_mm: Decimal) -> float:
    return float(
        (length_mm / Decimal("1000")).quantize(
            Decimal("0.1"), rounding=ROUND_CEILING
        )
    )


def test_specification_groups_use_placed_not_selected_quantities(
    documents: DesignDocs,
) -> None:
    actual = [
        (
            row.position,
            row.model,
            row.size,
            row.direction,
            row.quantity,
            row.selected_quantity,
        )
        for row in documents.specification
    ]
    assert actual == [
        (1, "АБН/АБР 400×150", "400×150", "supply", 1, 1),
        (2, "АБН/АБР 400×150", "400×150", "exhaust", 1, 1),
        (3, "РНБ/РНР 3000×400", "3000×400", "supply", 5, 9),
    ]

    terminal = _load_json(TERMINAL_RESPONSE)
    placement_count = sum(
        len(result["placements"]) for result in terminal["results"]
    )
    selected_count = sum(
        pick["nUnits"]
        for result in terminal["results"]
        for pick in (result["supplyPick"], result["exhaustPick"])
        if pick is not None
    )

    assert sum(row.quantity for row in documents.specification) == placement_count == 7
    assert (
        sum(row.selected_quantity for row in documents.specification)
        == selected_count
        == 11
    )


def test_specification_preserves_overflow_issues(documents: DesignDocs) -> None:
    overflow = next(
        row
        for row in documents.specification
        if row.model == "РНБ/РНР 3000×400" and row.direction == "supply"
    )

    assert overflow.quantity == 5
    assert overflow.selected_quantity == 9
    assert overflow.source_spaces == ("space-l-shaped", "space-overflow")
    assert overflow.issue_codes == ("TERMINAL_GRID_OVERFLOW",)
    assert "недоразмещено, см. TERMINAL_GRID_OVERFLOW" in overflow.note
    assert "Insufficient usable polygon" in overflow.note
    assert "L-shaped open office" in overflow.note


def test_overflow_issue_is_attached_only_to_underplaced_direction(
    tmp_path: Path,
) -> None:
    terminal = _load_json(TERMINAL_RESPONSE)
    rectangular = next(
        result
        for result in terminal["results"]
        if result["spaceId"] == "space-rectangular"
    )
    rectangular["supplyPick"]["nUnits"] = 2
    rectangular["issues"] = [
        {
            "code": "TERMINAL_GRID_OVERFLOW",
            "message": "supply: placed 1 of 2 selected terminals",
        }
    ]
    changed_path = tmp_path / "direction-specific-overflow.json"
    _write_json(changed_path, terminal)

    changed = DesignDocs.from_files(changed_path, ROUTE_RESPONSE)
    rows = {
        row.direction: row
        for row in changed.specification
        if row.model == "АБН/АБР 400×150"
    }
    assert "TERMINAL_GRID_OVERFLOW" in rows["supply"].note
    assert rows["exhaust"].note == ""
    assert rows["exhaust"].issue_codes == ()


def test_duct_schedule_matches_segments_rounding_categories_and_totals(
    documents: DesignDocs,
) -> None:
    expected_rows = [
        ("Ø100", "trunk", 1, 2000.0, 2.0),
        ("Ø160", "trunk", 3, 8750.0, 8.8),
        ("Ø160", "branch", 6, 18000.0, 18.0),
        ("Ø160", "connection", 6, 600.0, 0.6),
        ("Ø200", "trunk", 2, 3750.0, 3.8),
        ("Ø250", "trunk", 2, 3500.0, 3.5),
        ("Ø250", "riser", 1, 2100.0, 2.1),
        ("Ø250", "connection", 1, 2000.0, 2.0),
    ]
    actual_rows = [
        (
            row.size_label,
            row.segment_kind,
            row.segment_count,
            row.source_length_mm,
            row.length_m,
        )
        for row in documents.ducts
    ]
    assert actual_rows == expected_rows
    assert [row.position for row in documents.ducts] == list(range(1, 9))
    assert all(row.note == "Без монтажного запаса" for row in documents.ducts)
    assert all(row.unit == "м" for row in documents.ducts)

    route = _load_json(ROUTE_RESPONSE)
    raw_groups: dict[tuple[str, str], Decimal] = defaultdict(Decimal)
    raw_counts: dict[tuple[str, str], int] = defaultdict(int)
    raw_by_size: dict[str, Decimal] = defaultdict(Decimal)
    for segment in route["segments"]:
        size_label = (
            f"Ø{segment['size']['dMm']}"
            if segment["shape"] == "round"
            else f"{segment['size']['wMm']}×{segment['size']['hMm']}"
        )
        key = (size_label, segment["kind"])
        length_mm = Decimal(str(segment["lengthMm"]))
        raw_groups[key] += length_mm
        raw_counts[key] += 1
        raw_by_size[size_label] += length_mm

    for row in documents.ducts:
        key = (row.size_label, row.segment_kind)
        assert row.source_length_mm == float(raw_groups[key])
        assert row.segment_count == raw_counts[key]
        assert row.length_m == _ceil_tenth_m(raw_groups[key])

    assert {
        (row.size_label, row.segment_kind)
        for row in documents.ducts
        if row.segment_kind == "connection"
    } == {("Ø160", "connection"), ("Ø250", "connection")}

    expected_totals = [
        ("Ø100", 2000.0, 2.0),
        ("Ø160", 27350.0, 27.4),
        ("Ø200", 3750.0, 3.8),
        ("Ø250", 7600.0, 7.6),
    ]
    actual_totals = [
        (row.size_label, row.source_length_mm, row.rounded_length_m)
        for row in documents.duct_totals
    ]
    assert actual_totals == expected_totals
    for row in documents.duct_totals:
        assert row.source_length_mm == float(raw_by_size[row.size_label])
        assert row.rounded_length_m == _ceil_tenth_m(raw_by_size[row.size_label])

    payload = documents.as_dict()["ductSchedule"]
    assert payload["reserveAdded"] is False
    assert payload["note"] == "Без монтажного запаса"
    assert payload["totalSourceLengthM"] == 40.7
    assert payload["totalScheduledLengthM"] == 40.8


def test_duct_size_totals_equal_sum_of_displayed_category_rows(
    tmp_path: Path,
) -> None:
    route = _load_json(ROUTE_RESPONSE)
    for segment in route["segments"]:
        if segment["shape"] == "round" and segment["size"]["dMm"] == 250:
            segment["lengthMm"] = 1
    changed_path = tmp_path / "short-duct-groups.json"
    _write_json(changed_path, route)

    changed = DesignDocs.from_files(TERMINAL_RESPONSE, changed_path)
    displayed = sum(
        row.length_m for row in changed.ducts if row.size_label == "Ø250"
    )
    size_total = next(
        row.rounded_length_m
        for row in changed.duct_totals
        if row.size_label == "Ø250"
    )

    assert displayed == pytest.approx(0.3)
    assert size_total == pytest.approx(displayed)


def test_system_summary_is_read_literally_from_route_response(
    documents: DesignDocs,
) -> None:
    assert len(documents.systems) == 1
    system = documents.systems[0]

    assert system.system_id == "SA-01"
    assert system.airflow_m3h == 1080.0
    assert system.fan_pressure_pa == pytest.approx(125.878512153944)
    assert system.critical_path_segment_count == 12
    assert system.critical_path_length_m == 25.2
    assert system.terminal_count == 6

    payload = documents.as_dict()["designNote"]["systems"]
    assert payload == [
        {
            "systemId": "SA-01",
            "airflowM3h": 1080.0,
            "fanPressurePa": pytest.approx(125.878512153944),
            "dictatingBranch": {
                "segmentIds": list(system.critical_path_segment_ids),
                "segmentCount": 12,
                "lengthM": 25.2,
            },
            "terminalCount": 6,
        }
    ]


def test_json_export_is_deterministic(documents: DesignDocs, tmp_path: Path) -> None:
    first = tmp_path / "first.json"
    second = tmp_path / "second.json"

    documents.to_json(first)
    DesignDocs.from_files(TERMINAL_RESPONSE, ROUTE_RESPONSE).to_json(second)

    assert first.read_bytes() == second.read_bytes()
    assert first.read_bytes().endswith(b"\n")
    payload = _load_json(first)
    assert payload["kind"] == "hvac-design-documents"
    assert payload["status"] == "PRELIMINARY"


def test_velocity_norm_lookup_keeps_source_and_status_without_inventing_values(
    documents: DesignDocs,
) -> None:
    velocity_norms = documents.as_dict()["designNote"]["velocityNorms"]
    catalog = velocity_norms["catalog"]
    assert catalog["document"] == "ШНҚ 2.04.05-22"
    assert catalog["edition"]
    assert catalog["velocityEntryStatuses"] == ["unverified"]
    assert velocity_norms["buildingLookup"] == "all"
    assert velocity_norms["lookups"]

    for lookup in velocity_norms["lookups"]:
        assert lookup["lookupStatus"] == "NO_APPLICABLE_VALUE"
        assert lookup["references"] == []
        assert lookup["catalogSource"] == {
            "document": catalog["document"],
            "edition": catalog["edition"],
        }
        assert lookup["catalogStatuses"] == ["unverified"]

    assert ENGINEER_REVIEW_NOTE in velocity_norms["note"]


@pytest.mark.parametrize(
    ("fixture_name", "field", "value", "error_fragment"),
    [
        ("terminal", "kind", "hvac-route-network-response", "kind"),
        ("route", "kind", "hvac-terminal-layout-response", "kind"),
        ("terminal", "status", "READY", "status"),
        ("route", "status", "FAILED", "status"),
    ],
)
def test_foreign_kind_and_non_preliminary_status_are_rejected(
    tmp_path: Path,
    fixture_name: str,
    field: str,
    value: str,
    error_fragment: str,
) -> None:
    terminal_path = TERMINAL_RESPONSE
    route_path = ROUTE_RESPONSE
    source = terminal_path if fixture_name == "terminal" else route_path
    changed = _load_json(source)
    changed[field] = value
    changed_path = tmp_path / f"invalid-{fixture_name}-{field}.json"
    _write_json(changed_path, changed)

    if fixture_name == "terminal":
        terminal_path = changed_path
    else:
        route_path = changed_path

    with pytest.raises(DesignDocsInputError, match=error_fragment):
        DesignDocs.from_files(terminal_path, route_path)


def test_non_utf8_response_is_rejected_as_input_error(tmp_path: Path) -> None:
    invalid = tmp_path / "not-utf8.json"
    invalid.write_bytes(b"\xff\xfe\x00")

    with pytest.raises(DesignDocsInputError, match="Не удалось прочитать"):
        DesignDocs.from_files(invalid, ROUTE_RESPONSE)


def test_xlsx_has_two_expected_tables_without_reserve(
    documents: DesignDocs, tmp_path: Path
) -> None:
    from openpyxl import load_workbook

    output = tmp_path / "design-docs.xlsx"
    documents.to_xlsx(output)
    workbook = load_workbook(output, data_only=True)

    assert workbook.sheetnames == ["Спецификация", "Воздуховоды"]
    specification = workbook["Спецификация"]
    ducts = workbook["Воздуховоды"]
    assert [cell.value for cell in specification[5]] == [
        "Поз.",
        "Наименование",
        "Тип, марка",
        "Кол-во",
        "Ед.",
        "Примечание",
    ]
    assert [cell.value for cell in ducts[5]] == [
        "Поз.",
        "Сечение",
        "Категория",
        "Участков",
        "Длина, м",
        "Ед.",
        "Примечание",
    ]
    assert specification.max_row == 8
    assert ducts.max_row == 18

    duct_body = list(ducts.iter_rows(min_row=6, values_only=True))
    assert {
        row[2]
        for row in duct_body
        if row[2] not in {"Итого по сечению", "ИТОГО ПО ВЕДОМОСТИ"}
    } == {
        "Ствол",
        "Ответвления",
        "Стояки",
        "Подключения",
    }
    assert sum(row[2] == "Подключения" for row in duct_body) == 2
    assert all(row[6] == "Без монтажного запаса" for row in duct_body)
    assert "Без монтажного запаса" in str(ducts["A2"].value)
    section_totals = {
        row[1]: row[4] for row in duct_body if row[2] == "Итого по сечению"
    }
    assert section_totals == {
        "Ø100": 2.0,
        "Ø160": 27.4,
        "Ø200": 3.8,
        "Ø250": 7.6,
    }
    grand_total = next(
        row for row in duct_body if row[2] == "ИТОГО ПО ВЕДОМОСТИ"
    )
    assert grand_total[3] == 22
    assert grand_total[4] == 40.8


def test_docx_contains_exact_engineer_review_phrase(
    documents: DesignDocs, tmp_path: Path
) -> None:
    from docx import Document

    output = tmp_path / "design-docs.docx"
    documents.to_docx(output)
    document = Document(output)

    paragraphs = [paragraph.text for paragraph in document.paragraphs]
    assert "Вентиляция: принятые решения" in paragraphs
    assert f"Важно: {ENGINEER_REVIEW_NOTE}." in paragraphs
    assert ENGINEER_REVIEW_NOTE == "нормативные значения не сверены инженером"
    assert any("Без монтажного запаса" in paragraph for paragraph in paragraphs)
    assert any("ШНҚ 2.04.05-22" in paragraph for paragraph in paragraphs)
    assert any("unverified" in paragraph for paragraph in paragraphs)


def test_cli_succeeds_once_then_refuses_without_changing_outputs(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    output_directory = tmp_path / "result"
    arguments = [
        "--terminal",
        str(TERMINAL_RESPONSE),
        "--route",
        str(ROUTE_RESPONSE),
        "--out",
        str(output_directory),
    ]

    assert main(arguments) == 0
    first_capture = capsys.readouterr()
    output_names = (JSON_OUTPUT_NAME, XLSX_OUTPUT_NAME, DOCX_OUTPUT_NAME)
    assert sorted(path.name for path in output_directory.iterdir()) == sorted(output_names)
    before = {
        name: (output_directory / name).read_bytes()
        for name in output_names
    }
    assert all(before.values())
    assert all(name in first_capture.out for name in output_names)
    assert first_capture.err == ""

    assert main(arguments) == 2
    second_capture = capsys.readouterr()
    after = {
        name: (output_directory / name).read_bytes()
        for name in output_names
    }
    assert after == before
    assert sorted(path.name for path in output_directory.iterdir()) == sorted(output_names)
    assert second_capture.out == ""
    assert "design-docs: error:" in second_capture.err


def test_cli_does_not_publish_partial_directory_on_export_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    capsys: pytest.CaptureFixture[str],
) -> None:
    output_directory = tmp_path / "failed-result"

    def fail_xlsx(_documents: DesignDocs, _path: Path) -> None:
        raise RuntimeError("synthetic xlsx failure")

    monkeypatch.setattr(DesignDocs, "to_xlsx", fail_xlsx)
    result = main(
        [
            "--terminal",
            str(TERMINAL_RESPONSE),
            "--route",
            str(ROUTE_RESPONSE),
            "--out",
            str(output_directory),
        ]
    )

    assert result == 2
    assert not output_directory.exists()
    assert not list(tmp_path.glob(".design-docs-*"))
    assert "synthetic xlsx failure" in capsys.readouterr().err
