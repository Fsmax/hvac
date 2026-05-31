# -*- coding: utf-8 -*-
"""Golden smoke-тест экспорта в Excel.

Фиксирует набор листов и открываемость файла, чтобы рефакторинг
io_excel (разбиение монолита на пакет) не менял внешнее поведение.
"""
import pytest

from hvac.project import HVACProject
from hvac.models import Space
from hvac.io_excel import export_to_excel

pytest.importorskip("openpyxl")
from openpyxl import load_workbook


# Полный набор листов, который выдаёт _rich_project() (зафиксирован до
# рефакторинга io_excel). Любое расхождение — регрессия экспорта.
EXPECTED_SHEETS = [
    "Параметры", "Конструкции", "Теплопотери", "Теплопоступления",
    "Ограждения (наружные)", "Вентиляция", "Сводка по уровням",
    "Конструкции по этажам", "Зоны и системы", "Системы оборудования",
    "Проверки", "ГВС", "Энергопаспорт", "Воздуховоды", "Трубы отопления",
    "Трубы холодоснабжения", "Циркуляц. насосы", "Психрометрика AHU",
    "Насосы и баки", "Радиаторы", "Акустика", "Тёплый пол", "Фанкойлы",
    "VRF", "VRF внутренние блоки",
]


def _rich_project() -> HVACProject:
    """Проект, задействующий максимум листов экспорта."""
    p = HVACProject()
    p.params.apply_city("Ташкент")
    for i in range(6):
        sp = Space(
            space_id=f"r{i}", number=f"B01-{i:03d}",
            name="Офис" if i % 2 else "Санузел",
            level="L1" if i < 3 else "L2",
            area_m2=25, volume_m3=75, height_m=3,
            t_in_heat=20, t_in_cool=24, occupancy_people=2,
            heat_loss_w=1800 + 150 * i, heat_gain_w=2500 + 200 * i,
            heat_gain_sensible_w=1800 + 150 * i,
        )
        p.spaces.append(sp)
        p._space_by_id[sp.space_id] = sp

    for fn in (
        p.calculate_ventilation, p.auto_assign_zones, p.calculate_ahu_loads,
        p.auto_assign_smoke_systems, p.calculate_smoke_loads, p.calculate_dhw,
        p.calculate_energy_passport, p.check_condensation_risk, p.size_ducts,
        p.size_pipes, p.size_cooling_pipes, p.compute_ahu_processes,
        p.design_heating_hydraulics, p.select_radiators_for_all_spaces,
        p.analyze_acoustics_for_ahus, p.design_underfloor_loops,
        p.select_fancoils_for_project,
    ):
        fn()
    p.build_vrf_systems(group_by="all", indoor_family="Кассетный")
    return p


def test_export_creates_expected_sheets(tmp_path):
    out = tmp_path / "report.xlsx"
    export_to_excel(_rich_project(), str(out))
    assert out.exists()
    wb = load_workbook(str(out))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_params_sheet_has_city(tmp_path):
    out = tmp_path / "report.xlsx"
    export_to_excel(_rich_project(), str(out))
    wb = load_workbook(str(out))
    ws = wb["Параметры"]
    values = [c.value for row in ws.iter_rows() for c in row]
    assert any(v == "Ташкент" for v in values if isinstance(v, str))
