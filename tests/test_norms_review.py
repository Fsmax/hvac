# -*- coding: utf-8 -*-
"""Roundtrip-контракт инженерной сверки каталогов норм (ТЗ-22)."""

from __future__ import annotations

import json
from copy import deepcopy
from datetime import datetime, timezone
from importlib.resources import files
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from hvac.norms_review import (
    NormsReviewError,
    apply_review,
    export_catalog,
    main,
)


EXPECTED_HEADERS = (
    "Ключ",
    "Значение/диапазон",
    "Единица",
    "Применимость",
    "Пункт",
    "Страница PDF",
    "noteRu",
    "Текущий статус",
    "Подтверждаю (да/нет/исправить)",
    "Исправленное значение",
)
KEY_HEADER = EXPECTED_HEADERS[0]
VALUE_HEADER = EXPECTED_HEADERS[1]
APPLIES_TO_HEADER = EXPECTED_HEADERS[3]
STATUS_HEADER = EXPECTED_HEADERS[7]
DECISION_HEADER = EXPECTED_HEADERS[8]
CORRECTED_VALUE_HEADER = EXPECTED_HEADERS[9]

YES_KEY = "placement.air_distributor.unregulated_zone.distance"
CORRECT_KEY = "velocity.smoke_pressurization.open_doorway.design"
NO_KEY = "opening.transfer.door_undercut.height.min"
EMPTY_KEY = "opening.transfer.grille.free_area.min"
UNREADABLE_KEY = (
    "placement.exhaust.upper_zone.explosive_non_hydrogen.ceiling_distance"
)


@pytest.fixture
def catalog_path(tmp_path: Path) -> Path:
    """Каждый тест работает только с копией поставляемого каталога."""

    resource = files("hvac.catalogs") / "data" / "shnq_2_04_05_22_ducts.json"
    destination = tmp_path / "shnq_2_04_05_22_ducts.json"
    destination.write_bytes(resource.read_bytes())
    return destination


def _read_catalog(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _entries_by_key(path: Path) -> dict[str, dict]:
    return {entry["key"]: entry for entry in _read_catalog(path)["entries"]}


def _review_sheet(path: Path) -> tuple[object, Worksheet]:
    workbook = load_workbook(path)
    worksheet = workbook.active
    assert worksheet is not None
    return workbook, worksheet


def _columns(worksheet: Worksheet) -> dict[str, int]:
    columns = {
        cell.value: cell.column
        for cell in worksheet[1]
        if isinstance(cell.value, str)
    }
    assert set(EXPECTED_HEADERS) <= set(columns)
    return columns


def _row_for_key(worksheet: Worksheet, key: str) -> int:
    columns = _columns(worksheet)
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row, columns[KEY_HEADER]).value == key:
            return row
    raise AssertionError(f"В экспортной таблице нет ключа {key!r}")


def _set_review(
    worksheet: Worksheet,
    key: str,
    decision: str | None,
    corrected_value: object | None = None,
) -> None:
    columns = _columns(worksheet)
    row = _row_for_key(worksheet, key)
    worksheet.cell(row, columns[DECISION_HEADER]).value = decision
    worksheet.cell(row, columns[CORRECTED_VALUE_HEADER]).value = corrected_value


def _parse_exported_json(value: object) -> object:
    """Экспорт использует JSON и для скаляра, и для структурных полей."""

    return json.loads(str(value))


def _parse_utc_timestamp(value: object) -> datetime:
    assert isinstance(value, str) and value.strip()
    normalized = value.removesuffix("Z") + ("+00:00" if value.endswith("Z") else "")
    parsed = datetime.fromisoformat(normalized)
    assert parsed.tzinfo is not None
    return parsed.astimezone(timezone.utc)


def test_export_contains_every_entry_and_required_review_columns(
    catalog_path: Path,
    tmp_path: Path,
) -> None:
    output_path = tmp_path / "norms-review.xlsx"
    raw = _read_catalog(catalog_path)

    assert export_catalog(catalog_path, output_path) == len(raw["entries"])
    assert output_path.is_file()

    workbook, worksheet = _review_sheet(output_path)
    try:
        headers = [cell.value for cell in worksheet[1]][: len(EXPECTED_HEADERS)]
        assert headers == list(EXPECTED_HEADERS)
        assert worksheet.max_row == len(raw["entries"]) + 1

        columns = _columns(worksheet)
        exported_keys = [
            worksheet.cell(row, columns[KEY_HEADER]).value
            for row in range(2, worksheet.max_row + 1)
        ]
        assert exported_keys == [entry["key"] for entry in raw["entries"]]

        for row in range(2, worksheet.max_row + 1):
            assert worksheet.cell(row, columns[DECISION_HEADER]).value is None
            assert worksheet.cell(row, columns[CORRECTED_VALUE_HEADER]).value is None

        sample = next(entry for entry in raw["entries"] if entry["key"] == NO_KEY)
        row = _row_for_key(worksheet, NO_KEY)
        assert _parse_exported_json(
            worksheet.cell(row, columns[VALUE_HEADER]).value
        ) == sample["range"]
        assert _parse_exported_json(
            worksheet.cell(row, columns[APPLIES_TO_HEADER]).value
        ) == sample["appliesTo"]
        assert worksheet.cell(row, columns["Единица"]).value == sample["unit"]
        assert worksheet.cell(row, columns["Пункт"]).value == sample["source"]["clause"]
        assert (
            worksheet.cell(row, columns["Страница PDF"]).value
            == sample["source"]["pagePdf"]
        )
        assert worksheet.cell(row, columns["noteRu"]).value == sample["noteRu"]
        assert worksheet.cell(row, columns[STATUS_HEADER]).value == sample["status"]
    finally:
        workbook.close()


def test_roundtrip_updates_only_explicit_yes_and_correct_actions(
    catalog_path: Path,
    tmp_path: Path,
) -> None:
    review_path = tmp_path / "norms-review.xlsx"
    before = deepcopy(_entries_by_key(catalog_path))
    total = export_catalog(catalog_path, review_path)

    workbook, worksheet = _review_sheet(review_path)
    _set_review(worksheet, YES_KEY, "  ДА  ")
    _set_review(worksheet, CORRECT_KEY, "исправить", 1.45)
    # Даже заполненное значение не даёт права менять запись при "нет"/пусто.
    _set_review(worksheet, NO_KEY, "нет", 999)
    _set_review(worksheet, EMPTY_KEY, None, 999)
    workbook.save(review_path)
    workbook.close()

    reviewed_at = datetime(2026, 8, 23, 12, 34, 56, tzinfo=timezone.utc)
    summary = apply_review(
        review_path,
        catalog_path,
        "Иванов И.И.",
        now=reviewed_at,
    )

    assert summary.total == total
    assert summary.verified == 2
    assert summary.corrected == 1
    assert summary.unchanged == total - 2

    after = _entries_by_key(catalog_path)

    yes = after[YES_KEY]
    assert set(yes) == set(before[YES_KEY]) | {"verifiedBy", "verifiedAt"}
    assert yes["verifiedBy"] == "Иванов И.И."
    assert _parse_utc_timestamp(yes["verifiedAt"]) == reviewed_at
    expected_yes = before[YES_KEY] | {
        "status": "verified",
        "verifiedBy": yes["verifiedBy"],
        "verifiedAt": yes["verifiedAt"],
    }
    assert yes == expected_yes

    corrected = after[CORRECT_KEY]
    assert set(corrected) == set(before[CORRECT_KEY]) | {"verifiedBy", "verifiedAt"}
    assert corrected["value"] == pytest.approx(1.45)
    assert corrected["source"] == before[CORRECT_KEY]["source"]
    assert corrected["verifiedBy"] == "Иванов И.И."
    assert _parse_utc_timestamp(corrected["verifiedAt"]) == reviewed_at
    expected_corrected = before[CORRECT_KEY] | {
        "value": 1.45,
        "status": "verified",
        "verifiedBy": corrected["verifiedBy"],
        "verifiedAt": corrected["verifiedAt"],
    }
    assert corrected == expected_corrected

    assert after[NO_KEY] == before[NO_KEY]
    assert after[EMPTY_KEY] == before[EMPTY_KEY]
    assert len(after) == len(before)


def test_range_correction_and_unreadable_confirmation_preserve_provenance(
    catalog_path: Path,
    tmp_path: Path,
) -> None:
    review_path = tmp_path / "range-review.xlsx"
    before = deepcopy(_entries_by_key(catalog_path))
    export_catalog(catalog_path, review_path)
    corrected_range = {
        "min": 0.025,
        "max": None,
        "minInclusive": True,
        "maxInclusive": False,
    }

    workbook, worksheet = _review_sheet(review_path)
    _set_review(
        worksheet,
        NO_KEY,
        "исправить",
        json.dumps(corrected_range, ensure_ascii=False, separators=(",", ":")),
    )
    # "Да" подтверждает и явно неоднозначную запись, не выдумывая число.
    _set_review(worksheet, UNREADABLE_KEY, "да")
    workbook.save(review_path)
    workbook.close()

    reviewed_at = datetime(2026, 8, 23, 13, 0, tzinfo=timezone.utc)
    summary = apply_review(
        review_path,
        catalog_path,
        "Инженер ОВ",
        now=reviewed_at,
    )

    assert summary.verified == 2
    assert summary.corrected == 1
    after = _entries_by_key(catalog_path)

    corrected = after[NO_KEY]
    assert "value" not in corrected
    assert corrected["range"] == corrected_range
    assert corrected["source"] == before[NO_KEY]["source"]
    assert corrected["status"] == "verified"

    unreadable = after[UNREADABLE_KEY]
    assert unreadable["value"] is None
    assert unreadable["source"] == before[UNREADABLE_KEY]["source"]
    assert unreadable["noteRu"] == before[UNREADABLE_KEY]["noteRu"]
    assert unreadable["status"] == "verified"
    assert unreadable["verifiedBy"] == "Инженер ОВ"
    assert _parse_utc_timestamp(unreadable["verifiedAt"]) == reviewed_at


@pytest.mark.parametrize(
    ("decision", "corrected_value"),
    [
        ("пожалуй", None),
        ("исправить", None),
    ],
    ids=["unknown-decision", "missing-corrected-value"],
)
def test_apply_validates_the_whole_sheet_before_writing(
    catalog_path: Path,
    tmp_path: Path,
    decision: str,
    corrected_value: object | None,
) -> None:
    review_path = tmp_path / "invalid-review.xlsx"
    export_catalog(catalog_path, review_path)

    workbook, worksheet = _review_sheet(review_path)
    # Валидная подпись в предыдущей строке не должна примениться частично.
    _set_review(worksheet, YES_KEY, "да")
    _set_review(worksheet, CORRECT_KEY, decision, corrected_value)
    workbook.save(review_path)
    workbook.close()

    before_bytes = catalog_path.read_bytes()
    with pytest.raises(NormsReviewError) as error:
        apply_review(review_path, catalog_path, "Инженер")

    assert CORRECT_KEY in str(error.value)
    assert catalog_path.read_bytes() == before_bytes


def test_apply_rejects_tampered_or_duplicated_export_rows(
    catalog_path: Path,
    tmp_path: Path,
) -> None:
    for case in ("tampered-value", "duplicate-key"):
        review_path = tmp_path / f"{case}.xlsx"
        export_catalog(catalog_path, review_path)
        workbook, worksheet = _review_sheet(review_path)
        columns = _columns(worksheet)
        row = _row_for_key(worksheet, CORRECT_KEY)
        if case == "tampered-value":
            worksheet.cell(row, columns[VALUE_HEADER]).value = "999"
        else:
            worksheet.cell(row, columns[KEY_HEADER]).value = YES_KEY
        worksheet.cell(row, columns[DECISION_HEADER]).value = "да"
        workbook.save(review_path)
        workbook.close()

        before_bytes = catalog_path.read_bytes()
        with pytest.raises(NormsReviewError):
            apply_review(review_path, catalog_path, "Инженер")
        assert catalog_path.read_bytes() == before_bytes


def test_apply_rejects_review_exported_from_a_stale_catalog(
    catalog_path: Path,
    tmp_path: Path,
) -> None:
    review_path = tmp_path / "stale-review.xlsx"
    export_catalog(catalog_path, review_path)

    workbook, worksheet = _review_sheet(review_path)
    _set_review(worksheet, YES_KEY, "да")
    workbook.save(review_path)
    workbook.close()

    externally_changed = _read_catalog(catalog_path)
    externally_changed["entries"][0]["noteRu"] += " Внешнее изменение после экспорта."
    catalog_path.write_text(
        json.dumps(externally_changed, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    stale_bytes = catalog_path.read_bytes()

    with pytest.raises(NormsReviewError):
        apply_review(review_path, catalog_path, "Инженер")
    assert catalog_path.read_bytes() == stale_bytes


def test_cli_export_and_apply_print_a_summary(
    catalog_path: Path,
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    review_path = tmp_path / "cli-review.xlsx"

    assert main([
        "export",
        "--catalog",
        str(catalog_path),
        "--out",
        str(review_path),
    ]) == 0
    export_output = capsys.readouterr()
    assert export_output.err == ""
    assert str(review_path) in export_output.out

    workbook, worksheet = _review_sheet(review_path)
    _set_review(worksheet, YES_KEY, "да")
    workbook.save(review_path)
    workbook.close()

    assert main([
        "apply",
        "--catalog",
        str(catalog_path),
        "--in",
        str(review_path),
        "--verified-by",
        "CLI Engineer",
    ]) == 0
    apply_output = capsys.readouterr()
    assert apply_output.err == ""
    assert apply_output.out.strip()
    assert "1" in apply_output.out

    updated = _entries_by_key(catalog_path)[YES_KEY]
    assert updated["status"] == "verified"
    assert updated["verifiedBy"] == "CLI Engineer"
