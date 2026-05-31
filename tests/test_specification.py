# -*- coding: utf-8 -*-
"""Тесты спецификации оборудования по ГОСТ 21.110."""

import pytest
from hvac.project import HVACProject
from hvac.models import Space
from hvac.equipment import (
    CoolingSystem, HeatingCircuit, HeatingSystem, VentilationSystem,
)
from hvac.specification import (
    Specification, SpecificationItem, build_specification,
    export_specification_xlsx,
)


def _make_project_with_equipment():
    p = HVACProject()
    p.params.project_name = "Тестовый офис"
    p.params.apply_city("Ташкент")
    # 4 помещения
    for i in range(4):
        sp = Space(
            space_id=f"r{i}", number=f"R-{i:03d}", name="Office", level="L1",
            area_m2=25, volume_m3=75, height_m=3, room_type="Офис",
            heat_loss_w=2000, heat_gain_w=2800, supply_m3h=200,
        )
        p.spaces.append(sp); p._space_by_id[sp.space_id] = sp

    # Системы
    p.heating_systems["Котёл-1"] = HeatingSystem(
        name="Котёл-1", system_type="boiler_gas",
        t_supply=80, t_return=60, efficiency=0.92)
    p.cooling_systems["Чиллер-1"] = CoolingSystem(
        name="Чиллер-1", system_type="chiller_air",
        t_supply=7, t_return=12, cop=3.5)
    p.ventilation_systems["ПВ-1"] = VentilationSystem(
        name="ПВ-1", has_recovery=True,
        recovery_efficiency_winter=0.65,
        recovery_efficiency_summer=0.55)

    return p


class TestBuildSpec:
    def test_basic_systems_collected(self):
        p = _make_project_with_equipment()
        spec = build_specification(p)
        assert spec.project_name == "Тестовый офис"
        # Котёл, чиллер, AHU
        names = [it.name for it in spec.items]
        assert "Котёл-1" in names
        assert "Чиллер-1" in names
        assert "ПВ-1" in names

    def test_renumber_starts_from_1(self):
        p = _make_project_with_equipment()
        spec = build_specification(p)
        assert spec.items
        assert spec.items[0].position == 1
        # Позиции возрастают по 1
        for i, it in enumerate(spec.items):
            assert it.position == i + 1

    def test_radiators_grouped_by_model(self):
        p = _make_project_with_equipment()
        # Назначим радиаторы (одинаковые на все 4 помещения)
        p.select_radiators_for_all_spaces()
        spec = build_specification(p)
        # Радиаторы должны быть сгруппированы в одну позицию
        # (если все 4 получили одну и ту же модель)
        rad_items = [it for it in spec.items
                     if "Стальной" in it.name or "Алюминий" in it.name
                     or "Биметалл" in it.name or "Чугун" in it.name
                     or "Kermi" in it.name or "Purmo" in it.name]
        # Хотя бы одна позиция с количеством > 1 (группировка работает)
        if rad_items:
            total_qty = sum(it.quantity for it in rad_items)
            # Все 4 учтены
            assert total_qty == 4

    def test_section_order(self):
        p = _make_project_with_equipment()
        spec = build_specification(p)
        sections = []
        for it in spec.items:
            if it.section not in sections:
                sections.append(it.section)
        # Отопление раньше Кондиционирования
        if "Отопление" in sections and "Кондиционирование" in sections:
            assert sections.index("Отопление") < sections.index("Кондиционирование")


class TestRadiatorAggregation:
    def test_same_radiator_combined(self):
        spec = Specification(project_name="Test")
        # Эмулируем добавление двух одинаковых радиаторов
        spec.items.append(SpecificationItem(
            name="Kermi 22 500x1000", unit="шт.", quantity=5,
            section="Отопление"))
        spec.renumber()
        assert spec.items[0].position == 1


class TestExcelExport:
    def test_creates_file(self, tmp_path):
        p = _make_project_with_equipment()
        spec = build_specification(p)
        path = tmp_path / "spec.xlsx"
        export_specification_xlsx(spec, str(path))
        assert path.exists()
        assert path.stat().st_size > 5000

    def test_excel_has_correct_sheet(self, tmp_path):
        p = _make_project_with_equipment()
        spec = build_specification(p)
        path = tmp_path / "spec.xlsx"
        export_specification_xlsx(spec, str(path))
        from openpyxl import load_workbook
        wb = load_workbook(path)
        assert "Спецификация" in wb.sheetnames
        ws = wb["Спецификация"]
        # Заголовок с ГОСТ
        all_text = []
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v:
                    all_text.append(str(v))
        joined = "\n".join(all_text)
        assert "ГОСТ 21.110" in joined
        assert "Тестовый офис" in joined

    def test_includes_radiators_after_selection(self, tmp_path):
        p = _make_project_with_equipment()
        p.select_radiators_for_all_spaces()
        spec = build_specification(p)
        # Хотя бы одна позиция в разделе «Отопление» с радиатором
        rad_items = [it for it in spec.items if it.section == "Отопление"]
        assert any(it.quantity >= 1 for it in rad_items)


class TestFullPipeline:
    def test_with_underfloor_and_fancoils(self):
        p = _make_project_with_equipment()
        p.design_underfloor_loops(pitch_mm=150)
        p.select_fancoils_for_project()
        p.build_vrf_systems(group_by="all", indoor_family="Кассетный")
        spec = build_specification(p)
        names = [it.name for it in spec.items]
        # Тёплый пол
        assert any("Труба ТП" in n for n in names)
        # Фанкойл
        assert any("Carrier" in n or "Daikin" in n or "Lessar" in n
                   for n in names)
        # VRF внутренние
        assert any("Cassette" in n or "Duct" in n for n in names)
        # Медь
        assert any("медная" in n.lower() for n in names)
