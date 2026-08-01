# -*- coding: utf-8 -*-
r"""Раскладка ALL-BLOCKS по 3 листам-башням мастер-таблицы HLGC.

    HTL -> TOWER A,  RES -> TOWER B,  OFF (вкл. OFC) -> TOWER C.

Подиум (B1/B2, ~286 помещ.) отдельным листом НЕ выносится — остаётся только
в сводном листе "HLGC" (решение заказчика).

Движок — openpyxl (COM/Excel на этот файл не открывается headless из-за его
содержимого; ваши скрипты по той же причине используют openpyxl). Лист-башня
не копируется целиком (это медленно), а «засевается» шапкой + одной строкой-
образцом из HLGC; затем export_to_hlgc(mode="rebuild") заполняет его строками
своего блока и клонирует формат образца на все строки. Блок помещения берётся
по токену башни в его УРОВНЕ (io_hlgc.space_block).

Мастер НЕ трогаем — результат в отдельный *_towers.xlsx.

Запуск:  D:\HVAC\.venv\Scripts\python.exe D:\HVAC\tools\export_hlgc_towers.py
"""
import json
import os
import sys
from copy import copy

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, r"D:\HVAC")

from openpyxl import load_workbook

from hvac.project import HVACProject
from hvac.io_json import load_project
from hvac.io_hlgc import (
    export_to_hlgc, HLGC_SHEET_NAME, DATA_START_ROW, space_block,
)

# ------------------------------------------------------------------ конфиг
PROJECT = r"D:\HVAC\ALL-BLOCKS_fixed.hvac.json"
MASTER  = r"D:\Chorsu HLGC Design Table_ALL-BLOCKS.xlsx"          # источник (не меняем)
OUT     = r"D:\Chorsu HLGC Design Table_ALL-BLOCKS_towers.xlsx"   # результат
BLOCK_TO_SHEET = [("HTL", "TOWER A"), ("RES", "TOWER B"), ("OFF", "TOWER C")]

_TEMPLATE_ROW = DATA_START_ROW + 1     # строка-образец формата данных (13)


def _seed_header(wb, src_title: str, dst_title: str) -> None:
    """«Засеивает» лист dst_title шапкой (строки 1..DATA_START_ROW) и одной
    строкой-образцом из src_title: значения + стили + ширины колонок + высоты
    строк + объединения в пределах шапки. Дальше rebuild клонирует образец."""
    src = wb[src_title]
    if dst_title in wb.sheetnames:
        del wb[dst_title]
    dst = wb.create_sheet(dst_title)
    for col, dim in src.column_dimensions.items():
        dst.column_dimensions[col].width = dim.width
        dst.column_dimensions[col].hidden = dim.hidden
    max_col = src.max_column
    for r in range(1, _TEMPLATE_ROW + 1):
        if r in src.row_dimensions:
            dst.row_dimensions[r].height = src.row_dimensions[r].height
        for c in range(1, max_col + 1):
            s = src.cell(r, c)
            d = dst.cell(r, c)
            d.value = s.value
            if s.has_style:
                d._style = copy(s._style)
    for rng in list(src.merged_cells.ranges):
        if rng.max_row <= _TEMPLATE_ROW:
            dst.merge_cells(str(rng))


def main() -> None:
    # 1) проект + свежий расчёт; флаги отапл./охл. восстанавливаем с диска
    #    (recalc не должен их трогать, но на всякий случай — как в других скриптах)
    disk = json.load(open(PROJECT, encoding="utf-8"))
    flags = {(s.get("number") or "").strip().upper():
             (bool(s.get("is_heated", True)), bool(s.get("is_cooled", True)))
             for s in disk.get("spaces", [])}
    p = HVACProject()
    load_project(p, PROJECT)
    p.recalculate()
    for sp in p.spaces:
        f = flags.get((sp.number or "").strip().upper())
        if f:
            sp.is_heated, sp.is_cooled = f
    print(f"проект: {len(p.spaces)} помещ.")
    for block, sheet in BLOCK_TO_SHEET:
        n = sum(1 for sp in p.spaces if space_block(sp) == block)
        print(f"  {block:4} -> {sheet:8}: {n} помещ.")

    # 2) засеять 3 листа-башни шапкой из HLGC
    base, ext = os.path.splitext(OUT)
    seeded = f"{base}.seeded{ext}"
    wb = load_workbook(MASTER, keep_links=False)
    for _block, sheet in BLOCK_TO_SHEET:
        _seed_header(wb, HLGC_SHEET_NAME, sheet)
    wb.save(seeded)
    wb.close()
    print("листы-башни засеяны шапкой из HLGC")

    # 3) залить каждый лист своим блоком (rebuild оставит только его помещения)
    cur = seeded
    temps = [seeded]
    for i, (block, sheet) in enumerate(BLOCK_TO_SHEET):
        nxt = OUT if i == len(BLOCK_TO_SHEET) - 1 else f"{base}.step{i}{ext}"
        res = export_to_hlgc(p, source_path=cur, output_path=nxt,
                             engine="openpyxl", mode="rebuild",
                             sheet_name=sheet, block_prefixes=[block],
                             preserve_formulas=True)
        print(f"  {sheet}: rows={res['rows_matched']} cleared={res['rows_cleared']} "
              f"cells={res['cells_written']}")
        cur = nxt
        if nxt != OUT:
            temps.append(nxt)

    for t in temps:
        try:
            os.unlink(t)
        except OSError:
            pass
    print(f"\nГОТОВО -> {OUT}")


if __name__ == "__main__":
    main()
