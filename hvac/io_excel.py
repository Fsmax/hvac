# -*- coding: utf-8 -*-
"""Экспорт результатов в Excel (6 листов)."""

from __future__ import annotations
from typing import Dict, List
from hvac.project import HVACProject
from hvac.models import Space


def export_to_excel(project: HVACProject, path: str) -> None:
    """Сохраняет результаты расчёта в xlsx."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("Не установлен openpyxl. Выполните: pip install openpyxl")

    wb = Workbook()
    thin = Side(border_style="thin", color="888888")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    sum_fill = PatternFill("solid", fgColor="DCE6F1")
    sum_font = Font(bold=True)

    def style_header(row):
        for c in row:
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = border

    def autofit(ws, n_cols):
        for i in range(1, n_cols + 1):
            letter = get_column_letter(i)
            max_len = 8
            for cell in ws[letter]:
                v = cell.value
                if v is None:
                    continue
                max_len = max(max_len, min(len(str(v)) + 2, 40))
            ws.column_dimensions[letter].width = max_len

    total_loss = sum(s.heat_loss_w for s in project.spaces)
    total_gain = sum(s.heat_gain_w for s in project.spaces)
    total_area = sum(s.area_m2 for s in project.spaces)

    # ===== Параметры =====
    ws = wb.active
    ws.title = "Параметры"
    p = project.params
    ws.append(["Параметр", "Значение"])
    style_header(ws[1])
    for k, v in [
        ("Проект", p.project_name), ("Город", p.city),
        ("Расч. наружная зимой, °C", p.t_out_heating),
        ("Расч. наружная летом, °C", p.t_out_cooling),
        ("Суточная амплитуда летом, K", p.daily_amplitude),
        ("Пиковая солнечная радиация, Вт/м²", p.solar_intensity_w_m2),
        ("ГСОП (база 18°C)", p.gsop_18),
        ("Методика", p.methodology),
        ("Коэффициент инфильтрации k", p.inf_correction_k),
        ("Запас на отопление", p.safety_margin_heating),
        ("Запас на охлаждение", p.safety_margin_cooling),
        ("Всего помещений", len(project.spaces)),
        ("Уникальных типов конструкций", len(project.constructions)),
    ]:
        ws.append([k, v])
    autofit(ws, 2)

    # ===== Конструкции =====
    ws = wb.create_sheet("Конструкции")
    headers = ["Ключ", "Категория", "Семейство", "Тип", "Толщина, мм",
               "U, Вт/(м²·К)", "SHGC", "Примечание"]
    ws.append(headers)
    style_header(ws[1])
    for con in sorted(project.constructions.values(),
                      key=lambda x: (x.category, x.key)):
        ws.append([con.key, con.category, con.family, con.type_name,
                   con.thickness_mm, con.u_value, con.shgc, con.note])
    autofit(ws, len(headers))

    # ===== Теплопотери =====
    ws = wb.create_sheet("Теплопотери")
    cols = ["№", "Имя", "Уровень", "Тип помещения", "Площадь, м²",
            "Объём, м³", "tв, °C", "Угловое",
            "Через ограждения, Вт", "Инфильтрация, Вт", "Бытовые, Вт",
            "ИТОГО, Вт", "Уд., Вт/м²"]
    ws.append(cols)
    style_header(ws[1])
    for sp in project.spaces:
        br = sp.heat_loss_breakdown
        total = br.get("ИТОГО", 0.0)
        ud = (total / sp.area_m2) if sp.area_m2 > 0 else 0
        ws.append([
            sp.number, sp.name, sp.level, sp.room_type,
            round(sp.area_m2, 2), round(sp.volume_m3, 2),
            sp.t_in_heat, "да" if sp.is_corner else "—",
            round(br.get("Через ограждения", 0), 1),
            round(br.get("Инфильтрация", 0), 1),
            round(br.get("Бытовые (−)", 0), 1),
            round(total, 1), round(ud, 1),
        ])
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="ИТОГО").font = sum_font
    ws.cell(row=row, column=5, value=round(total_area, 2)).font = sum_font
    ws.cell(row=row, column=12, value=round(total_loss, 1)).font = sum_font
    ws.cell(row=row, column=13,
            value=round(total_loss / total_area if total_area else 0, 1)).font = sum_font
    for c in ws[row]:
        c.fill = sum_fill
    autofit(ws, len(cols))

    # ===== Теплопоступления =====
    ws = wb.create_sheet("Теплопоступления")
    cols = ["№", "Имя", "Уровень", "Тип помещения", "S, м²",
            "Объём, м³", "tв, °C", "Чел.",
            "Через огражд., Вт", "Солнце, Вт", "Люди, Вт", "Освещ., Вт",
            "Оборуд., Вт", "Вентиляция, Вт",
            "Sensible, Вт", "Latent, Вт",
            "ИТОГО, Вт", "Уд., Вт/м²"]
    ws.append(cols)
    style_header(ws[1])
    for sp in project.spaces:
        br = sp.heat_gain_breakdown
        total = br.get("ИТОГО", 0.0)
        ud = (total / sp.area_m2) if sp.area_m2 > 0 else 0
        ws.append([
            sp.number, sp.name, sp.level, sp.room_type,
            round(sp.area_m2, 2), round(sp.volume_m3, 2),
            sp.t_in_cool, sp.occupancy_people,
            round(br.get("Через ограждения", 0), 1),
            round(br.get("Солнечная радиация", 0), 1),
            round(br.get("Люди", 0), 1),
            round(br.get("Освещение", 0), 1),
            round(br.get("Оборудование", 0), 1),
            round(br.get("Инфильтрация/вентиляция", 0), 1),
            round(sp.heat_gain_sensible_w, 1),
            round(sp.heat_gain_latent_w, 1),
            round(total, 1), round(ud, 1),
        ])
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="ИТОГО").font = sum_font
    ws.cell(row=row, column=5, value=round(total_area, 2)).font = sum_font
    ws.cell(row=row, column=15,
            value=round(sum(s.heat_gain_sensible_w for s in project.spaces), 1)).font = sum_font
    ws.cell(row=row, column=16,
            value=round(sum(s.heat_gain_latent_w for s in project.spaces), 1)).font = sum_font
    ws.cell(row=row, column=17, value=round(total_gain, 1)).font = sum_font
    ws.cell(row=row, column=18,
            value=round(total_gain / total_area if total_area else 0, 1)).font = sum_font
    for c in ws[row]:
        c.fill = sum_fill
    autofit(ws, len(cols))

    # ===== Ограждения =====
    from hvac.parsers import effective_orientation
    tn_offset = getattr(project.params, "true_north_offset_deg", 0.0) or 0.0

    ws = wb.create_sheet("Ограждения (наружные)")
    cols = ["№ помещ.", "Помещение", "Тип элемента", "Конструкция",
            "Площадь чистая, м²", "U, Вт/(м²·К)",
            "Ориентация (raw)", "Эфф. ориентация",
            "Q зимой, Вт", "Q летом, Вт"]
    ws.append(cols)
    style_header(ws[1])
    # Подсказка о повороте True North в заголовке листа
    if tn_offset != 0:
        ws.cell(row=2, column=1).value = (
            f"Применён поворот True North = {tn_offset:+.0f}°. "
            f"«Raw» — как в CSV из Revit. «Эфф.» — после поправки "
            f"(именно эта используется в расчёте солнца).")
        ws.cell(row=2, column=1).font = Font(italic=True, color="A00000")
        ws.merge_cells(start_row=2, start_column=1, end_row=2,
                       end_column=len(cols))

    for el in project.elements:
        if not el.is_exterior or el.net_area_m2 <= 0:
            continue
        sp = project.get_space(el.space_id)
        if not sp:
            continue
        dt_h = sp.t_in_heat - project.params.t_out_heating
        dt_c = (project.params.t_out_cooling - sp.t_in_cool
                + project.params.daily_amplitude * 0.3)
        q_h = el.u_value * el.net_area_m2 * max(dt_h, 0)
        q_c = el.u_value * el.net_area_m2 * dt_c
        # Raw ориентация
        orient_raw = el.orientation or "—"
        if el.orientation_deg is not None:
            orient_raw = f"{el.orientation} ({el.orientation_deg:.0f}°)"
        # Эффективная ориентация с учётом поворота True North
        eff = effective_orientation(el.orientation, el.orientation_deg,
                                     tn_offset)
        if tn_offset != 0 and el.orientation_deg is not None:
            eff_deg = (el.orientation_deg + tn_offset) % 360
            orient_eff = f"{eff} ({eff_deg:.0f}°)"
        else:
            orient_eff = eff or "—"

        ws.append([
            sp.number, sp.name,
            "Стена" if el.row_type == "external_wall" else "Проём",
            el.construction_key,
            round(el.net_area_m2, 2), round(el.u_value, 3),
            orient_raw, orient_eff,
            round(q_h, 1), round(q_c, 1),
        ])
    autofit(ws, len(cols))

    # ===== Вентиляция =====
    ws = wb.create_sheet("Вентиляция")
    cols = ["№", "Имя", "Уровень", "Тип помещения", "S, м²", "V, м³",
            "Чел.", "Норма свеж., м³/ч·чел",
            "Supply, м³/ч", "Exhaust, м³/ч", "Hood, м³/ч",
            "ACH, 1/ч", "Метод расчёта", "Норматив", "Изм. вручную"]
    ws.append(cols)
    style_header(ws[1])
    total_supply = total_exh = total_hood = 0.0
    for sp in project.spaces:
        br = sp.ventilation_breakdown or {}
        total_supply += sp.supply_m3h
        total_exh += sp.exhaust_m3h
        total_hood += sp.hood_m3h
        ws.append([
            sp.number, sp.name, sp.level, sp.room_type,
            round(sp.area_m2, 2), round(sp.volume_m3, 2),
            sp.occupancy_people,
            br.get("fresh_air_per_person", 0),
            round(sp.supply_m3h, 1), round(sp.exhaust_m3h, 1),
            round(sp.hood_m3h, 1),
            round(sp.ach_calculated, 2),
            br.get("method", "") if not sp.vent_user_modified else "—",
            br.get("note", "") if not sp.vent_user_modified else "ручная правка",
            "да" if sp.vent_user_modified else "—",
        ])
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="ИТОГО").font = sum_font
    ws.cell(row=row, column=9, value=round(total_supply, 1)).font = sum_font
    ws.cell(row=row, column=10, value=round(total_exh, 1)).font = sum_font
    ws.cell(row=row, column=11, value=round(total_hood, 1)).font = sum_font
    for c in ws[row]:
        c.fill = sum_fill
    autofit(ws, len(cols))

    # ===== Сводка по уровням =====
    ws = wb.create_sheet("Сводка по уровням")
    cols = ["Уровень", "Кол-во помещений", "Площадь, м²",
            "Теплопотери, Вт", "Теплопоступл., Вт",
            "Уд. потери, Вт/м²", "Уд. поступл., Вт/м²",
            "Σ Supply, м³/ч", "Σ Exhaust, м³/ч",
            "Стекло Revit, шт", "WWR-оценка, шт", "Без остекл., шт",
            "Σ Стекло, м²", "Σ Наруж. стены, м²"]
    ws.append(cols)
    style_header(ws[1])
    by_level: Dict[str, List[Space]] = {}
    for sp in project.spaces:
        by_level.setdefault(sp.level, []).append(sp)
    # Карта space_id → наружные элементы (для подсчёта площадей по этажу)
    ext_elems_by_space: Dict[str, List] = {}
    for el in project.elements:
        if el.is_exterior:
            ext_elems_by_space.setdefault(el.space_id, []).append(el)
    for lvl in sorted(by_level.keys()):
        items = by_level[lvl]
        a = sum(s.area_m2 for s in items)
        ll = sum(s.heat_loss_w for s in items)
        gg = sum(s.heat_gain_w for s in items)
        sup = sum(s.supply_m3h for s in items)
        exh = sum(s.exhaust_m3h for s in items)
        n_real = sum(1 for s in items
                     if getattr(s, "glazing_source", "none") == "real")
        n_wwr = sum(1 for s in items
                    if getattr(s, "glazing_source", "none") == "wwr")
        n_none = sum(1 for s in items
                     if getattr(s, "glazing_source", "none") == "none")
        # Σ стекла и Σ наружных стен по этажу (по фактической геометрии CSV).
        # Если фасад одинаковый, эти суммы должны совпадать на всех типовых
        # этажах — расхождение указывает на потерю элементов в выгрузке Revit.
        glass_area = 0.0
        wall_area = 0.0
        for s in items:
            for el in ext_elems_by_space.get(s.space_id, []):
                con = project.constructions.get(el.construction_key)
                is_glazed = con is not None and con.shgc > 0
                if is_glazed:
                    glass_area += el.net_area_m2
                elif el.row_type == "external_wall":
                    wall_area += el.net_area_m2
        ws.append([
            lvl, len(items), round(a, 2),
            round(ll, 1), round(gg, 1),
            round(ll / a if a else 0, 1), round(gg / a if a else 0, 1),
            round(sup, 0), round(exh, 0),
            n_real, n_wwr, n_none,
            round(glass_area, 1), round(wall_area, 1),
        ])
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="ВСЕГО").font = sum_font
    ws.cell(row=row, column=2, value=len(project.spaces)).font = sum_font
    ws.cell(row=row, column=3, value=round(total_area, 2)).font = sum_font
    ws.cell(row=row, column=4, value=round(total_loss, 1)).font = sum_font
    ws.cell(row=row, column=5, value=round(total_gain, 1)).font = sum_font
    for c in ws[row]:
        c.fill = sum_fill
    autofit(ws, len(cols))

    # ===== Конструкции по этажам =====
    # Сводная таблица: для каждой пары (тип конструкции, уровень) —
    # суммарная площадь наружных элементов. Если фасад типовой, но в
    # выгрузке на разных этажах указаны разные типы конструкций
    # (например, на одних этажах витраж, а на других — Basic Wall),
    # «лишний» тип сразу вылезет с площадью на проблемных этажах
    # и нулём на остальных.
    ws = wb.create_sheet("Конструкции по этажам")
    # Соберём pivot: con_key → {level → area}
    by_con_lvl: Dict[str, Dict[str, float]] = {}
    for el in project.elements:
        if not el.is_exterior or el.net_area_m2 <= 0:
            continue
        sp = project.get_space(el.space_id)
        if not sp:
            continue
        key = el.construction_key or "<пусто>"
        by_con_lvl.setdefault(key, {}).setdefault(sp.level, 0.0)
        by_con_lvl[key][sp.level] += el.net_area_m2
    levels_sorted = sorted(by_level.keys())
    header = ["Конструкция", "SHGC", "U, Вт/(м²·К)", "Σ всего, м²"] + levels_sorted
    ws.append(header)
    style_header(ws[1])
    # Сортируем по убыванию суммарной площади
    rows_data = []
    for key, lvl_map in by_con_lvl.items():
        total = sum(lvl_map.values())
        con = project.constructions.get(key)
        shgc = con.shgc if con else ""
        u_val = con.u_value if con else ""
        rows_data.append((key, shgc, u_val, total, lvl_map))
    rows_data.sort(key=lambda r: -r[3])
    for key, shgc, u_val, total, lvl_map in rows_data:
        row = [key, shgc, u_val, round(total, 1)]
        for lvl in levels_sorted:
            row.append(round(lvl_map.get(lvl, 0.0), 1) if lvl in lvl_map else "")
        ws.append(row)
    autofit(ws, len(header))

    # ===== Зоны и системы =====
    has_any_zones = any(sp.system_heating or sp.system_cooling
                        or sp.system_ventilation for sp in project.spaces)
    if has_any_zones:
        from hvac.project import (suggest_ahu_size, suggest_boiler_size,
                                   suggest_chiller_size)
        ws = wb.create_sheet("Зоны и системы")
        ws.append(["СВОДКА ПО ЗОНАМ ДЛЯ ПОДБОРА ОБОРУДОВАНИЯ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.append([])

        # Отопление
        ws.append(["ОТОПЛЕНИЕ (для подбора котлов)"])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        ws.append(["Зона / котёл", "Помещ.", "Площадь, м²",
                   "Q, кВт", "Уд., Вт/м²", "Типовой котёл"])
        style_header(ws[ws.max_row])
        s = project.get_zone_summary("heating")
        total_h = 0
        for zone, d in sorted(s.items(), key=lambda x: -x[1]["q_heating_w"]):
            ud = d["q_heating_w"] / d["area_m2"] if d["area_m2"] else 0
            ws.append([zone, d["n_spaces"], round(d["area_m2"], 1),
                       round(d["q_heating_w"] / 1000, 2), round(ud, 1),
                       suggest_boiler_size(d["q_heating_w"])])
            total_h += d["q_heating_w"]
        ws.append(["ИТОГО", "", "", round(total_h / 1000, 2), "",
                   suggest_boiler_size(total_h)])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        for c in ws[ws.max_row]:
            c.fill = sum_fill
        ws.append([])

        # Охлаждение
        ws.append(["ОХЛАЖДЕНИЕ (для подбора чиллеров)"])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        ws.append(["Зона / чиллер", "Помещ.", "Площадь, м²",
                   "Sensible, кВт", "Latent, кВт", "Total, кВт",
                   "Уд., Вт/м²", "Типовой чиллер"])
        style_header(ws[ws.max_row])
        s = project.get_zone_summary("cooling")
        for zone, d in sorted(s.items(), key=lambda x: -x[1]["q_cooling_w"]):
            ud = d["q_cooling_w"] / d["area_m2"] if d["area_m2"] else 0
            ws.append([zone, d["n_spaces"], round(d["area_m2"], 1),
                       round(d["q_sensible_w"] / 1000, 2),
                       round(d["q_latent_w"] / 1000, 2),
                       round(d["q_cooling_w"] / 1000, 2),
                       round(ud, 1),
                       suggest_chiller_size(d["q_cooling_w"])])
        ws.append([])

        # Вентиляция
        ws.append(["ВЕНТИЛЯЦИЯ (для подбора AHU)"])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        ws.append(["Установка", "Помещ.", "Площадь, м²",
                   "Supply, м³/ч", "Exhaust, м³/ч", "Hood, м³/ч",
                   "Типовой AHU"])
        style_header(ws[ws.max_row])
        s = project.get_zone_summary("ventilation")
        for zone, d in sorted(s.items(), key=lambda x: -x[1]["supply_m3h"]):
            ws.append([zone, d["n_spaces"], round(d["area_m2"], 1),
                       round(d["supply_m3h"], 0),
                       round(d["exhaust_m3h"], 0),
                       round(d["hood_m3h"], 0),
                       suggest_ahu_size(d["supply_m3h"]) + " м³/ч"])
        autofit(ws, 8)

    # ===== Системы оборудования =====
    has_equipment = (project.ventilation_systems or project.heating_systems
                     or project.cooling_systems)
    if has_equipment:
        ws = wb.create_sheet("Системы оборудования")
        ws.append(["ПАРАМЕТРЫ СИСТЕМ ОБОРУДОВАНИЯ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.append([])

        # Приточные установки
        if project.ventilation_systems:
            ws.append(["ПРИТОЧНЫЕ УСТАНОВКИ (AHU)"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Тип", "Рекуп.", "η зимой", "η летом",
                       "tпод зим °C", "tпод лет °C", "wпод лет г/кг",
                       "Примечание"])
            style_header(ws[ws.max_row])
            for name, ahu in sorted(project.ventilation_systems.items()):
                ws.append([
                    ahu.name, ahu.system_type,
                    "да" if ahu.has_recovery else "—",
                    f"{ahu.recovery_efficiency_winter*100:.0f}%" if ahu.has_recovery else "—",
                    f"{ahu.recovery_efficiency_summer*100:.0f}%" if ahu.has_recovery else "—",
                    ahu.t_supply_winter, ahu.t_supply_summer,
                    ahu.w_supply_summer, ahu.note,
                ])

            # Рассчитанные нагрузки на AHU
            ws.append([])
            ws.append(["НАГРУЗКИ ОТ ПРИТОЧКИ (расчёт)"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Установка", "L подача, м³/ч",
                       "Q калорифер, кВт", "Q охлад. явная, кВт",
                       "Q охлад. скрытая, кВт", "Q охлад. total, кВт"])
            style_header(ws[ws.max_row])
            loads = project.calculate_ahu_loads()
            sum_h = sum_cs = sum_cl = 0
            for name, d in sorted(loads.items(),
                                   key=lambda x: -x[1]["q_heater_w"]):
                if d["supply_m3h"] < 1:
                    continue
                ws.append([
                    name, round(d["supply_m3h"], 0),
                    round(d["q_heater_w"] / 1000, 2),
                    round(d["q_cooler_sens_w"] / 1000, 2),
                    round(d["q_cooler_lat_w"] / 1000, 2),
                    round(d["q_cooler_total_w"] / 1000, 2),
                ])
                sum_h += d["q_heater_w"]
                sum_cs += d["q_cooler_sens_w"]
                sum_cl += d["q_cooler_lat_w"]
            ws.append(["ИТОГО", "", round(sum_h/1000, 2),
                       round(sum_cs/1000, 2), round(sum_cl/1000, 2),
                       round((sum_cs+sum_cl)/1000, 2)])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            for c in ws[ws.max_row]:
                c.fill = sum_fill
            ws.append([])

        # Котлы
        if project.heating_systems:
            ws.append(["КОТЛЫ / ИСТОЧНИКИ ТЕПЛА"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Тип", "tпод °C", "tобр °C",
                       "Топливо", "КПД", "Примечание"])
            style_header(ws[ws.max_row])
            for name, h in sorted(project.heating_systems.items()):
                ws.append([
                    h.name, h.system_type, h.t_supply, h.t_return,
                    h.fuel, f"{h.efficiency*100:.0f}%", h.note,
                ])
            ws.append([])

        # Чиллеры
        if project.cooling_systems:
            ws.append(["ЧИЛЛЕРЫ / ИСТОЧНИКИ ХОЛОДА"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Тип", "tпод °C", "tобр °C",
                       "COP", "Хладагент", "Примечание"])
            style_header(ws[ws.max_row])
            for name, c in sorted(project.cooling_systems.items()):
                ws.append([
                    c.name, c.system_type, c.t_supply, c.t_return,
                    c.cop, c.refrigerant, c.note,
                ])
        autofit(ws, 9)

    # ===== Дымоудаление и подпор воздуха =====
    if project.smoke_systems:
        ws = wb.create_sheet("Дымоудаление")
        ws.append(["СИСТЕМЫ ДЫМОУДАЛЕНИЯ И ПОДПОРА ВОЗДУХА (КМК / СП 7.13130)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.append([])

        # Расчёт нагрузок (single zone — стандартный сценарий)
        loads = project.calculate_smoke_loads(fire_mode="single_zone")

        # СДУ
        smoke_systems = {k: v for k, v in loads.items()
                         if v["system_type"] == "smoke_removal"}
        if smoke_systems:
            ws.append(["СИСТЕМЫ ДЫМОУДАЛЕНИЯ (СДУ)"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Назначение", "Площадь м²", "Зон",
                       "Норма, м³/ч·м²", "L зоны м³/ч",
                       "L системы м³/ч", "L компенс. м³/ч",
                       "t дыма °C", "Огнестойк.", "Примечание"])
            style_header(ws[ws.max_row])
            total_smoke = total_makeup = 0
            for name, d in sorted(smoke_systems.items()):
                ws.append([
                    name, d["purpose"], round(d["served_area_m2"], 0),
                    d["n_zones"], d.get("norm_per_m2", "—"),
                    round(d["L_per_zone_m3h"], 0),
                    round(d["L_smoke_m3h"], 0),
                    round(d["L_makeup_m3h"], 0),
                    d["t_smoke_C"], d["fire_rating"], d["note"],
                ])
                total_smoke += d["L_smoke_m3h"]
                total_makeup += d["L_makeup_m3h"]
            ws.append(["ИТОГО (один пожар)", "", "", "", "", "",
                       round(total_smoke, 0), round(total_makeup, 0),
                       "", "", ""])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            for c in ws[ws.max_row]:
                c.fill = sum_fill
            ws.append([])

        # СПВ
        pres_systems = {k: v for k, v in loads.items()
                        if v["system_type"] == "air_supply"}
        if pres_systems:
            ws.append(["СИСТЕМЫ ПОДПОРА ВОЗДУХА (СПВ)"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Назначение", "L подпора, м³/ч",
                       "Давление, Па", "Примечание"])
            style_header(ws[ws.max_row])
            total_pres = 0
            for name, d in sorted(pres_systems.items()):
                ws.append([
                    name, d["purpose"],
                    round(d["L_smoke_m3h"], 0),
                    d["pressure_pa"], d["note"],
                ])
                total_pres += d["L_smoke_m3h"]
            ws.append(["ИТОГО СПВ", "", round(total_pres, 0), "", ""])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            for c in ws[ws.max_row]:
                c.fill = sum_fill
        autofit(ws, 11)

    # ===== Проверки данных =====
    validation_results = project.validate_detailed()
    if validation_results:
        ws = wb.create_sheet("Проверки")
        cols = ["№", "Серьёзность", "Категория", "Сообщение", "Помещение"]
        ws.append(cols)
        style_header(ws[1])

        severity_fill = {
            "error":   PatternFill("solid", fgColor="F4B7B6"),
            "warning": PatternFill("solid", fgColor="FFE69A"),
            "info":    PatternFill("solid", fgColor="D9E1F2"),
        }
        severity_text = {"error": "Ошибка", "warning": "Внимание", "info": "Инфо"}

        for i, r in enumerate(validation_results, start=1):
            sp = project.get_space(r.get("space_id", ""))
            sp_label = f"{sp.number} {sp.name}" if sp else "—"
            row_num = ws.max_row + 1
            ws.append([
                i,
                severity_text.get(r["severity"], r["severity"]),
                r["category"],
                r["msg"],
                sp_label,
            ])
            # Раскрасить строку по серьёзности
            fill = severity_fill.get(r["severity"])
            if fill:
                for c in ws[row_num]:
                    c.fill = fill
        autofit(ws, len(cols))

    # ====================================================================
    # ===== Расширения v3.7 =====
    # ====================================================================

    # ===== Лист "ГВС" =====
    if project.dhw_systems:
        ws = wb.create_sheet("ГВС")
        ws.append(["ГОРЯЧЕЕ ВОДОСНАБЖЕНИЕ (СП 30.13330.2020)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        cols = ["Система", "Тип нагревателя", "Потребителей",
                "V сут, м³/сут", "V час макс, м³/ч",
                "Q пик, кВт", "Q с цирк., кВт", "Q нагревателя, кВт",
                "V бака, м³", "Циркуляция", "η нагр."]
        ws.append(cols)
        style_header(ws[ws.max_row])

        for name, sys in sorted(project.dhw_systems.items()):
            ws.append([
                name,
                sys.heater_type,
                sys.n_consumers,
                round(sys.v_daily_total_m3, 2),
                round(sys.v_hourly_max_m3, 2),
                round(sys.q_peak_w / 1000, 1),
                round(sys.q_with_circulation_w / 1000, 1),
                round(sys.q_heater_size_w / 1000, 1),
                round(sys.storage_recommended_m3, 2),
                "Да" if sys.has_circulation else "Нет",
                round(sys.efficiency, 2),
            ])

        # Итого
        total_v_d = sum(s.v_daily_total_m3 for s in project.dhw_systems.values())
        total_v_h = sum(s.v_hourly_max_m3 for s in project.dhw_systems.values())
        total_q = sum(s.q_heater_size_w for s in project.dhw_systems.values())
        ws.append(["ИТОГО", "", "",
                   round(total_v_d, 2), round(total_v_h, 2),
                   "", "", round(total_q / 1000, 1), "", "", ""])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        for c in ws[ws.max_row]:
            c.fill = sum_fill
        autofit(ws, len(cols))

    # ===== Лист "Энергопаспорт" =====
    if project.energy_passport:
        ep = project.energy_passport
        ws = wb.create_sheet("Энергопаспорт")
        ws.append(["ЭНЕРГЕТИЧЕСКИЙ ПАСПОРТ (СП 50.13330)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        sections = [
            ("Исходные данные", [
                ("Объект", ep.project_name),
                ("Город", ep.city),
                ("Тип здания", ep.building_type),
                ("ГСОП (база +18°C), °C·сут", ep.gsop_18),
                ("Расчётная зимняя tн, °C", ep.t_out_heating),
                ("Отапливаемая площадь, м²", round(ep.total_area_m2, 1)),
                ("Объём, м³", round(ep.total_volume_m3, 0)),
                ("Помещений", ep.n_spaces),
            ]),
            ("Пиковые нагрузки", [
                ("Q отопления, кВт", round(ep.q_peak_heating_w / 1000, 1)),
                ("Q охлаждения, кВт", round(ep.q_peak_cooling_w / 1000, 1)),
                ("Q нагрев приточек, кВт", round(ep.q_peak_ventilation_heating_w / 1000, 1)),
                ("Q ГВС, кВт", round(ep.q_peak_dhw_w / 1000, 1)),
            ]),
            ("Отопительный сезон", [
                ("Длительность, сут", ep.z_heating_days),
                ("Средняя tн, °C", ep.t_avg_heating),
                ("Коэф. регулирования", ep.k_regulation),
                ("Внутр. теплопост., Вт/м²", ep.internal_gain_w_m2),
            ]),
            ("Годовое потребление, МВт·ч/год", [
                ("Отопление", round(ep.e_heating_kwh_year / 1000, 1)),
                ("Нагрев приточек", round(ep.e_ventilation_kwh_year / 1000, 1)),
                ("Охлаждение (электроэнергия)", round(ep.e_cooling_kwh_year / 1000, 1)),
                ("ГВС", round(ep.e_dhw_kwh_year / 1000, 1)),
                ("Внутр. теплопост. (использованные)",
                 round(ep.e_internal_gains_kwh_year * ep.k_internal_use / 1000, 1)),
            ]),
            ("Удельные показатели", [
                ("qh у удельный, кВт·ч/(м²·год)", round(ep.qh_specific_kwh_m2, 1)),
                ("qh н нормативный, кВт·ч/(м²·год)", round(ep.qh_normative_kwh_m2, 1)),
                ("Отклонение от нормы, %", round(ep.deviation_percent, 1)),
            ]),
        ]
        for title, rows in sections:
            ws.append([title])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=1).fill = sum_fill
            for label, value in rows:
                ws.append([label, value])
            ws.append([])

        # Класс энергоэффективности — крупно
        ws.append(["КЛАСС ЭНЕРГОЭФФЕКТИВНОСТИ",
                   ep.energy_class, ep.energy_class_description])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True, size=12)
            c.fill = sum_fill

        autofit(ws, 3)

    # ===== Лист "Точка росы" =====
    if project.condensation_results:
        ws = wb.create_sheet("Точка росы")
        ws.append(["ПРОВЕРКА ОГРАЖДЕНИЙ НА КОНДЕНСАЦИЮ (СП 50.13330 Прил. Е)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        cols = ["Помещение", "Категория", "Конструкция",
                "U, Вт/(м²·К)", "tв, °C", "tн, °C", "RH вн., %",
                "τ_int, °C", "t_d, °C", "Δt факт., К", "Δt норм., К",
                "Запас до конд., К", "Статус"]
        ws.append(cols)
        style_header(ws[ws.max_row])

        fill_cond = PatternFill("solid", fgColor="F4B7B6")  # красный
        fill_norm = PatternFill("solid", fgColor="FFE69A")  # жёлтый

        for c in project.condensation_results:
            row_num = ws.max_row + 1
            status = ("КОНДЕНСАТ" if c.condensation_risk
                      else "Норматив" if c.normative_fail else "OK")
            ws.append([
                f"{c.space_number} {c.space_name}",
                c.category, c.construction_key,
                round(c.u_value, 2),
                round(c.t_in, 1), round(c.t_out, 1),
                round(c.rh_in, 0),
                round(c.t_surface, 2),
                round(c.t_dew, 2),
                round(c.dt_actual, 2),
                round(c.dt_normative, 1),
                round(c.margin_to_dew, 2),
                status,
            ])
            if c.condensation_risk:
                for cell in ws[row_num]:
                    cell.fill = fill_cond
            elif c.normative_fail:
                for cell in ws[row_num]:
                    cell.fill = fill_norm

        # Сводка
        from hvac.dew_point import total_problems
        prob = total_problems(project.condensation_results)
        ws.append([])
        ws.append(["ИТОГО:", f"всего проверено {prob['total']},",
                   f"конденсат {prob['condensation']},",
                   f"норматив СП 50 не выполняется {prob['normative_fail']},",
                   f"OK {prob['ok']}"])
        ws.cell(row=ws.max_row, column=1).font = sum_font

        autofit(ws, len(cols))

    # ===== Лист "Воздуховоды" =====
    if project.duct_networks:
        ws = wb.create_sheet("Воздуховоды")
        ws.append(["ПОДБОР СЕЧЕНИЙ ВОЗДУХОВОДОВ (СП 60.13330)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        for sys_name, net in sorted(project.duct_networks.items()):
            ws.append([f"СИСТЕМА: {sys_name}",
                       f"Σ Q = {net.total_flow_m3h:.0f} м³/ч",
                       f"Δp подбора AHU = {net.total_pressure_loss_pa:.0f} Па",
                       f"Точек = {net.n_terminals}"])
            for c in ws[ws.max_row]:
                c.font = sum_font
                c.fill = sum_fill
            ws.append(["Участок", "Тип", "Роль", "L, м³/ч",
                       "Форма", "Размер, мм", "v, м/с", "d_гидр, мм",
                       "Δp_тр, Па", "Δp_мест, Па", "Δp_сум, Па", "Примечание"])
            style_header(ws[ws.max_row])

            for sec in net.sections:
                size_str = (f"Ø{int(sec.diameter_mm)}" if sec.shape == "round"
                            else f"{int(sec.width_mm)}×{int(sec.height_mm)}")
                ws.append([
                    sec.id, sec.section_type, sec.role,
                    round(sec.flow_m3h, 0), sec.shape, size_str,
                    round(sec.velocity_m_s, 2),
                    round(sec.hydraulic_diameter_mm, 0),
                    round(sec.pressure_loss_friction_pa, 1),
                    round(sec.pressure_loss_local_pa, 1),
                    round(sec.pressure_loss_total_pa, 1),
                    sec.note,
                ])
            ws.append([])

        autofit(ws, 12)

    # ===== Лист "Трубы отопления" =====
    if project.pipe_networks:
        ws = wb.create_sheet("Трубы отопления")
        ws.append(["ПОДБОР СЕЧЕНИЙ ТРУБ ОТОПЛЕНИЯ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        for sys_name, net in sorted(project.pipe_networks.items()):
            pump_head_m = (getattr(net, "pump_head_m", 0.0) or
                           net.total_pressure_loss_pa / (977.7 * 9.81) * 1.3)
            pump_flow = (getattr(net, "pump_flow_m3_h", 0.0) or
                         net.total_flow_kg_h / 977.7)
            pump_model = getattr(net, "pump_model", "") or "—"
            t_label = (f"{net.t_supply_c:.0f}/{net.t_return_c:.0f}"
                       if hasattr(net, "t_supply_c") else f"Δt={net.delta_t_k:.0f}")
            ws.append([
                f"КОНТУР: {sys_name}",
                f"Σ Q = {net.total_heat_load_w / 1000:.1f} кВт",
                f"G = {net.total_flow_kg_h:.0f} кг/ч ({pump_flow:.2f} м³/ч)",
                f"t = {t_label} °C",
                f"Δp = {net.total_pressure_loss_pa / 1000:.1f} кПа",
                f"Насос: {pump_model} (H={pump_head_m:.1f} м)",
            ])
            for c in ws[ws.max_row]:
                c.font = sum_font
                c.fill = sum_fill
            ws.append(["Участок", "Тип", "Q, Вт", "G, кг/ч", "V, м³/ч",
                       "Материал", "DN", "d_вн, мм", "v, м/с",
                       "L, м", "Z (Σζ)", "ΔH, м",
                       "Δp_тр, Па", "Δp_мест, Па", "Δp_сум, Па", "Примечание"])
            style_header(ws[ws.max_row])

            for sec in net.sections:
                ws.append([
                    sec.id, sec.section_type,
                    round(sec.heat_load_w, 0),
                    round(sec.flow_kg_h, 1),
                    round(sec.flow_m3_h, 3),
                    sec.pipe_material,
                    int(sec.dn_mm),
                    round(sec.inner_diameter_mm, 1),
                    round(sec.velocity_m_s, 2),
                    round(sec.length_m, 1),
                    round(getattr(sec, "local_zeta_sum", 0.0), 1),
                    round(getattr(sec, "elevation_m", 0.0), 1),
                    round(sec.pressure_loss_friction_pa, 1),
                    round(sec.pressure_loss_local_pa, 1),
                    round(sec.pressure_loss_total_pa, 1),
                    sec.note,
                ])
            ws.append([])

        autofit(ws, 16)

    # ===== Лист "Трубы холодоснабжения" =====
    if getattr(project, "cooling_pipe_networks", None):
        ws = wb.create_sheet("Трубы холодоснабжения")
        ws.append(["ПОДБОР СЕЧЕНИЙ ТРУБ ХОЛОДОСНАБЖЕНИЯ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        for sys_name, net in sorted(project.cooling_pipe_networks.items()):
            pump_head_m = getattr(net, "pump_head_m", 0.0)
            pump_flow = getattr(net, "pump_flow_m3_h", 0.0)
            pump_model = getattr(net, "pump_model", "") or "—"
            t_label = f"{net.t_supply_c:.0f}/{net.t_return_c:.0f}"
            ws.append([
                f"КОНТУР: {sys_name}",
                f"Σ Q = {net.total_heat_load_w / 1000:.1f} кВт",
                f"G = {net.total_flow_kg_h:.0f} кг/ч ({pump_flow:.2f} м³/ч)",
                f"t = {t_label} °C",
                f"Δp = {net.total_pressure_loss_pa / 1000:.1f} кПа",
                f"Насос: {pump_model} (H={pump_head_m:.1f} м)",
            ])
            for c in ws[ws.max_row]:
                c.font = sum_font
                c.fill = sum_fill
            ws.append(["Участок", "Тип", "Q, Вт", "G, кг/ч", "V, м³/ч",
                       "Материал", "DN", "d_вн, мм", "v, м/с",
                       "L, м", "Z (Σζ)", "ΔH, м",
                       "Δp_тр, Па", "Δp_мест, Па", "Δp_сум, Па", "Примечание"])
            style_header(ws[ws.max_row])

            for sec in net.sections:
                ws.append([
                    sec.id, sec.section_type,
                    round(sec.heat_load_w, 0),
                    round(sec.flow_kg_h, 1),
                    round(sec.flow_m3_h, 3),
                    sec.pipe_material,
                    int(sec.dn_mm),
                    round(sec.inner_diameter_mm, 1),
                    round(sec.velocity_m_s, 2),
                    round(sec.length_m, 1),
                    round(getattr(sec, "local_zeta_sum", 0.0), 1),
                    round(getattr(sec, "elevation_m", 0.0), 1),
                    round(sec.pressure_loss_friction_pa, 1),
                    round(sec.pressure_loss_local_pa, 1),
                    round(sec.pressure_loss_total_pa, 1),
                    sec.note,
                ])
            ws.append([])
        autofit(ws, 16)

    # ===== Лист "Контуры ИТП" =====
    has_heating_circuits = bool(getattr(project, "heating_circuits", {}))
    has_cooling_circuits = bool(getattr(project, "cooling_circuits", {}))
    if has_heating_circuits or has_cooling_circuits:
        ws = wb.create_sheet("Контуры ИТП")
        ws.append(["КОНТУРЫ ИНДИВИДУАЛЬНОГО ТЕПЛОВОГО ПУНКТА"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        # Heating circuits
        if has_heating_circuits:
            ws.append(["Контуры отопления"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            ws.append(["Контур", "ИТП", "Тип", "t_под, °C", "t_обр, °C",
                       "Смеситель", "Q, кВт", "Расход, м³/ч",
                       "Насос (модель)", "Напор, м", "Привязан AHU", "Примечание"])
            style_header(ws[ws.max_row])
            for cname, circ in sorted(project.heating_circuits.items()):
                net = project.pipe_networks.get(cname)
                q_kw = net.total_heat_load_w / 1000.0 if net else 0.0
                flow = getattr(net, "pump_flow_m3_h", 0.0) if net else 0.0
                pump = getattr(net, "pump_model", "") if net else (circ.pump_model or "—")
                head = getattr(net, "pump_head_m", 0.0) if net else circ.pump_head_m
                ws.append([
                    cname, circ.parent_system, circ.circuit_type,
                    circ.t_supply, circ.t_return,
                    "✓" if circ.has_mixing_node else "—",
                    round(q_kw, 1), round(flow, 2),
                    pump or "—", round(head, 1),
                    circ.serves_ahu or "—", circ.note,
                ])
            ws.append([])

        # Cooling circuits
        if has_cooling_circuits:
            ws.append(["Контуры холодоснабжения"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            ws.append(["Контур", "Источник", "Тип", "t_под, °C", "t_обр, °C",
                       "Изоляция", "Q, кВт", "Расход, м³/ч",
                       "Насос (модель)", "Напор, м", "Привязан AHU", "Примечание"])
            style_header(ws[ws.max_row])
            for cname, circ in sorted(project.cooling_circuits.items()):
                net = project.cooling_pipe_networks.get(cname)
                q_kw = net.total_heat_load_w / 1000.0 if net else 0.0
                flow = getattr(net, "pump_flow_m3_h", 0.0) if net else 0.0
                pump = getattr(net, "pump_model", "") if net else (circ.pump_model or "—")
                head = getattr(net, "pump_head_m", 0.0) if net else circ.pump_head_m
                ws.append([
                    cname, circ.parent_system, circ.circuit_type,
                    circ.t_supply, circ.t_return,
                    "✓" if circ.insulated else "—",
                    round(q_kw, 1), round(flow, 2),
                    pump or "—", round(head, 1),
                    circ.serves_ahu or "—", circ.note,
                ])
            ws.append([])
        autofit(ws, 12)

    # ===== Лист "Циркуляц. насосы" — сводка по всем контурам =====
    all_nets = []
    for n in getattr(project, "pipe_networks", {}).values():
        all_nets.append(("Отопление", n))
    for n in getattr(project, "cooling_pipe_networks", {}).values():
        all_nets.append(("Холод", n))
    if all_nets:
        ws = wb.create_sheet("Циркуляц. насосы")
        ws.append(["ПОДБОР ЦИРКУЛЯЦИОННЫХ НАСОСОВ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Среда", "Контур", "ИТП/источник", "Тип контура",
                   "Q, кВт", "Расход, м³/ч", "Δp сети, кПа",
                   "Напор насоса (с зап. 1.3), м", "Подобранная модель"])
        style_header(ws[ws.max_row])
        for medium_label, net in all_nets:
            pump_head_m = getattr(net, "pump_head_m", 0.0) or (
                net.total_pressure_loss_pa / (977.7 * 9.81) * 1.3)
            pump_flow = (getattr(net, "pump_flow_m3_h", 0.0) or
                         net.total_flow_kg_h / 977.7)
            ws.append([
                medium_label,
                net.system_name,
                getattr(net, "parent_system", "") or "—",
                getattr(net, "circuit_type", "") or "—",
                round(net.total_heat_load_w / 1000.0, 1),
                round(pump_flow, 2),
                round(net.total_pressure_loss_pa / 1000.0, 2),
                round(pump_head_m, 1),
                getattr(net, "pump_model", "") or "—",
            ])
        autofit(ws, 9)

    # ====================================================================
    # ===== Расширения v4.1: подробная инженерия =====
    # ====================================================================

    # ===== Психрометрика AHU =====
    if getattr(project, "ahu_processes", None):
        ws = wb.create_sheet("Психрометрика AHU")
        headers = ["AHU", "Режим", "Точка процесса",
                   "T, °C", "W, г/кг", "RH, %", "H, кДж/кг", "Td, °C"]
        ws.append(headers)
        style_header(ws[1])
        # Сводка мощностей по режимам в начале листа
        for ahu_name, by_mode in project.ahu_processes.items():
            for mode in ("winter", "transitional", "summer"):
                proc = by_mode.get(mode)
                if proc is None:
                    continue
                for point_name, st in proc.points.items():
                    ws.append([
                        ahu_name, mode, point_name,
                        round(st.t_c, 1),
                        round(st.w_g_kg, 2),
                        round(st.rh * 100, 1),
                        round(st.h_kj_kg, 2),
                        round(st.t_dp_c, 1),
                    ])
                # Строка-итог мощностей
                ws.append([
                    ahu_name, mode, "Σ мощность",
                    "", "", "",
                    (f"Калорифер {proc.q_heater_kw:.1f} кВт; "
                     f"Охладитель {proc.q_cooler_total_kw:.1f} кВт "
                     f"(явная {proc.q_cooler_sensible_kw:.1f} / "
                     f"скрытая {proc.q_cooler_latent_kw:.1f}); "
                     f"Увлажнитель {proc.q_humidifier_kw:.1f} кВт"),
                    "",
                ])
                last = ws.max_row
                for c in ws[last]:
                    c.fill = sum_fill
                    c.font = sum_font
        autofit(ws, len(headers))

    # ===== Аэродинамика воздуховодов (детальная) =====
    if getattr(project, "duct_networks_detailed", None):
        ws = wb.create_sheet("Аэродинамика (детально)")
        headers = ["Система", "Участок", "Родитель", "Назначение",
                   "L, м³/ч", "Дл., м", "Форма", "Размер, мм",
                   "v, м/с", "Δp тр., Па", "Δp мест., Па", "Σ Δp, Па"]
        ws.append(headers)
        style_header(ws[1])
        for name, net in project.duct_networks_detailed.items():
            for e in net.edges.values():
                size = (f"Ø{e.diameter_mm:.0f}" if e.shape == "round"
                        else f"{e.width_mm:.0f}×{e.height_mm:.0f}")
                ws.append([
                    name, e.edge_id, e.parent_id or "(корень)",
                    e.terminal_name or "проход",
                    round(e.flow_m3_h, 0),
                    round(e.length_m, 1),
                    e.shape, size,
                    round(e.velocity_m_s, 2),
                    round(e.dp_friction_pa, 1),
                    round(e.dp_local_pa, 1),
                    round(e.dp_total_pa, 1),
                ])
            # Сводка по вентилятору
            ws.append([
                name, "ИТОГО", "", "ветви: " + str(len(net.branches)),
                round(net.fan_flow_m3_h, 0), "", "", "",
                "—",
                "—", "—",
                round(net.fan_pressure_required_pa, 0),
            ])
            last = ws.max_row
            for c in ws[last]:
                c.fill = sum_fill
                c.font = sum_font
        autofit(ws, len(headers))

    # ===== Насосы и расширительные баки =====
    if getattr(project, "heating_hydraulics_results", None):
        ws = wb.create_sheet("Насосы и баки")
        headers = ["Контур", "Q, м³/ч", "H, м",
                   "Насос (модель)", "Q насоса", "H насоса", "P, Вт",
                   "Объём системы, л", "ΔV расш., л",
                   "Объём бака расч., л", "Бак модель",
                   "P_max, бар", "P_init, бар", "Подпитка, л/сут"]
        ws.append(headers)
        style_header(ws[1])
        for name, r in project.heating_hydraulics_results.items():
            ws.append([
                name,
                round(r.pump.flow_m3_h, 2),
                round(r.pump.head_m, 2),
                r.pump.selected_model or "—",
                round(r.pump.selected_flow_m3_h, 2),
                round(r.pump.selected_head_m, 2),
                round(r.pump.selected_power_w, 0),
                round(r.expansion_tank.system_volume_l, 1),
                round(r.expansion_tank.expansion_volume_l, 1),
                round(r.expansion_tank.required_tank_volume_l, 1),
                r.expansion_tank.selected_model or "—",
                round(r.expansion_tank.p_max_bar, 2),
                round(r.expansion_tank.p_init_bar, 2),
                round(r.makeup.daily_makeup_l, 1),
            ])
        autofit(ws, len(headers))

    # ===== Радиаторы по помещениям =====
    if getattr(project, "radiator_picks", None):
        ws = wb.create_sheet("Радиаторы")
        headers = ["№ помещ.", "Название", "Уровень", "Q треб., Вт",
                   "Модель", "Семейство",
                   "Высота, мм", "Длина, мм",
                   "Секций", "Q факт., Вт", "Запас, %", "V воды, л"]
        ws.append(headers)
        style_header(ws[1])
        for sp in project.spaces:
            pick = project.radiator_picks.get(sp.space_id)
            if pick is None:
                continue
            v_water = (pick.model.water_volume_l * pick.sections
                       if pick.model.is_sectional
                       else pick.model.water_volume_l)
            ws.append([
                sp.number, sp.name, sp.level,
                round(sp.heat_loss_w, 0),
                pick.model.name, pick.model.family,
                pick.model.height_mm, pick.model.length_mm,
                pick.sections if pick.model.is_sectional else "—",
                round(pick.actual_power_w, 0),
                round(pick.margin_pct, 1),
                round(v_water, 2),
            ])
        # Сводка: Σ воды
        if project.radiator_picks:
            total_water = 0.0
            for sp in project.spaces:
                pick = project.radiator_picks.get(sp.space_id)
                if pick is None:
                    continue
                vv = (pick.model.water_volume_l * pick.sections
                      if pick.model.is_sectional
                      else pick.model.water_volume_l)
                total_water += vv
            ws.append(["ИТОГО", "", "", "", "", "", "", "", "",
                       "", "", round(total_water, 1)])
            last = ws.max_row
            for c in ws[last]:
                c.fill = sum_fill
                c.font = sum_font
        autofit(ws, len(headers))

    # ===== Акустика =====
    if getattr(project, "acoustics_results", None):
        ws = wb.create_sheet("Акустика")
        headers = ["AHU", "Норма Lp, дБА", "Расчёт Lp, дБА", "Запас, дБА",
                   "Глушитель нужен?", "Глушитель", "Длина, мм",
                   "ΔP глушителя, Па"]
        ws.append(headers)
        style_header(ws[1])
        for name, a in project.acoustics_results.items():
            sil = a.silencer_selected
            ws.append([
                name,
                round(a.lpa_required_dba, 1),
                round(a.lpa_at_terminal, 1),
                round(a.margin_dba, 1),
                "Да" if a.silencer_required else "Нет",
                sil.name if sil else "—",
                sil.length_mm if sil else "—",
                round(sil.pressure_drop_pa, 0) if sil else "—",
            ])
        autofit(ws, len(headers))

    # ===== Тёплый пол =====
    if getattr(project, "underfloor_loops", None):
        ws = wb.create_sheet("Тёплый пол")
        headers = ["№", "Помещение", "Уровень", "Площадь, м²", "Q треб., Вт",
                    "Шаг, мм", "Покрытие", "Зона",
                    "T под./об., °C", "T пов., °C", "Лимит T, °C",
                    "Q факт., Вт/м²", "Q факт., Вт",
                    "Длина трубы, м", "G, кг/ч", "Δp, кПа",
                    "Трубка", "Замечания"]
        ws.append(headers)
        style_header(ws[1])
        total_pipe_m = 0.0
        for sp in project.spaces:
            loop = project.underfloor_loops.get(sp.space_id)
            if loop is None:
                continue
            total_pipe_m += loop.pipe_length_m
            ws.append([
                sp.number, sp.name, sp.level,
                round(loop.area_m2, 1),
                round(loop.q_required_w, 0),
                loop.pitch_mm, loop.cover, loop.zone,
                f"{loop.t_supply_c:.0f}/{loop.t_return_c:.0f}",
                round(loop.t_floor_surface_c, 1),
                round(loop.t_floor_limit_c, 1),
                round(loop.q_actual_w_m2, 1),
                round(loop.q_actual_w, 0),
                round(loop.pipe_length_m, 1),
                round(loop.flow_kg_h, 1),
                round(loop.pressure_drop_kpa, 1),
                loop.pipe.name if loop.pipe else "—",
                "; ".join(loop.warnings) or "—",
            ])
        if project.underfloor_loops:
            ws.append(["ИТОГО", "", "", "", "", "", "", "", "", "", "",
                        "", "", round(total_pipe_m, 0), "", "", "", ""])
            last = ws.max_row
            for c in ws[last]:
                c.fill = sum_fill
                c.font = sum_font
        autofit(ws, len(headers))

    # ===== Фанкойлы по помещениям =====
    if getattr(project, "fancoil_picks", None):
        ws = wb.create_sheet("Фанкойлы")
        headers = ["№", "Помещение", "Q_холод треб., Вт", "Q_тепло треб., Вт",
                    "Модель", "Семейство", "Труб",
                    "Q_холод факт., Вт", "Q_тепло факт., Вт",
                    "Запас холод, %", "Запас тепло, %",
                    "Воздух, м³/ч", "Шум, дБА"]
        ws.append(headers)
        style_header(ws[1])
        for sp in project.spaces:
            pick = project.fancoil_picks.get(sp.space_id)
            if pick is None:
                continue
            ws.append([
                sp.number, sp.name,
                round(sp.heat_gain_w, 0),
                round(sp.heat_loss_w, 0),
                pick.model.name, pick.model.family, pick.model.pipes,
                round(pick.actual_cool_w, 0),
                round(pick.actual_heat_w, 0),
                round(pick.cool_margin_pct, 1),
                round(pick.heat_margin_pct, 1),
                round(pick.model.air_flow_m3_h, 0),
                round(pick.model.noise_db_a, 0),
            ])
        autofit(ws, len(headers))

    # ===== VRF/VRV =====
    if getattr(project, "vrf_systems", None):
        ws = wb.create_sheet("VRF")
        # Сводка систем
        headers = ["Система", "Внешний блок", "Внутренних блоков",
                    "Σ индекс", "K соединения",
                    "Q_холод (кат), кВт", "Q_тепло (кат), кВт",
                    "Q_холод корр., кВт",
                    "Магистраль, м", "Макс. до внутреннего, м",
                    "Δh, м", "Корректировка"]
        ws.append(headers)
        style_header(ws[1])
        for name, sys in project.vrf_systems.items():
            out_name = sys.outdoor.name if sys.outdoor else "—"
            q_cool_cat = (sys.outdoor.q_cool_w / 1000.0
                          if sys.outdoor else 0.0)
            q_heat_cat = (sys.outdoor.q_heat_w / 1000.0
                          if sys.outdoor else 0.0)
            ws.append([
                name, out_name,
                len(sys.indoors), sys.total_indoor_capacity_index,
                round(sys.combination_ratio, 2),
                round(q_cool_cat, 1), round(q_heat_cat, 1),
                round(sys.corrected_cool_w / 1000.0, 1),
                round(sys.main_pipe_length_m, 0),
                round(sys.max_pipe_length_to_indoor_m, 0),
                round(sys.max_height_diff_m, 0),
                round(sys.capacity_correction_factor, 3),
            ])
        autofit(ws, len(headers))

        # Внутренние блоки и медные трубы — следующая страница / диапазон
        ws2 = wb.create_sheet("VRF внутренние блоки")
        headers2 = ["Система", "Помещение", "Внутренний блок",
                    "Семейство", "Индекс",
                    "Q_холод, Вт", "Q_тепло, Вт",
                    "Длина трассы, м",
                    "Ø жидкость, мм", "Ø газ, мм"]
        ws2.append(headers2)
        style_header(ws2[1])
        from hvac.vrf import pipe_diameters_by_index
        for name, sys in project.vrf_systems.items():
            for a in sys.indoors:
                liq, gas = pipe_diameters_by_index(a.indoor.capacity_index)
                ws2.append([
                    name,
                    a.space_id or "—",
                    a.indoor.name, a.indoor.family,
                    a.indoor.capacity_index,
                    round(a.indoor.q_cool_w, 0),
                    round(a.indoor.q_heat_w, 0),
                    round(a.pipe_length_m or
                          sys.max_pipe_length_to_indoor_m, 1),
                    liq, gas,
                ])
        autofit(ws2, len(headers2))

    wb.save(path)
