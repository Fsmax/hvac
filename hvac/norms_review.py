# -*- coding: utf-8 -*-
"""Экспорт и применение инженерной сверки каталогов норм через XLSX.

Модуль не подтверждает нормативные значения сам. Он переносит записи каталога
в таблицу, принимает только явные решения инженера и фиксирует подпись в JSON.
Защищённые поля таблицы и SHA-256 исходного каталога проверяются до записи,
поэтому устаревшая или отредактированная вне review-колонок таблица отклоняется.

CLI::

    python -m hvac.norms_review export --out norms-review.xlsx
    python -m hvac.norms_review apply --in norms-review.xlsx

Для тестовой копии или будущего каталога того же формата обе команды принимают
``--catalog path/to/catalog.json``. Имя проверяющего можно передать через
``--verified-by``; без него используется имя текущего пользователя ОС.
"""

from __future__ import annotations

import argparse
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime, timezone
import getpass
from hashlib import sha256
from io import BytesIO
import json
import math
import os
from pathlib import Path
import re
import sys
from tempfile import NamedTemporaryFile
from typing import Any, Mapping, Sequence, cast


DEFAULT_CATALOG_PATH = (
    Path(__file__).resolve().parent
    / "catalogs"
    / "data"
    / "shnq_2_04_05_22_ducts.json"
)

REVIEW_SHEET = "Сверка норм"
META_SHEET = "_meta"
REVIEW_SCHEMA = "hvac.norms-review/1.0"
MAX_CATALOG_BYTES = 16 * 1024 * 1024
MAX_WORKBOOK_BYTES = 32 * 1024 * 1024

HEADERS = (
    "Ключ",
    "Значение/диапазон",
    "Единица",
    "Применимость",
    "Пункт",
    "Страница PDF",
    "noteRu",
    "Текущий статус",
    "Подтверждаю (да/нет/исправить)",
    "Исправленное значение",
)

_PROTECTED_HEADER_COUNT = 8
_KEY_RE = re.compile(r"^[a-z0-9]+(?:[._][a-z0-9]+)*$")
_INTEGER_RE = re.compile(r"^[+-]?(?:0|[1-9][0-9]*)$")
_DECIMAL_COMMA_RE = re.compile(
    r"^[+-]?(?:(?:0|[1-9][0-9]*),[0-9]+|(?:0|[1-9][0-9]*)(?:[eE][+-]?[0-9]+))$"
)
_TOP_LEVEL_KEYS = {"schemaVersion", "document", "edition", "entries"}
_ENTRY_BASE_KEYS = {"key", "unit", "appliesTo", "source", "status", "noteRu"}
_VERIFICATION_KEYS = {"verifiedBy", "verifiedAt"}
_SOURCE_KEYS = {"document", "edition", "clause", "pagePdf", "table"}
_RANGE_KEYS = {"min", "max", "minInclusive", "maxInclusive"}
_ALLOWED_STATUSES = {"unverified", "unreadable", "verified"}


class NormsReviewError(ValueError):
    """Входной каталог или таблица сверки нарушает безопасный контракт."""


@dataclass(frozen=True, slots=True)
class ReviewSummary:
    """Счётчики полностью применённой таблицы сверки."""

    total: int
    verified: int
    corrected: int
    unchanged: int


@dataclass(frozen=True, slots=True)
class _LoadedCatalog:
    path: Path
    raw_bytes: bytes
    digest: str
    data: dict[str, Any]


@dataclass(frozen=True, slots=True)
class _ReviewAction:
    key: str
    corrected: bool
    value: Any = None


def _fail(message: str) -> NormsReviewError:
    return NormsReviewError(message)


def _reject_constant(value: str) -> None:
    raise _fail(f"Недопустимая JSON-константа: {value}")


def _object_without_duplicates(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise _fail(f"Повторяющийся JSON-ключ: {key}")
        result[key] = value
    return result


def _as_object(value: Any, location: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise _fail(f"{location}: ожидался объект")
    return cast(dict[str, Any], value)


def _as_non_empty_string(value: Any, location: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise _fail(f"{location}: ожидалась непустая строка")
    return value


def _as_number(value: Any, location: str) -> int | float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise _fail(f"{location}: ожидалось число")
    if not math.isfinite(float(value)):
        raise _fail(f"{location}: число должно быть конечным")
    return value


def _expect_exact_keys(
    value: Mapping[str, Any], expected: set[str], location: str
) -> None:
    actual = set(value)
    if actual != expected:
        missing = sorted(expected - actual)
        extra = sorted(actual - expected)
        raise _fail(f"{location}: неверные поля; отсутствуют {missing}, лишние {extra}")


def _validate_aware_timestamp(value: Any, location: str) -> str:
    timestamp = _as_non_empty_string(value, location)
    try:
        parsed = datetime.fromisoformat(timestamp)
    except ValueError as exc:
        raise _fail(f"{location}: ожидалась дата ISO 8601") from exc
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise _fail(f"{location}: дата ISO 8601 должна содержать часовой пояс")
    return timestamp


def _validate_range(value: Any, location: str) -> dict[str, Any]:
    data = _as_object(value, location)
    _expect_exact_keys(data, _RANGE_KEYS, location)

    minimum = data["min"]
    maximum = data["max"]
    if minimum is not None:
        _as_number(minimum, f"{location}.min")
    if maximum is not None:
        _as_number(maximum, f"{location}.max")
    if minimum is None and maximum is None:
        raise _fail(f"{location}: должна быть задана хотя бы одна граница")
    if minimum is not None and maximum is not None and minimum > maximum:
        raise _fail(f"{location}: min не может быть больше max")

    min_inclusive = data["minInclusive"]
    max_inclusive = data["maxInclusive"]
    if not isinstance(min_inclusive, bool) or not isinstance(max_inclusive, bool):
        raise _fail(f"{location}: признаки включённости должны быть bool")
    if minimum is None and min_inclusive:
        raise _fail(f"{location}: minInclusive должен быть false без min")
    if maximum is None and max_inclusive:
        raise _fail(f"{location}: maxInclusive должен быть false без max")
    return data


def _validate_applies_to(value: Any, location: str) -> dict[str, Any]:
    data = _as_object(value, location)
    if not data:
        raise _fail(f"{location}: объект не должен быть пустым")
    for dimension, raw_values in data.items():
        _as_non_empty_string(dimension, f"{location} key")
        if not isinstance(raw_values, list) or not raw_values:
            raise _fail(f"{location}.{dimension}: ожидался непустой массив")
        normalized: set[str] = set()
        for index, item in enumerate(raw_values):
            text = _as_non_empty_string(item, f"{location}.{dimension}[{index}]")
            token = text.strip().casefold()
            if token in normalized:
                raise _fail(f"{location}.{dimension}: повторяющиеся значения")
            normalized.add(token)
    return data


def _validate_catalog(data: Any) -> dict[str, Any]:
    catalog = _as_object(data, "catalog")
    _expect_exact_keys(catalog, _TOP_LEVEL_KEYS, "catalog")
    _as_non_empty_string(catalog["schemaVersion"], "schemaVersion")
    document = _as_non_empty_string(catalog["document"], "document")
    edition = _as_non_empty_string(catalog["edition"], "edition")

    raw_entries = catalog["entries"]
    if not isinstance(raw_entries, list) or not raw_entries:
        raise _fail("entries: ожидался непустой массив")

    seen_keys: set[str] = set()
    for index, raw_entry in enumerate(raw_entries):
        location = f"entries[{index}]"
        entry = _as_object(raw_entry, location)
        has_value = "value" in entry
        has_range = "range" in entry
        if has_value == has_range:
            raise _fail(f"{location}: требуется ровно одно из value и range")

        status = entry.get("status")
        if status not in _ALLOWED_STATUSES:
            raise _fail(f"{location}.status: недопустимый статус {status!r}")
        expected_keys = _ENTRY_BASE_KEYS | ({"value"} if has_value else {"range"})
        if status == "verified":
            expected_keys |= _VERIFICATION_KEYS
        _expect_exact_keys(entry, expected_keys, location)

        key = _as_non_empty_string(entry["key"], f"{location}.key")
        if not _KEY_RE.fullmatch(key):
            raise _fail(f"{location}.key: нестабильный ASCII-ключ {key!r}")
        if key in seen_keys:
            raise _fail(f"{location}.key: повторяющийся ключ {key!r}")
        seen_keys.add(key)

        _as_non_empty_string(entry["unit"], f"{location}.unit")
        _as_non_empty_string(entry["noteRu"], f"{location}.noteRu")
        _validate_applies_to(entry["appliesTo"], f"{location}.appliesTo")

        source = _as_object(entry["source"], f"{location}.source")
        _expect_exact_keys(source, _SOURCE_KEYS, f"{location}.source")
        if _as_non_empty_string(source["document"], f"{location}.source.document") != document:
            raise _fail(f"{location}.source.document: не совпадает с каталогом")
        if _as_non_empty_string(source["edition"], f"{location}.source.edition") != edition:
            raise _fail(f"{location}.source.edition: не совпадает с каталогом")
        _as_non_empty_string(source["clause"], f"{location}.source.clause")
        page = source["pagePdf"]
        if isinstance(page, bool) or not isinstance(page, int) or page <= 0:
            raise _fail(f"{location}.source.pagePdf: ожидался положительный int")
        table = source["table"]
        if table is not None and (not isinstance(table, str) or not table.strip()):
            raise _fail(f"{location}.source.table: ожидалась строка или null")

        if has_value:
            raw_value = entry["value"]
            if raw_value is None:
                if status not in {"unreadable", "verified"}:
                    raise _fail(
                        f"{location}: value:null допустим только для unreadable/verified"
                    )
            else:
                _as_number(raw_value, f"{location}.value")
                if status == "unreadable":
                    raise _fail(f"{location}: числовое value несовместимо с unreadable")
        else:
            _validate_range(entry["range"], f"{location}.range")
            if status == "unreadable":
                raise _fail(f"{location}: range несовместим с unreadable")

        if status == "verified":
            _as_non_empty_string(entry["verifiedBy"], f"{location}.verifiedBy")
            _validate_aware_timestamp(entry["verifiedAt"], f"{location}.verifiedAt")

    return catalog


def _load_catalog(path: str | Path) -> _LoadedCatalog:
    catalog_path = Path(path).resolve()
    try:
        raw = catalog_path.read_bytes()
    except OSError as exc:
        raise _fail(f"Не удалось прочитать каталог {catalog_path}: {exc}") from exc
    if len(raw) > MAX_CATALOG_BYTES:
        raise _fail(
            f"Каталог слишком велик: {len(raw)} байт; предел {MAX_CATALOG_BYTES}"
        )
    try:
        text = raw.decode("utf-8")
    except UnicodeDecodeError as exc:
        raise _fail(f"Каталог должен быть UTF-8: {catalog_path}") from exc
    try:
        decoded = json.loads(
            text,
            parse_constant=_reject_constant,
            object_pairs_hook=_object_without_duplicates,
        )
    except NormsReviewError:
        raise
    except (json.JSONDecodeError, TypeError) as exc:
        raise _fail(f"Некорректный JSON {catalog_path}: {exc}") from exc

    data = _validate_catalog(decoded)
    return _LoadedCatalog(
        path=catalog_path,
        raw_bytes=raw,
        digest=sha256(raw).hexdigest(),
        data=data,
    )


def _canonical_json(value: Any) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    )


def _entry_snapshot(entry: Mapping[str, Any]) -> tuple[Any, ...]:
    value = entry["value"] if "value" in entry else entry["range"]
    source = cast(Mapping[str, Any], entry["source"])
    return (
        entry["key"],
        _canonical_json(value),
        entry["unit"],
        _canonical_json(entry["appliesTo"]),
        source["clause"],
        source["pagePdf"],
        entry["noteRu"],
        entry["status"],
    )


def _plain_cell(cell: Any, value: Any) -> None:
    """Записать строку как текст, даже если она начинается со знака '='."""

    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"


def export_catalog(catalog_path: str | Path, output_path: str | Path) -> int:
    """Экспортировать каталог в новую XLSX-таблицу и вернуть число записей."""

    catalog = _load_catalog(catalog_path)
    output = Path(output_path).resolve()
    if output.exists():
        raise _fail(f"Файл уже существует, перезапись запрещена: {output}")
    if output.suffix.casefold() != ".xlsx":
        raise _fail(f"Файл сверки должен иметь расширение .xlsx: {output}")

    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as exc:  # pragma: no cover - project dependency
        raise RuntimeError("Для сверки норм требуется openpyxl") from exc

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = REVIEW_SHEET
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{len(catalog.data['entries']) + 1}"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    protected_fill = PatternFill("solid", fgColor="E7E6E6")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    fix_fill = PatternFill("solid", fgColor="E2F0D9")
    warning_fill = PatternFill("solid", fgColor="FCE4D6")

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(1, column)
        _plain_cell(cell, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.cell(1, 9).comment = Comment(
        "Допустимо только: да, нет, исправить. Пустая ячейка ничего не меняет.",
        "HVAC",
    )
    sheet.cell(1, 10).comment = Comment(
        "Для value укажите число. Для range укажите JSON с min, max, "
        "minInclusive, maxInclusive. Используется только при решении «исправить».",
        "HVAC",
    )

    entries = cast(list[dict[str, Any]], catalog.data["entries"])
    for row_index, entry in enumerate(entries, start=2):
        for column, value in enumerate(_entry_snapshot(entry), start=1):
            cell = sheet.cell(row_index, column)
            _plain_cell(cell, value)
            cell.fill = protected_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row_index, 9).fill = review_fill
        sheet.cell(row_index, 10).fill = fix_fill
        sheet.cell(row_index, 9).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.cell(row_index, 10).alignment = Alignment(vertical="top", wrap_text=True)

    decision_validation = DataValidation(
        type="list", formula1='"да,нет,исправить"', allow_blank=True
    )
    decision_validation.error = "Выберите: да, нет или исправить"
    decision_validation.errorTitle = "Недопустимое решение"
    decision_validation.prompt = "Явно выберите решение инженера"
    decision_validation.promptTitle = "Сверка нормы"
    decision_validation.showErrorMessage = True
    decision_validation.showInputMessage = True
    sheet.add_data_validation(decision_validation)
    decision_validation.add(f"I2:I{len(entries) + 1}")

    sheet.conditional_formatting.add(
        f"A2:J{len(entries) + 1}",
        FormulaRule(formula=['$I2="исправить"'], fill=warning_fill),
    )
    widths = (48, 40, 14, 55, 20, 14, 85, 18, 27, 42)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + column)].width = width
    sheet.row_dimensions[1].height = 42

    meta = workbook.create_sheet(META_SHEET)
    metadata: tuple[tuple[str, Any], ...] = (
        ("reviewSchema", REVIEW_SCHEMA),
        ("catalogSha256", catalog.digest),
        ("catalogFileName", catalog.path.name),
        ("catalogSchemaVersion", catalog.data["schemaVersion"]),
        ("document", catalog.data["document"]),
        ("edition", catalog.data["edition"]),
        ("entryCount", len(entries)),
    )
    for row, (name, value) in enumerate(metadata, start=1):
        _plain_cell(meta.cell(row, 1), name)
        _plain_cell(meta.cell(row, 2), value)
    meta.sheet_state = "veryHidden"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    payload = buffer.getvalue()
    output.parent.mkdir(parents=True, exist_ok=True)
    try:
        with output.open("xb") as stream:
            stream.write(payload)
    except FileExistsError as exc:
        raise _fail(f"Файл уже существует, перезапись запрещена: {output}") from exc
    except OSError as exc:
        raise _fail(f"Не удалось записать XLSX {output}: {exc}") from exc
    return len(entries)


def _metadata(sheet: Any) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for row in sheet.iter_rows(min_col=1, max_col=2, values_only=False):
        name = row[0].value
        value = row[1].value
        if name is None and value is None:
            continue
        if not isinstance(name, str) or not name.strip():
            raise _fail("Служебный лист содержит некорректный ключ")
        if name in result:
            raise _fail(f"Служебный лист содержит повторяющийся ключ {name!r}")
        if row[0].data_type == "f" or row[1].data_type == "f":
            raise _fail("Формулы на служебном листе запрещены")
        result[name] = value
    return result


def _validate_workbook_metadata(meta: Mapping[str, Any], catalog: _LoadedCatalog) -> None:
    expected = {
        "reviewSchema": REVIEW_SCHEMA,
        "catalogSha256": catalog.digest,
        "catalogFileName": catalog.path.name,
        "catalogSchemaVersion": catalog.data["schemaVersion"],
        "document": catalog.data["document"],
        "edition": catalog.data["edition"],
        "entryCount": len(catalog.data["entries"]),
    }
    if set(meta) != set(expected):
        raise _fail("Служебные поля XLSX отсутствуют или изменены")
    for key, expected_value in expected.items():
        if meta[key] != expected_value:
            if key == "catalogSha256":
                raise _fail("Каталог изменён после экспорта; выполните новый export")
            raise _fail(f"Служебное поле XLSX изменено: {key}")


def _decision(value: Any, key: str) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        raise _fail(f"{key}: решение должно быть текстом да/нет/исправить")
    normalized = value.strip().casefold()
    if normalized not in {"", "да", "нет", "исправить"}:
        raise _fail(f"{key}: неизвестное решение {value!r}")
    return normalized


def _parse_json_cell(value: str, key: str) -> Any:
    try:
        return json.loads(
            value,
            parse_constant=_reject_constant,
            object_pairs_hook=_object_without_duplicates,
        )
    except NormsReviewError:
        raise
    except (json.JSONDecodeError, TypeError) as exc:
        raise _fail(f"{key}: некорректное исправленное значение: {exc}") from exc


def _parse_scalar_correction(value: Any, key: str) -> int | float:
    if isinstance(value, bool):
        raise _fail(f"{key}: исправленное value должно быть числом")
    if isinstance(value, (int, float)):
        return _as_number(value, f"{key}: исправленное value")
    if not isinstance(value, str) or not value.strip():
        raise _fail(f"{key}: для «исправить» требуется исправленное значение")

    text = value.strip()
    if _DECIMAL_COMMA_RE.fullmatch(text):
        text = text.replace(",", ".")
    if _INTEGER_RE.fullmatch(text):
        parsed: Any = int(text)
    else:
        parsed = _parse_json_cell(text, key)
    return _as_number(parsed, f"{key}: исправленное value")


def _parse_correction(value: Any, entry: Mapping[str, Any]) -> Any:
    key = cast(str, entry["key"])
    if value is None or (isinstance(value, str) and not value.strip()):
        raise _fail(f"{key}: для «исправить» требуется исправленное значение")
    if "value" in entry:
        return _parse_scalar_correction(value, key)
    if not isinstance(value, str):
        raise _fail(f"{key}: исправленный range должен быть JSON-объектом")
    parsed = _parse_json_cell(value.strip(), key)
    return _validate_range(parsed, f"{key}: исправленный range")


def _formula_is_forbidden(cell: Any, key: str, header: str) -> None:
    if cell.data_type == "f" or (
        isinstance(cell.value, str) and cell.value.startswith("=")
    ):
        raise _fail(f"{key}: формулы в колонке «{header}» запрещены")


def _collect_actions(sheet: Any, catalog: _LoadedCatalog) -> list[_ReviewAction]:
    actual_headers = tuple(sheet.cell(1, column).value for column in range(1, 11))
    if actual_headers != HEADERS:
        raise _fail("Заголовки таблицы отсутствуют, переименованы или переставлены")

    entries = cast(list[dict[str, Any]], catalog.data["entries"])
    expected_by_key = {cast(str, entry["key"]): entry for entry in entries}
    seen: set[str] = set()
    actions: list[_ReviewAction] = []

    for row in range(2, sheet.max_row + 1):
        cells = [sheet.cell(row, column) for column in range(1, 11)]
        if all(cell.value is None for cell in cells):
            continue
        raw_key = cells[0].value
        if not isinstance(raw_key, str) or not raw_key:
            raise _fail(f"Строка {row}: отсутствует ключ записи")
        key = raw_key
        if key in seen:
            raise _fail(f"Строка {row}: повторяющийся ключ {key!r}")
        seen.add(key)
        entry = expected_by_key.get(key)
        if entry is None:
            raise _fail(f"Строка {row}: неизвестный ключ {key!r}")

        expected_snapshot = _entry_snapshot(entry)
        for column in range(_PROTECTED_HEADER_COUNT):
            cell = cells[column]
            _formula_is_forbidden(cell, key, HEADERS[column])
            if cell.value != expected_snapshot[column]:
                raise _fail(
                    f"{key}: защищённая колонка «{HEADERS[column]}» изменена"
                )

        _formula_is_forbidden(cells[8], key, HEADERS[8])
        _formula_is_forbidden(cells[9], key, HEADERS[9])
        decision = _decision(cells[8].value, key)
        if decision == "да":
            actions.append(_ReviewAction(key=key, corrected=False))
        elif decision == "исправить":
            actions.append(
                _ReviewAction(
                    key=key,
                    corrected=True,
                    value=_parse_correction(cells[9].value, entry),
                )
            )
        # Для «нет» и пустой ячейки исправленное значение намеренно игнорируется.

    if seen != set(expected_by_key):
        missing = sorted(set(expected_by_key) - seen)
        raise _fail(f"Из таблицы удалены записи каталога: {missing}")
    return actions


def _review_timestamp(now: datetime | None) -> str:
    timestamp = datetime.now(timezone.utc) if now is None else now
    if timestamp.tzinfo is None or timestamp.utcoffset() is None:
        raise _fail("verifiedAt требует timezone-aware datetime")
    utc = timestamp.astimezone(timezone.utc).replace(microsecond=0)
    return utc.isoformat().replace("+00:00", "Z")


def _atomic_write_catalog(
    catalog: _LoadedCatalog,
    updated: dict[str, Any],
) -> None:
    serialized = (
        json.dumps(updated, ensure_ascii=False, indent=2, allow_nan=False) + "\n"
    ).encode("utf-8")

    temporary_name: str | None = None
    try:
        with NamedTemporaryFile(
            mode="wb",
            prefix=f".{catalog.path.name}.",
            suffix=".tmp",
            dir=catalog.path.parent,
            delete=False,
        ) as stream:
            temporary_name = stream.name
            stream.write(serialized)
            stream.flush()
            os.fsync(stream.fileno())

        # Проверяем сериализованный кандидат и повторно проверяем stale-state
        # непосредственно перед атомарной заменой исходного файла.
        candidate = _load_catalog(temporary_name)
        if candidate.data != updated:
            raise _fail("Readback временного JSON не совпал с подготовленным каталогом")
        try:
            current_digest = sha256(catalog.path.read_bytes()).hexdigest()
        except OSError as exc:
            raise _fail(f"Не удалось повторно прочитать каталог: {exc}") from exc
        if current_digest != catalog.digest:
            raise _fail("Каталог изменён во время apply; запись отменена")

        os.replace(temporary_name, catalog.path)
        temporary_name = None
        readback = _load_catalog(catalog.path)
        if readback.raw_bytes != serialized or readback.data != updated:
            raise _fail("Readback обновлённого JSON не совпал с применённой записью")
    finally:
        if temporary_name is not None:
            try:
                Path(temporary_name).unlink(missing_ok=True)
            except OSError:
                pass


def apply_review(
    input_path: str | Path,
    catalog_path: str | Path = DEFAULT_CATALOG_PATH,
    verified_by: str | None = None,
    *,
    now: datetime | None = None,
) -> ReviewSummary:
    """Применить явные решения из XLSX к каталогу одной атомарной записью."""

    review_path = Path(input_path).resolve()
    try:
        size = review_path.stat().st_size
    except OSError as exc:
        raise _fail(f"Не удалось прочитать XLSX {review_path}: {exc}") from exc
    if size > MAX_WORKBOOK_BYTES:
        raise _fail(f"XLSX слишком велик: {size} байт; предел {MAX_WORKBOOK_BYTES}")

    signer = getpass.getuser() if verified_by is None else verified_by
    if not isinstance(signer, str) or not signer.strip():
        raise _fail("Имя проверяющего не должно быть пустым")
    signer = signer.strip()
    timestamp = _review_timestamp(now)
    catalog = _load_catalog(catalog_path)

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - project dependency
        raise RuntimeError("Для сверки норм требуется openpyxl") from exc
    try:
        workbook = load_workbook(review_path, data_only=False, read_only=False)
    except Exception as exc:
        raise _fail(f"Не удалось открыть XLSX {review_path}: {exc}") from exc

    try:
        if REVIEW_SHEET not in workbook.sheetnames or META_SHEET not in workbook.sheetnames:
            raise _fail("XLSX не является экспортом hvac.norms_review")
        _validate_workbook_metadata(_metadata(workbook[META_SHEET]), catalog)
        actions = _collect_actions(workbook[REVIEW_SHEET], catalog)
    finally:
        workbook.close()

    corrected_count = sum(action.corrected for action in actions)
    total = len(catalog.data["entries"])
    if not actions:
        return ReviewSummary(total=total, verified=0, corrected=0, unchanged=total)

    updated = deepcopy(catalog.data)
    updated_entries = {
        cast(str, entry["key"]): entry
        for entry in cast(list[dict[str, Any]], updated["entries"])
    }
    source_before = {
        key: deepcopy(entry["source"]) for key, entry in updated_entries.items()
    }
    for action in actions:
        entry = updated_entries[action.key]
        if action.corrected:
            if "value" in entry:
                entry["value"] = action.value
            else:
                entry["range"] = action.value
        entry["status"] = "verified"
        entry["verifiedBy"] = signer
        entry["verifiedAt"] = timestamp

    for key, entry in updated_entries.items():
        if entry["source"] != source_before[key]:
            raise _fail(f"{key}: source изменён во время apply")
    _validate_catalog(updated)
    _atomic_write_catalog(catalog, updated)

    return ReviewSummary(
        total=total,
        verified=len(actions),
        corrected=corrected_count,
        unchanged=total - len(actions),
    )


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="python -m hvac.norms_review",
        description="Экспорт и применение инженерной сверки каталогов норм",
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export", help="создать XLSX для сверки")
    export_parser.add_argument("--out", required=True, type=Path, help="новый XLSX")
    export_parser.add_argument(
        "--catalog",
        type=Path,
        default=DEFAULT_CATALOG_PATH,
        help="JSON-каталог того же шаблона",
    )

    apply_parser = subparsers.add_parser("apply", help="применить подписанную XLSX")
    apply_parser.add_argument("--in", dest="input_path", required=True, type=Path)
    apply_parser.add_argument(
        "--catalog",
        type=Path,
        default=DEFAULT_CATALOG_PATH,
        help="JSON-каталог того же шаблона",
    )
    apply_parser.add_argument(
        "--verified-by",
        "--by",
        dest="verified_by",
        help="ФИО/идентификатор инженера; по умолчанию пользователь ОС",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        if args.command == "export":
            count = export_catalog(args.catalog, args.out)
            print(f"Экспортировано записей: {count}")
            print(Path(args.out).resolve())
            return 0

        summary = apply_review(
            args.input_path,
            args.catalog,
            args.verified_by,
        )
        print(f"Каталог: {Path(args.catalog).resolve()}")
        print(f"Всего записей: {summary.total}")
        print(f"Подтверждено: {summary.verified}")
        print(f"Исправлено: {summary.corrected}")
        print(f"Без изменений: {summary.unchanged}")
        return 0
    except (NormsReviewError, OSError, RuntimeError) as exc:
        print(f"norms-review: error: {exc}", file=sys.stderr)
        return 2


__all__ = [
    "DEFAULT_CATALOG_PATH",
    "HEADERS",
    "NormsReviewError",
    "REVIEW_SHEET",
    "ReviewSummary",
    "apply_review",
    "export_catalog",
    "main",
]


if __name__ == "__main__":  # pragma: no cover - exercised through CLI tests
    raise SystemExit(main())
