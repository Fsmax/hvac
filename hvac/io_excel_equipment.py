# -*- coding: utf-8 -*-
"""Экспорт СВОДНОЙ ТАБЛИЦЫ ОБОРУДОВАНИЯ в Excel.

Содержит несколько листов:
1. «Сводная по помещениям» — главная таблица: каждая строка = помещение,
   колонки описывают всё, что в нём установлено (отопит. прибор, охлад. прибор,
   приток, вытяжка) + расчётные нагрузки + покрытие.
2. «Радиаторы и фанкойлы» — спецификация: уникальные модели приборов
   отопления + общее количество по всему зданию.
3. «Охлаждение» — спецификация моделей охлаждения.
4. «Воздухораспределители» — спецификация решёток / диффузоров.
5. «По системам» — группировка по system_heating / system_cooling /
   system_ventilation.
"""

from __future__ import annotations
from collections import defaultdict
from typing import Dict, List
from hvac.project import HVACProject


def export_equipment_summary(project: HVACProject, path: str) -> None:
    """Создаёт Excel-файл со сводной таблицей оборудования."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("Не установлен openpyxl. Выполните: pip install openpyxl")

    wb = Workbook()
    thin = Side(border_style="thin", color="888888")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    sum_fill = PatternFill("solid", fgColor="DCE6F1")
    sum_font = Font(bold=True)

    def style_header(row):
        for c in row:
            c.fill = head_fill
            c.font = head_font
            c.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)
            c.border = border

    def autofit(ws, max_cols):
        for i in range(1, max_cols + 1):
            letter = get_column_letter(i)
            max_len = 8
            for cell in ws[letter]:
                v = cell.value
                if v is None:
                    continue
                max_len = max(max_len, min(len(str(v)) + 2, 45))
            ws.column_dimensions[letter].width = max_len

    # ===========================================================
    # Лист 1: Сводная таблица по помещениям
    # ===========================================================
    ws = wb.active
    ws.title = "Сводная по помещениям"

    headers = [
        "№ помещ.", "Имя", "Уровень", "Тип", "S, м²", "V, м³",
        "Q_отопл, Вт", "Отопит. прибор", "Модель", "Q одн., Вт", "Кол.",
        "Q общ., Вт", "Покрыт. %",
        "Q_охл, Вт", "Охлад. прибор", "Модель", "Q одн., Вт", "Кол.",
        "Q общ., Вт", "Покрыт. %",
        "Приток, м³/ч", "Тип притока", "Модель", "L одн.", "Кол.", "L общ.",
        "Вытяж., м³/ч", "Тип вытяжки", "Модель", "L одн.", "Кол.", "L общ.",
        "Сист. отопл.", "Сист. охл.", "Сист. вент.", "Примечание",
    ]
    ws.append(headers)
    style_header(ws[1])

    # Сортировка: по уровню, потом по номеру
    sorted_spaces = sorted(project.spaces,
                            key=lambda s: (s.level, s.number))

    for sp in sorted_spaces:
        eq = sp.room_equipment
        if eq is None:
            # Пустая строка — оборудование не назначено
            ws.append([
                sp.number, sp.name, sp.level, sp.room_type,
                round(sp.area_m2, 1), round(sp.volume_m3, 1),
                round(sp.heat_loss_w, 0),
                "—", "", "", "", "", "",
                round(sp.heat_gain_w, 0),
                "—", "", "", "", "", "",
                round(sp.supply_m3h, 0),
                "—", "", "", "", "",
                round(sp.exhaust_m3h, 0),
                "—", "", "", "", "",
                sp.system_heating, sp.system_cooling, sp.system_ventilation, "",
            ])
        else:
            ws.append([
                sp.number, sp.name, sp.level, sp.room_type,
                round(sp.area_m2, 1), round(sp.volume_m3, 1),
                # Отопление
                round(sp.heat_loss_w, 0),
                eq.heating_terminal_type,
                eq.heating_terminal_model,
                round(eq.heating_terminal_power_w, 0) if eq.heating_terminal_power_w else "",
                eq.heating_terminal_qty if eq.heating_terminal_qty else "",
                round(eq.heating_total_w, 0) if eq.heating_total_w else "",
                round(eq.coverage_heating(sp.heat_loss_w), 0)
                    if eq.heating_total_w > 0 and sp.heat_loss_w > 0 else "",
                # Охлаждение
                round(sp.heat_gain_w, 0),
                eq.cooling_terminal_type,
                eq.cooling_terminal_model,
                round(eq.cooling_terminal_power_w, 0) if eq.cooling_terminal_power_w else "",
                eq.cooling_terminal_qty if eq.cooling_terminal_qty else "",
                round(eq.cooling_total_w, 0) if eq.cooling_total_w else "",
                round(eq.coverage_cooling(sp.heat_gain_w), 0)
                    if eq.cooling_total_w > 0 and sp.heat_gain_w > 0 else "",
                # Приток
                round(sp.supply_m3h, 0),
                eq.supply_terminal_type,
                eq.supply_terminal_model,
                round(eq.supply_terminal_flow_m3h, 0) if eq.supply_terminal_flow_m3h else "",
                eq.supply_terminal_qty if eq.supply_terminal_qty else "",
                round(eq.supply_total_m3h, 0) if eq.supply_total_m3h else "",
                # Вытяжка
                round(sp.exhaust_m3h, 0),
                eq.exhaust_terminal_type,
                eq.exhaust_terminal_model,
                round(eq.exhaust_terminal_flow_m3h, 0) if eq.exhaust_terminal_flow_m3h else "",
                eq.exhaust_terminal_qty if eq.exhaust_terminal_qty else "",
                round(eq.exhaust_total_m3h, 0) if eq.exhaust_total_m3h else "",
                # Системы
                sp.system_heating, sp.system_cooling, sp.system_ventilation,
                eq.notes,
            ])

    # Итоговая строка
    total_loss = sum(s.heat_loss_w for s in project.spaces)
    total_gain = sum(s.heat_gain_w for s in project.spaces)
    total_supply = sum(s.supply_m3h for s in project.spaces)
    total_exhaust = sum(s.exhaust_m3h for s in project.spaces)

    total_heat_inst = sum(
        s.room_equipment.heating_total_w for s in project.spaces
        if s.room_equipment is not None)
    total_cool_inst = sum(
        s.room_equipment.cooling_total_w for s in project.spaces
        if s.room_equipment is not None)
    total_sup_inst = sum(
        s.room_equipment.supply_total_m3h for s in project.spaces
        if s.room_equipment is not None)
    total_exh_inst = sum(
        s.room_equipment.exhaust_total_m3h for s in project.spaces
        if s.room_equipment is not None)

    summary_row = ["ИТОГО", "", "", "",
                    round(sum(s.area_m2 for s in project.spaces), 1),
                    round(sum(s.volume_m3 for s in project.spaces), 1),
                    round(total_loss, 0), "", "", "", "",
                    round(total_heat_inst, 0),
                    f"{(100*total_heat_inst/total_loss):.0f}" if total_loss else "",
                    round(total_gain, 0), "", "", "", "",
                    round(total_cool_inst, 0),
                    f"{(100*total_cool_inst/total_gain):.0f}" if total_gain else "",
                    round(total_supply, 0), "", "", "", "",
                    round(total_sup_inst, 0),
                    round(total_exhaust, 0), "", "", "", "",
                    round(total_exh_inst, 0),
                    "", "", "", ""]
    ws.append(summary_row)
    for c in ws[ws.max_row]:
        c.fill = sum_fill
        c.font = sum_font
        c.border = border

    autofit(ws, len(headers))
    ws.freeze_panes = "F2"   # фиксируем шапку и первые 5 колонок

    # ===========================================================
    # Лист 2: Спецификация — Радиаторы / Фанкойлы / Тёплые полы
    # ===========================================================
    ws = wb.create_sheet("Радиаторы и фанкойлы")
    ws.append(["Тип прибора", "Модель", "Q одного, Вт",
                "Кол-во шт.", "Σ мощность, Вт", "Помещения"])
    style_header(ws[1])

    # Группируем: (тип, модель, мощность) → список (помещение, qty)
    heat_groups: Dict = defaultdict(list)
    for sp in project.spaces:
        if sp.room_equipment is None:
            continue
        eq = sp.room_equipment
        if eq.heating_terminal_qty <= 0 or eq.heating_terminal_type == "—":
            continue
        key = (eq.heating_terminal_type, eq.heating_terminal_model,
               eq.heating_terminal_power_w)
        heat_groups[key].append((sp, eq.heating_terminal_qty))

    for (t, m, p), instances in sorted(heat_groups.items()):
        total_qty = sum(q for _, q in instances)
        rooms = "; ".join(f"{sp.number}({q})" for sp, q in instances)
        ws.append([t, m, round(p, 0), total_qty,
                    round(p * total_qty, 0), rooms])

    if heat_groups:
        ws.append([])
        ws.append(["ИТОГО", "", "", sum(sum(q for _, q in v)
                                          for v in heat_groups.values()),
                   round(sum(p * sum(q for _, q in v)
                              for (_, _, p), v in heat_groups.items()), 0), ""])
        for c in ws[ws.max_row]:
            c.fill = sum_fill
            c.font = sum_font
    autofit(ws, 6)

    # ===========================================================
    # Лист 3: Спецификация — Охлаждение
    # ===========================================================
    ws = wb.create_sheet("Охлаждение")
    ws.append(["Тип прибора", "Модель", "Q одного, Вт",
                "Кол-во шт.", "Σ мощность, Вт", "Помещения"])
    style_header(ws[1])

    cool_groups: Dict = defaultdict(list)
    for sp in project.spaces:
        if sp.room_equipment is None:
            continue
        eq = sp.room_equipment
        if eq.cooling_terminal_qty <= 0 or eq.cooling_terminal_type == "—":
            continue
        key = (eq.cooling_terminal_type, eq.cooling_terminal_model,
               eq.cooling_terminal_power_w)
        cool_groups[key].append((sp, eq.cooling_terminal_qty))

    for (t, m, p), instances in sorted(cool_groups.items()):
        total_qty = sum(q for _, q in instances)
        rooms = "; ".join(f"{sp.number}({q})" for sp, q in instances)
        ws.append([t, m, round(p, 0), total_qty,
                    round(p * total_qty, 0), rooms])

    if cool_groups:
        ws.append([])
        ws.append(["ИТОГО", "", "", sum(sum(q for _, q in v)
                                          for v in cool_groups.values()),
                   round(sum(p * sum(q for _, q in v)
                              for (_, _, p), v in cool_groups.items()), 0), ""])
        for c in ws[ws.max_row]:
            c.fill = sum_fill
            c.font = sum_font
    autofit(ws, 6)

    # ===========================================================
    # Лист 4: Спецификация — Воздухораспределители
    # ===========================================================
    ws = wb.create_sheet("Воздухораспределители")
    ws.append(["Назначение", "Тип", "Модель", "L одного, м³/ч",
                "Кол-во шт.", "Σ расход, м³/ч", "Помещения"])
    style_header(ws[1])

    air_groups: Dict = defaultdict(list)
    for sp in project.spaces:
        if sp.room_equipment is None:
            continue
        eq = sp.room_equipment
        if eq.supply_terminal_qty > 0 and eq.supply_terminal_type != "—":
            akey = ("Приток", eq.supply_terminal_type,
                    eq.supply_terminal_model, eq.supply_terminal_flow_m3h)
            air_groups[akey].append((sp, eq.supply_terminal_qty))
        if eq.exhaust_terminal_qty > 0 and eq.exhaust_terminal_type != "—":
            akey = ("Вытяжка", eq.exhaust_terminal_type,
                    eq.exhaust_terminal_model, eq.exhaust_terminal_flow_m3h)
            air_groups[akey].append((sp, eq.exhaust_terminal_qty))

    for (purpose, t, m, l), instances in sorted(air_groups.items()):
        total_qty = sum(q for _, q in instances)
        rooms = "; ".join(f"{sp.number}({q})" for sp, q in instances)
        ws.append([purpose, t, m, round(l, 0), total_qty,
                    round(l * total_qty, 0), rooms])

    if air_groups:
        ws.append([])
        ws.append(["ИТОГО", "", "", "",
                    sum(sum(q for _, q in v) for v in air_groups.values()),
                    round(sum(l * sum(q for _, q in v)
                               for (_, _, _, l), v in air_groups.items()), 0),
                    ""])
        for c in ws[ws.max_row]:
            c.fill = sum_fill
            c.font = sum_font
    autofit(ws, 7)

    # ===========================================================
    # Лист 5: Группировка по системам (system_heating / cooling / vent)
    # ===========================================================
    ws = wb.create_sheet("По системам")
    ws.append(["Система", "Тип", "Помещений", "Σ Q расч., Вт",
                "Σ Q установл., Вт", "Покрытие %"])
    style_header(ws[1])

    # Системы отопления
    by_sys_h: Dict[str, List] = defaultdict(list)
    for sp in project.spaces:
        if sp.system_heating:
            by_sys_h[sp.system_heating].append(sp)
    for sys_name, sps in sorted(by_sys_h.items()):
        q_design = sum(s.heat_loss_w for s in sps)
        q_inst = sum(s.room_equipment.heating_total_w for s in sps
                     if s.room_equipment is not None)
        cov = f"{(100*q_inst/q_design):.0f}" if q_design > 0 else ""
        ws.append([sys_name, "Отопление", len(sps),
                    round(q_design, 0), round(q_inst, 0), cov])

    # Системы охлаждения
    by_sys_c: Dict[str, List] = defaultdict(list)
    for sp in project.spaces:
        if sp.system_cooling:
            by_sys_c[sp.system_cooling].append(sp)
    for sys_name, sps in sorted(by_sys_c.items()):
        q_design = sum(s.heat_gain_w for s in sps)
        q_inst = sum(s.room_equipment.cooling_total_w for s in sps
                     if s.room_equipment is not None)
        cov = f"{(100*q_inst/q_design):.0f}" if q_design > 0 else ""
        ws.append([sys_name, "Охлаждение", len(sps),
                    round(q_design, 0), round(q_inst, 0), cov])

    # Системы вентиляции
    by_sys_v: Dict[str, List] = defaultdict(list)
    for sp in project.spaces:
        if sp.system_ventilation:
            by_sys_v[sp.system_ventilation].append(sp)
    for sys_name, sps in sorted(by_sys_v.items()):
        L_design = sum(s.supply_m3h for s in sps)
        L_inst = sum(s.room_equipment.supply_total_m3h for s in sps
                     if s.room_equipment is not None)
        cov = f"{(100*L_inst/L_design):.0f}" if L_design > 0 else ""
        ws.append([sys_name, "Вентиляция (приток)", len(sps),
                    round(L_design, 0), round(L_inst, 0), cov])

    autofit(ws, 6)

    # ===========================================================
    # Сохранение
    # ===========================================================
    wb.save(path)
