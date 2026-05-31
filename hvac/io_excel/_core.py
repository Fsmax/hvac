# -*- coding: utf-8 -*-
"""Базовые листы: параметры, конструкции, теплопотери/поступления,
ограждения, вентиляция, сводки, зоны/системы, дымоудаление, проверки."""
from __future__ import annotations

from typing import Dict, List

from openpyxl.styles import Font, PatternFill

from hvac.models import Space
from hvac.io_excel._common import style_header, autofit, sum_fill, sum_font


def write_core_sheets(wb, project) -> None:
    total_loss = sum(s.heat_loss_w for s in project.spaces)
    total_gain = sum(s.heat_gain_w for s in project.spaces)
    total_area = sum(s.area_m2 for s in project.spaces)

    # ===== Параметры =====
    ws = wb.active
    ws.title = "Параметры"
    p = project.params
    ws.append(["Параметр", "Значение"])
    style_header(ws[1])
    for k, v in [
        ("Проект", p.project_name), ("Город", p.city),
        ("Расч. наружная зимой, °C", p.t_out_heating),
        ("Расч. наружная летом, °C", p.t_out_cooling),
        ("Суточная амплитуда летом, K", p.daily_amplitude),
        ("Пиковая солнечная радиация, Вт/м²", p.solar_intensity_w_m2),
        ("ГСОП (база 18°C)", p.gsop_18),
        ("Методика", p.methodology),
        ("Коэффициент инфильтрации k", p.inf_correction_k),
        ("Запас на отопление", p.safety_margin_heating),
        ("Запас на охлаждение", p.safety_margin_cooling),
        ("Всего помещений", len(project.spaces)),
        ("Уникальных типов конструкций", len(project.constructions)),
    ]:
        ws.append([k, v])
    autofit(ws, 2)

    # ===== Конструкции =====
    ws = wb.create_sheet("Конструкции")
    headers = ["Ключ", "Категория", "Семейство", "Тип", "Толщина, мм",
               "U, Вт/(м²·К)", "SHGC", "Примечание"]
    ws.append(headers)
    style_header(ws[1])
    for con in sorted(project.constructions.values(),
                      key=lambda x: (x.category, x.key)):
        ws.append([con.key, con.category, con.family, con.type_name,
                   con.thickness_mm, con.u_value, con.shgc, con.note])
    autofit(ws, len(headers))

    # ===== Теплопотери =====
    ws = wb.create_sheet("Теплопотери")
    cols = ["№", "Имя", "Уровень", "Тип помещения", "Площадь, м²",
            "Объём, м³", "tв, °C", "Угловое",
            "Через ограждения, Вт", "Инфильтрация, Вт", "Бытовые, Вт",
            "ИТОГО, Вт", "Уд., Вт/м²"]
    ws.append(cols)
    style_header(ws[1])
    for sp in project.spaces:
        br = sp.heat_loss_breakdown
        total = br.get("ИТОГО", 0.0)
        ud = (total / sp.area_m2) if sp.area_m2 > 0 else 0
        ws.append([
            sp.number, sp.name, sp.level, sp.room_type,
            round(sp.area_m2, 2), round(sp.volume_m3, 2),
            sp.t_in_heat, "да" if sp.is_corner else "—",
            round(br.get("Через ограждения", 0), 1),
            round(br.get("Инфильтрация", 0), 1),
            round(br.get("Бытовые (−)", 0), 1),
            round(total, 1), round(ud, 1),
        ])
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="ИТОГО").font = sum_font
    ws.cell(row=row, column=5, value=round(total_area, 2)).font = sum_font
    ws.cell(row=row, column=12, value=round(total_loss, 1)).font = sum_font
    ws.cell(row=row, column=13,
            value=round(total_loss / total_area if total_area else 0, 1)).font = sum_font
    for c in ws[row]:
        c.fill = sum_fill
    autofit(ws, len(cols))

    # ===== Теплопоступления =====
    ws = wb.create_sheet("Теплопоступления")
    cols = ["№", "Имя", "Уровень", "Тип помещения", "S, м²",
            "Объём, м³", "tв, °C", "Чел.",
            "Через огражд., Вт", "Солнце, Вт", "Люди, Вт", "Освещ., Вт",
            "Оборуд., Вт", "Вентиляция, Вт",
            "Sensible, Вт", "Latent, Вт",
            "ИТОГО, Вт", "Уд., Вт/м²"]
    ws.append(cols)
    style_header(ws[1])
    for sp in project.spaces:
        br = sp.heat_gain_breakdown
        total = br.get("ИТОГО", 0.0)
        ud = (total / sp.area_m2) if sp.area_m2 > 0 else 0
        ws.append([
            sp.number, sp.name, sp.level, sp.room_type,
            round(sp.area_m2, 2), round(sp.volume_m3, 2),
            sp.t_in_cool, sp.occupancy_people,
            round(br.get("Через ограждения", 0), 1),
            round(br.get("Солнечная радиация", 0), 1),
            round(br.get("Люди", 0), 1),
            round(br.get("Освещение", 0), 1),
            round(br.get("Оборудование", 0), 1),
            round(br.get("Инфильтрация/вентиляция", 0), 1),
            round(sp.heat_gain_sensible_w, 1),
            round(sp.heat_gain_latent_w, 1),
            round(total, 1), round(ud, 1),
        ])
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="ИТОГО").font = sum_font
    ws.cell(row=row, column=5, value=round(total_area, 2)).font = sum_font
    ws.cell(row=row, column=15,
            value=round(sum(s.heat_gain_sensible_w for s in project.spaces), 1)).font = sum_font
    ws.cell(row=row, column=16,
            value=round(sum(s.heat_gain_latent_w for s in project.spaces), 1)).font = sum_font
    ws.cell(row=row, column=17, value=round(total_gain, 1)).font = sum_font
    ws.cell(row=row, column=18,
            value=round(total_gain / total_area if total_area else 0, 1)).font = sum_font
    for c in ws[row]:
        c.fill = sum_fill
    autofit(ws, len(cols))

    # ===== Ограждения =====
    from hvac.parsers import effective_orientation
    tn_offset = getattr(project.params, "true_north_offset_deg", 0.0) or 0.0

    ws = wb.create_sheet("Ограждения (наружные)")
    cols = ["№ помещ.", "Помещение", "Тип элемента", "Конструкция",
            "Площадь чистая, м²", "U, Вт/(м²·К)",
            "Ориентация (raw)", "Эфф. ориентация",
            "Q зимой, Вт", "Q летом, Вт"]
    ws.append(cols)
    style_header(ws[1])
    # Подсказка о повороте True North в заголовке листа
    if tn_offset != 0:
        ws.cell(row=2, column=1).value = (
            f"Применён поворот True North = {tn_offset:+.0f}°. "
            f"«Raw» — как в CSV из Revit. «Эфф.» — после поправки "
            f"(именно эта используется в расчёте солнца).")
        ws.cell(row=2, column=1).font = Font(italic=True, color="A00000")
        ws.merge_cells(start_row=2, start_column=1, end_row=2,
                       end_column=len(cols))

    for el in project.elements:
        if not el.is_exterior or el.net_area_m2 <= 0:
            continue
        sp = project.get_space(el.space_id)
        if not sp:
            continue
        dt_h = sp.t_in_heat - project.params.t_out_heating
        dt_c = (project.params.t_out_cooling - sp.t_in_cool
                + project.params.daily_amplitude * 0.3)
        q_h = el.u_value * el.net_area_m2 * max(dt_h, 0)
        q_c = el.u_value * el.net_area_m2 * dt_c
        # Raw ориентация
        orient_raw = el.orientation or "—"
        if el.orientation_deg is not None:
            orient_raw = f"{el.orientation} ({el.orientation_deg:.0f}°)"
        # Эффективная ориентация с учётом поворота True North
        eff = effective_orientation(el.orientation, el.orientation_deg,
                                     tn_offset)
        if tn_offset != 0 and el.orientation_deg is not None:
            eff_deg = (el.orientation_deg + tn_offset) % 360
            orient_eff = f"{eff} ({eff_deg:.0f}°)"
        else:
            orient_eff = eff or "—"

        ws.append([
            sp.number, sp.name,
            "Стена" if el.row_type == "external_wall" else "Проём",
            el.construction_key,
            round(el.net_area_m2, 2), round(el.u_value, 3),
            orient_raw, orient_eff,
            round(q_h, 1), round(q_c, 1),
        ])
    autofit(ws, len(cols))

    # ===== Вентиляция =====
    ws = wb.create_sheet("Вентиляция")
    cols = ["№", "Имя", "Уровень", "Тип помещения", "S, м²", "V, м³",
            "Чел.", "Норма свеж., м³/ч·чел",
            "Supply, м³/ч", "Exhaust, м³/ч", "Hood, м³/ч",
            "ACH, 1/ч", "Метод расчёта", "Норматив", "Изм. вручную"]
    ws.append(cols)
    style_header(ws[1])
    total_supply = total_exh = total_hood = 0.0
    for sp in project.spaces:
        br = sp.ventilation_breakdown or {}
        total_supply += sp.supply_m3h
        total_exh += sp.exhaust_m3h
        total_hood += sp.hood_m3h
        ws.append([
            sp.number, sp.name, sp.level, sp.room_type,
            round(sp.area_m2, 2), round(sp.volume_m3, 2),
            sp.occupancy_people,
            br.get("fresh_air_per_person", 0),
            round(sp.supply_m3h, 1), round(sp.exhaust_m3h, 1),
            round(sp.hood_m3h, 1),
            round(sp.ach_calculated, 2),
            br.get("method", "") if not sp.vent_user_modified else "—",
            br.get("note", "") if not sp.vent_user_modified else "ручная правка",
            "да" if sp.vent_user_modified else "—",
        ])
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="ИТОГО").font = sum_font
    ws.cell(row=row, column=9, value=round(total_supply, 1)).font = sum_font
    ws.cell(row=row, column=10, value=round(total_exh, 1)).font = sum_font
    ws.cell(row=row, column=11, value=round(total_hood, 1)).font = sum_font
    for c in ws[row]:
        c.fill = sum_fill
    autofit(ws, len(cols))

    # ===== Сводка по уровням =====
    ws = wb.create_sheet("Сводка по уровням")
    cols = ["Уровень", "Кол-во помещений", "Площадь, м²",
            "Теплопотери, Вт", "Теплопоступл., Вт",
            "Уд. потери, Вт/м²", "Уд. поступл., Вт/м²",
            "Σ Supply, м³/ч", "Σ Exhaust, м³/ч",
            "Стекло Revit, шт", "WWR-оценка, шт", "Без остекл., шт",
            "Σ Стекло, м²", "Σ Наруж. стены, м²"]
    ws.append(cols)
    style_header(ws[1])
    by_level: Dict[str, List[Space]] = {}
    for sp in project.spaces:
        by_level.setdefault(sp.level, []).append(sp)
    # Карта space_id → наружные элементы (для подсчёта площадей по этажу)
    ext_elems_by_space: Dict[str, List] = {}
    for el in project.elements:
        if el.is_exterior:
            ext_elems_by_space.setdefault(el.space_id, []).append(el)
    for lvl in sorted(by_level.keys()):
        items = by_level[lvl]
        a = sum(s.area_m2 for s in items)
        ll = sum(s.heat_loss_w for s in items)
        gg = sum(s.heat_gain_w for s in items)
        sup = sum(s.supply_m3h for s in items)
        exh = sum(s.exhaust_m3h for s in items)
        n_real = sum(1 for s in items
                     if getattr(s, "glazing_source", "none") == "real")
        n_wwr = sum(1 for s in items
                    if getattr(s, "glazing_source", "none") == "wwr")
        n_none = sum(1 for s in items
                     if getattr(s, "glazing_source", "none") == "none")
        # Σ стекла и Σ наружных стен по этажу (по фактической геометрии CSV).
        # Если фасад одинаковый, эти суммы должны совпадать на всех типовых
        # этажах — расхождение указывает на потерю элементов в выгрузке Revit.
        glass_area = 0.0
        wall_area = 0.0
        for s in items:
            for el in ext_elems_by_space.get(s.space_id, []):
                con = project.constructions.get(el.construction_key)
                is_glazed = con is not None and con.shgc > 0
                if is_glazed:
                    glass_area += el.net_area_m2
                elif el.row_type == "external_wall":
                    wall_area += el.net_area_m2
        ws.append([
            lvl, len(items), round(a, 2),
            round(ll, 1), round(gg, 1),
            round(ll / a if a else 0, 1), round(gg / a if a else 0, 1),
            round(sup, 0), round(exh, 0),
            n_real, n_wwr, n_none,
            round(glass_area, 1), round(wall_area, 1),
        ])
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value="ВСЕГО").font = sum_font
    ws.cell(row=row, column=2, value=len(project.spaces)).font = sum_font
    ws.cell(row=row, column=3, value=round(total_area, 2)).font = sum_font
    ws.cell(row=row, column=4, value=round(total_loss, 1)).font = sum_font
    ws.cell(row=row, column=5, value=round(total_gain, 1)).font = sum_font
    for c in ws[row]:
        c.fill = sum_fill
    autofit(ws, len(cols))

    # ===== Конструкции по этажам =====
    # Сводная таблица: для каждой пары (тип конструкции, уровень) —
    # суммарная площадь наружных элементов. Если фасад типовой, но в
    # выгрузке на разных этажах указаны разные типы конструкций
    # (например, на одних этажах витраж, а на других — Basic Wall),
    # «лишний» тип сразу вылезет с площадью на проблемных этажах
    # и нулём на остальных.
    ws = wb.create_sheet("Конструкции по этажам")
    # Соберём pivot: con_key → {level → area}
    by_con_lvl: Dict[str, Dict[str, float]] = {}
    for el in project.elements:
        if not el.is_exterior or el.net_area_m2 <= 0:
            continue
        sp = project.get_space(el.space_id)
        if not sp:
            continue
        key = el.construction_key or "<пусто>"
        by_con_lvl.setdefault(key, {}).setdefault(sp.level, 0.0)
        by_con_lvl[key][sp.level] += el.net_area_m2
    levels_sorted = sorted(by_level.keys())
    header = ["Конструкция", "SHGC", "U, Вт/(м²·К)", "Σ всего, м²"] + levels_sorted
    ws.append(header)
    style_header(ws[1])
    # Сортируем по убыванию суммарной площади
    rows_data = []
    for key, lvl_map in by_con_lvl.items():
        total = sum(lvl_map.values())
        con = project.constructions.get(key)
        shgc = con.shgc if con else ""
        u_val = con.u_value if con else ""
        rows_data.append((key, shgc, u_val, total, lvl_map))
    rows_data.sort(key=lambda r: -r[3])
    for key, shgc, u_val, total, lvl_map in rows_data:
        row = [key, shgc, u_val, round(total, 1)]
        for lvl in levels_sorted:
            row.append(round(lvl_map.get(lvl, 0.0), 1) if lvl in lvl_map else "")
        ws.append(row)
    autofit(ws, len(header))

    # ===== Зоны и системы =====
    has_any_zones = any(sp.system_heating or sp.system_cooling
                        or sp.system_ventilation for sp in project.spaces)
    if has_any_zones:
        from hvac.project import (suggest_ahu_size, suggest_boiler_size,
                                   suggest_chiller_size)
        ws = wb.create_sheet("Зоны и системы")
        ws.append(["СВОДКА ПО ЗОНАМ ДЛЯ ПОДБОРА ОБОРУДОВАНИЯ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.append([])

        # Отопление
        ws.append(["ОТОПЛЕНИЕ (для подбора котлов)"])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        ws.append(["Зона / котёл", "Помещ.", "Площадь, м²",
                   "Q, кВт", "Уд., Вт/м²", "Типовой котёл"])
        style_header(ws[ws.max_row])
        s = project.get_zone_summary("heating")
        total_h = 0
        for zone, d in sorted(s.items(), key=lambda x: -x[1]["q_heating_w"]):
            ud = d["q_heating_w"] / d["area_m2"] if d["area_m2"] else 0
            ws.append([zone, d["n_spaces"], round(d["area_m2"], 1),
                       round(d["q_heating_w"] / 1000, 2), round(ud, 1),
                       suggest_boiler_size(d["q_heating_w"])])
            total_h += d["q_heating_w"]
        ws.append(["ИТОГО", "", "", round(total_h / 1000, 2), "",
                   suggest_boiler_size(total_h)])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        for c in ws[ws.max_row]:
            c.fill = sum_fill
        ws.append([])

        # Охлаждение
        ws.append(["ОХЛАЖДЕНИЕ (для подбора чиллеров)"])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        ws.append(["Зона / чиллер", "Помещ.", "Площадь, м²",
                   "Sensible, кВт", "Latent, кВт", "Total, кВт",
                   "Уд., Вт/м²", "Типовой чиллер"])
        style_header(ws[ws.max_row])
        s = project.get_zone_summary("cooling")
        for zone, d in sorted(s.items(), key=lambda x: -x[1]["q_cooling_w"]):
            ud = d["q_cooling_w"] / d["area_m2"] if d["area_m2"] else 0
            ws.append([zone, d["n_spaces"], round(d["area_m2"], 1),
                       round(d["q_sensible_w"] / 1000, 2),
                       round(d["q_latent_w"] / 1000, 2),
                       round(d["q_cooling_w"] / 1000, 2),
                       round(ud, 1),
                       suggest_chiller_size(d["q_cooling_w"])])
        ws.append([])

        # Вентиляция
        ws.append(["ВЕНТИЛЯЦИЯ (для подбора AHU)"])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        ws.append(["Установка", "Помещ.", "Площадь, м²",
                   "Supply, м³/ч", "Exhaust, м³/ч", "Hood, м³/ч",
                   "Типовой AHU"])
        style_header(ws[ws.max_row])
        s = project.get_zone_summary("ventilation")
        for zone, d in sorted(s.items(), key=lambda x: -x[1]["supply_m3h"]):
            ws.append([zone, d["n_spaces"], round(d["area_m2"], 1),
                       round(d["supply_m3h"], 0),
                       round(d["exhaust_m3h"], 0),
                       round(d["hood_m3h"], 0),
                       suggest_ahu_size(d["supply_m3h"]) + " м³/ч"])
        autofit(ws, 8)

    # ===== Системы оборудования =====
    has_equipment = (project.ventilation_systems or project.heating_systems
                     or project.cooling_systems)
    if has_equipment:
        ws = wb.create_sheet("Системы оборудования")
        ws.append(["ПАРАМЕТРЫ СИСТЕМ ОБОРУДОВАНИЯ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.append([])

        # Приточные установки
        if project.ventilation_systems:
            ws.append(["ПРИТОЧНЫЕ УСТАНОВКИ (AHU)"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Тип", "Рекуп.", "η зимой", "η летом",
                       "tпод зим °C", "tпод лет °C", "wпод лет г/кг",
                       "Примечание"])
            style_header(ws[ws.max_row])
            for name, ahu in sorted(project.ventilation_systems.items()):
                ws.append([
                    ahu.name, ahu.system_type,
                    "да" if ahu.has_recovery else "—",
                    f"{ahu.recovery_efficiency_winter*100:.0f}%" if ahu.has_recovery else "—",
                    f"{ahu.recovery_efficiency_summer*100:.0f}%" if ahu.has_recovery else "—",
                    ahu.t_supply_winter, ahu.t_supply_summer,
                    ahu.w_supply_summer, ahu.note,
                ])

            # Рассчитанные нагрузки на AHU
            ws.append([])
            ws.append(["НАГРУЗКИ ОТ ПРИТОЧКИ (расчёт)"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Установка", "L подача, м³/ч",
                       "Q калорифер, кВт", "Q охлад. явная, кВт",
                       "Q охлад. скрытая, кВт", "Q охлад. total, кВт"])
            style_header(ws[ws.max_row])
            loads = project.calculate_ahu_loads()
            sum_h = sum_cs = sum_cl = 0
            for name, d in sorted(loads.items(),
                                   key=lambda x: -x[1]["q_heater_w"]):
                if d["supply_m3h"] < 1:
                    continue
                ws.append([
                    name, round(d["supply_m3h"], 0),
                    round(d["q_heater_w"] / 1000, 2),
                    round(d["q_cooler_sens_w"] / 1000, 2),
                    round(d["q_cooler_lat_w"] / 1000, 2),
                    round(d["q_cooler_total_w"] / 1000, 2),
                ])
                sum_h += d["q_heater_w"]
                sum_cs += d["q_cooler_sens_w"]
                sum_cl += d["q_cooler_lat_w"]
            ws.append(["ИТОГО", "", round(sum_h/1000, 2),
                       round(sum_cs/1000, 2), round(sum_cl/1000, 2),
                       round((sum_cs+sum_cl)/1000, 2)])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            for c in ws[ws.max_row]:
                c.fill = sum_fill
            ws.append([])

        # Котлы
        if project.heating_systems:
            ws.append(["КОТЛЫ / ИСТОЧНИКИ ТЕПЛА"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Тип", "tпод °C", "tобр °C",
                       "Топливо", "КПД", "Примечание"])
            style_header(ws[ws.max_row])
            for name, h in sorted(project.heating_systems.items()):
                ws.append([
                    h.name, h.system_type, h.t_supply, h.t_return,
                    h.fuel, f"{h.efficiency*100:.0f}%", h.note,
                ])
            ws.append([])

        # Чиллеры
        if project.cooling_systems:
            ws.append(["ЧИЛЛЕРЫ / ИСТОЧНИКИ ХОЛОДА"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Тип", "tпод °C", "tобр °C",
                       "COP", "Хладагент", "Примечание"])
            style_header(ws[ws.max_row])
            for name, c in sorted(project.cooling_systems.items()):
                ws.append([
                    c.name, c.system_type, c.t_supply, c.t_return,
                    c.cop, c.refrigerant, c.note,
                ])
        autofit(ws, 9)

    # ===== Дымоудаление и подпор воздуха =====
    if project.smoke_systems:
        ws = wb.create_sheet("Дымоудаление")
        ws.append(["СИСТЕМЫ ДЫМОУДАЛЕНИЯ И ПОДПОРА ВОЗДУХА (КМК / СП 7.13130)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws.append([])

        # Расчёт нагрузок (single zone — стандартный сценарий)
        loads = project.calculate_smoke_loads(fire_mode="single_zone")

        # СДУ
        smoke_systems = {k: v for k, v in loads.items()
                         if v["system_type"] == "smoke_removal"}
        if smoke_systems:
            ws.append(["СИСТЕМЫ ДЫМОУДАЛЕНИЯ (СДУ)"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Назначение", "Площадь м²", "Зон",
                       "Норма, м³/ч·м²", "L зоны м³/ч",
                       "L системы м³/ч", "L компенс. м³/ч",
                       "t дыма °C", "Огнестойк.", "Примечание"])
            style_header(ws[ws.max_row])
            total_smoke = total_makeup = 0
            for name, d in sorted(smoke_systems.items()):
                ws.append([
                    name, d["purpose"], round(d["served_area_m2"], 0),
                    d["n_zones"], d.get("norm_per_m2", "—"),
                    round(d["L_per_zone_m3h"], 0),
                    round(d["L_smoke_m3h"], 0),
                    round(d["L_makeup_m3h"], 0),
                    d["t_smoke_C"], d["fire_rating"], d["note"],
                ])
                total_smoke += d["L_smoke_m3h"]
                total_makeup += d["L_makeup_m3h"]
            ws.append(["ИТОГО (один пожар)", "", "", "", "", "",
                       round(total_smoke, 0), round(total_makeup, 0),
                       "", "", ""])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            for c in ws[ws.max_row]:
                c.fill = sum_fill
            ws.append([])

        # СПВ
        pres_systems = {k: v for k, v in loads.items()
                        if v["system_type"] == "air_supply"}
        if pres_systems:
            ws.append(["СИСТЕМЫ ПОДПОРА ВОЗДУХА (СПВ)"])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            ws.append(["Имя", "Назначение", "L подпора, м³/ч",
                       "Давление, Па", "Примечание"])
            style_header(ws[ws.max_row])
            total_pres = 0
            for name, d in sorted(pres_systems.items()):
                ws.append([
                    name, d["purpose"],
                    round(d["L_smoke_m3h"], 0),
                    d["pressure_pa"], d["note"],
                ])
                total_pres += d["L_smoke_m3h"]
            ws.append(["ИТОГО СПВ", "", round(total_pres, 0), "", ""])
            ws.cell(row=ws.max_row, column=1).font = sum_font
            for c in ws[ws.max_row]:
                c.fill = sum_fill
        autofit(ws, 11)

    # ===== Проверки данных =====
    validation_results = project.validate_detailed()
    if validation_results:
        ws = wb.create_sheet("Проверки")
        cols = ["№", "Серьёзность", "Категория", "Сообщение", "Помещение"]
        ws.append(cols)
        style_header(ws[1])

        severity_fill = {
            "error":   PatternFill("solid", fgColor="F4B7B6"),
            "warning": PatternFill("solid", fgColor="FFE69A"),
            "info":    PatternFill("solid", fgColor="D9E1F2"),
        }
        severity_text = {"error": "Ошибка", "warning": "Внимание", "info": "Инфо"}

        for i, r in enumerate(validation_results, start=1):
            sp = project.get_space(r.get("space_id", ""))
            sp_label = f"{sp.number} {sp.name}" if sp else "—"
            row_num = ws.max_row + 1
            ws.append([
                i,
                severity_text.get(r["severity"], r["severity"]),
                r["category"],
                r["msg"],
                sp_label,
            ])
            # Раскрасить строку по серьёзности
            fill = severity_fill.get(r["severity"])
            if fill:
                for c in ws[row_num]:
                    c.fill = fill
        autofit(ws, len(cols))
