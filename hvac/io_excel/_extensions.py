# -*- coding: utf-8 -*-
"""Листы расширений v3.7: ГВС, энергопаспорт, точка росы, воздуховоды,
трубы отопления/холодоснабжения, контуры ИТП, циркуляционные насосы."""
from __future__ import annotations

from openpyxl.styles import Font, PatternFill

from hvac.io_excel._common import style_header, autofit, sum_fill, sum_font


def write_extension_sheets(wb, project) -> None:
    # ====================================================================
    # ===== Расширения v3.7 =====
    # ====================================================================

    # ===== Лист "ГВС" =====
    if project.dhw_systems:
        ws = wb.create_sheet("ГВС")
        ws.append(["ГОРЯЧЕЕ ВОДОСНАБЖЕНИЕ (СП 30.13330.2020)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        cols = ["Система", "Тип нагревателя", "Потребителей",
                "V сут, м³/сут", "V час макс, м³/ч",
                "Q пик, кВт", "Q с цирк., кВт", "Q нагревателя, кВт",
                "V бака, м³", "Циркуляция", "η нагр."]
        ws.append(cols)
        style_header(ws[ws.max_row])

        for name, sys in sorted(project.dhw_systems.items()):
            ws.append([
                name,
                sys.heater_type,
                sys.n_consumers,
                round(sys.v_daily_total_m3, 2),
                round(sys.v_hourly_max_m3, 2),
                round(sys.q_peak_w / 1000, 1),
                round(sys.q_with_circulation_w / 1000, 1),
                round(sys.q_heater_size_w / 1000, 1),
                round(sys.storage_recommended_m3, 2),
                "Да" if sys.has_circulation else "Нет",
                round(sys.efficiency, 2),
            ])

        # Итого
        total_v_d = sum(s.v_daily_total_m3 for s in project.dhw_systems.values())
        total_v_h = sum(s.v_hourly_max_m3 for s in project.dhw_systems.values())
        total_q = sum(s.q_heater_size_w for s in project.dhw_systems.values())
        ws.append(["ИТОГО", "", "",
                   round(total_v_d, 2), round(total_v_h, 2),
                   "", "", round(total_q / 1000, 1), "", "", ""])
        ws.cell(row=ws.max_row, column=1).font = sum_font
        for c in ws[ws.max_row]:
            c.fill = sum_fill
        autofit(ws, len(cols))

    # ===== Лист "Энергопаспорт" =====
    if project.energy_passport:
        ep = project.energy_passport
        ws = wb.create_sheet("Энергопаспорт")
        ws.append(["ЭНЕРГЕТИЧЕСКИЙ ПАСПОРТ (СП 50.13330)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        sections = [
            ("Исходные данные", [
                ("Объект", ep.project_name),
                ("Город", ep.city),
                ("Тип здания", ep.building_type),
                ("ГСОП (t_в=+20°C, ≤8°C), °C·сут", ep.gsop_18),
                ("Расчётная зимняя tн, °C", ep.t_out_heating),
                ("Отапливаемая площадь, м²", round(ep.total_area_m2, 1)),
                ("Объём, м³", round(ep.total_volume_m3, 0)),
                ("Помещений", ep.n_spaces),
            ]),
            ("Пиковые нагрузки", [
                ("Q отопления, кВт", round(ep.q_peak_heating_w / 1000, 1)),
                ("Q охлаждения, кВт", round(ep.q_peak_cooling_w / 1000, 1)),
                ("Q нагрев приточек, кВт", round(ep.q_peak_ventilation_heating_w / 1000, 1)),
                ("Q ГВС, кВт", round(ep.q_peak_dhw_w / 1000, 1)),
            ]),
            ("Отопительный сезон", [
                ("Длительность, сут", ep.z_heating_days),
                ("Средняя tн, °C", ep.t_avg_heating),
                ("Коэф. регулирования", ep.k_regulation),
                ("Внутр. теплопост., Вт/м²", ep.internal_gain_w_m2),
            ]),
            ("Годовое потребление, МВт·ч/год", [
                ("Отопление", round(ep.e_heating_kwh_year / 1000, 1)),
                ("Нагрев приточек", round(ep.e_ventilation_kwh_year / 1000, 1)),
                ("Охлаждение (электроэнергия)", round(ep.e_cooling_kwh_year / 1000, 1)),
                ("ГВС", round(ep.e_dhw_kwh_year / 1000, 1)),
                ("Внутр. теплопост. (использованные)",
                 round(ep.e_internal_gains_kwh_year * ep.k_internal_use / 1000, 1)),
            ]),
            ("Удельные показатели", [
                ("qh у удельный, кВт·ч/(м²·год)", round(ep.qh_specific_kwh_m2, 1)),
                ("qh н нормативный, кВт·ч/(м²·год)", round(ep.qh_normative_kwh_m2, 1)),
                ("Отклонение от нормы, %", round(ep.deviation_percent, 1)),
            ]),
        ]
        for title, rows in sections:
            ws.append([title])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
            ws.cell(row=ws.max_row, column=1).fill = sum_fill
            for label, value in rows:
                ws.append([label, value])
            ws.append([])

        # Класс энергоэффективности — крупно
        ws.append(["КЛАСС ЭНЕРГОЭФФЕКТИВНОСТИ",
                   ep.energy_class, ep.energy_class_description])
        for c in ws[ws.max_row]:
            c.font = Font(bold=True, size=12)
            c.fill = sum_fill

        autofit(ws, 3)

    # ===== Лист "Точка росы" =====
    if project.condensation_results:
        ws = wb.create_sheet("Точка росы")
        ws.append(["ПРОВЕРКА ОГРАЖДЕНИЙ НА КОНДЕНСАЦИЮ (СП 50.13330 Прил. Е)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        cols = ["Помещение", "Категория", "Конструкция",
                "U, Вт/(м²·К)", "tв, °C", "tн, °C", "RH вн., %",
                "τ_int, °C", "t_d, °C", "Δt факт., К", "Δt норм., К",
                "Запас до конд., К", "Статус"]
        ws.append(cols)
        style_header(ws[ws.max_row])

        fill_cond = PatternFill("solid", fgColor="F4B7B6")  # красный
        fill_norm = PatternFill("solid", fgColor="FFE69A")  # жёлтый

        for c in project.condensation_results:
            row_num = ws.max_row + 1
            status = ("КОНДЕНСАТ" if c.condensation_risk
                      else "Норматив" if c.normative_fail else "OK")
            ws.append([
                f"{c.space_number} {c.space_name}",
                c.category, c.construction_key,
                round(c.u_value, 2),
                round(c.t_in, 1), round(c.t_out, 1),
                round(c.rh_in, 0),
                round(c.t_surface, 2),
                round(c.t_dew, 2),
                round(c.dt_actual, 2),
                round(c.dt_normative, 1),
                round(c.margin_to_dew, 2),
                status,
            ])
            if c.condensation_risk:
                for cell in ws[row_num]:
                    cell.fill = fill_cond
            elif c.normative_fail:
                for cell in ws[row_num]:
                    cell.fill = fill_norm

        # Сводка
        from hvac.dew_point import total_problems
        prob = total_problems(project.condensation_results)
        ws.append([])
        ws.append(["ИТОГО:", f"всего проверено {prob['total']},",
                   f"конденсат {prob['condensation']},",
                   f"норматив СП 50 не выполняется {prob['normative_fail']},",
                   f"OK {prob['ok']}"])
        ws.cell(row=ws.max_row, column=1).font = sum_font

        autofit(ws, len(cols))

    # ===== Лист "Воздуховоды" =====
    if project.duct_networks:
        ws = wb.create_sheet("Воздуховоды")
        ws.append(["ПОДБОР СЕЧЕНИЙ ВОЗДУХОВОДОВ (СП 60.13330)"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        for sys_name, net in sorted(project.duct_networks.items()):
            ws.append([f"СИСТЕМА: {sys_name}",
                       f"Σ Q = {net.total_flow_m3h:.0f} м³/ч",
                       f"Δp подбора AHU = {net.total_pressure_loss_pa:.0f} Па",
                       f"Точек = {net.n_terminals}"])
            for c in ws[ws.max_row]:
                c.font = sum_font
                c.fill = sum_fill
            ws.append(["Участок", "Тип", "Роль", "L, м³/ч",
                       "Форма", "Размер, мм", "v, м/с", "d_гидр, мм",
                       "Δp_тр, Па", "Δp_мест, Па", "Δp_сум, Па", "Примечание"])
            style_header(ws[ws.max_row])

            for sec in net.sections:
                size_str = (f"Ø{int(sec.diameter_mm)}" if sec.shape == "round"
                            else f"{int(sec.width_mm)}×{int(sec.height_mm)}")
                ws.append([
                    sec.id, sec.section_type, sec.role,
                    round(sec.flow_m3h, 0), sec.shape, size_str,
                    round(sec.velocity_m_s, 2),
                    round(sec.hydraulic_diameter_mm, 0),
                    round(sec.pressure_loss_friction_pa, 1),
                    round(sec.pressure_loss_local_pa, 1),
                    round(sec.pressure_loss_total_pa, 1),
                    sec.note,
                ])
            ws.append([])

        autofit(ws, 12)

    # ===== Лист "Трубы отопления" =====
    if project.pipe_networks:
        ws = wb.create_sheet("Трубы отопления")
        ws.append(["ПОДБОР СЕЧЕНИЙ ТРУБ ОТОПЛЕНИЯ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        for sys_name, net in sorted(project.pipe_networks.items()):
            pump_head_m = (getattr(net, "pump_head_m", 0.0) or
                           net.total_pressure_loss_pa / (977.7 * 9.81) * 1.3)
            pump_flow = (getattr(net, "pump_flow_m3_h", 0.0) or
                         net.total_flow_kg_h / 977.7)
            pump_model = getattr(net, "pump_model", "") or "—"
            t_label = (f"{net.t_supply_c:.0f}/{net.t_return_c:.0f}"
                       if hasattr(net, "t_supply_c") else f"Δt={net.delta_t_k:.0f}")
            ws.append([
                f"КОНТУР: {sys_name}",
                f"Σ Q = {net.total_heat_load_w / 1000:.1f} кВт",
                f"G = {net.total_flow_kg_h:.0f} кг/ч ({pump_flow:.2f} м³/ч)",
                f"t = {t_label} °C",
                f"Δp = {net.total_pressure_loss_pa / 1000:.1f} кПа",
                f"Насос: {pump_model} (H={pump_head_m:.1f} м)",
            ])
            for c in ws[ws.max_row]:
                c.font = sum_font
                c.fill = sum_fill
            ws.append(["Участок", "Тип", "Q, Вт", "G, кг/ч", "V, м³/ч",
                       "Материал", "DN", "d_вн, мм", "v, м/с",
                       "L, м", "Z (Σζ)", "ΔH, м",
                       "Δp_тр, Па", "Δp_мест, Па", "Δp_сум, Па", "Примечание"])
            style_header(ws[ws.max_row])

            for sec in net.sections:
                ws.append([
                    sec.id, sec.section_type,
                    round(sec.heat_load_w, 0),
                    round(sec.flow_kg_h, 1),
                    round(sec.flow_m3_h, 3),
                    sec.pipe_material,
                    int(sec.dn_mm),
                    round(sec.inner_diameter_mm, 1),
                    round(sec.velocity_m_s, 2),
                    round(sec.length_m, 1),
                    round(getattr(sec, "local_zeta_sum", 0.0), 1),
                    round(getattr(sec, "elevation_m", 0.0), 1),
                    round(sec.pressure_loss_friction_pa, 1),
                    round(sec.pressure_loss_local_pa, 1),
                    round(sec.pressure_loss_total_pa, 1),
                    sec.note,
                ])
            ws.append([])

        autofit(ws, 16)

    # ===== Лист "Трубы холодоснабжения" =====
    if getattr(project, "cooling_pipe_networks", None):
        ws = wb.create_sheet("Трубы холодоснабжения")
        ws.append(["ПОДБОР СЕЧЕНИЙ ТРУБ ХОЛОДОСНАБЖЕНИЯ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        for sys_name, net in sorted(project.cooling_pipe_networks.items()):
            pump_head_m = getattr(net, "pump_head_m", 0.0)
            pump_flow = getattr(net, "pump_flow_m3_h", 0.0)
            pump_model = getattr(net, "pump_model", "") or "—"
            t_label = f"{net.t_supply_c:.0f}/{net.t_return_c:.0f}"
            ws.append([
                f"КОНТУР: {sys_name}",
                f"Σ Q = {net.total_heat_load_w / 1000:.1f} кВт",
                f"G = {net.total_flow_kg_h:.0f} кг/ч ({pump_flow:.2f} м³/ч)",
                f"t = {t_label} °C",
                f"Δp = {net.total_pressure_loss_pa / 1000:.1f} кПа",
                f"Насос: {pump_model} (H={pump_head_m:.1f} м)",
            ])
            for c in ws[ws.max_row]:
                c.font = sum_font
                c.fill = sum_fill
            ws.append(["Участок", "Тип", "Q, Вт", "G, кг/ч", "V, м³/ч",
                       "Материал", "DN", "d_вн, мм", "v, м/с",
                       "L, м", "Z (Σζ)", "ΔH, м",
                       "Δp_тр, Па", "Δp_мест, Па", "Δp_сум, Па", "Примечание"])
            style_header(ws[ws.max_row])

            for sec in net.sections:
                ws.append([
                    sec.id, sec.section_type,
                    round(sec.heat_load_w, 0),
                    round(sec.flow_kg_h, 1),
                    round(sec.flow_m3_h, 3),
                    sec.pipe_material,
                    int(sec.dn_mm),
                    round(sec.inner_diameter_mm, 1),
                    round(sec.velocity_m_s, 2),
                    round(sec.length_m, 1),
                    round(getattr(sec, "local_zeta_sum", 0.0), 1),
                    round(getattr(sec, "elevation_m", 0.0), 1),
                    round(sec.pressure_loss_friction_pa, 1),
                    round(sec.pressure_loss_local_pa, 1),
                    round(sec.pressure_loss_total_pa, 1),
                    sec.note,
                ])
            ws.append([])
        autofit(ws, 16)

    # ===== Лист "Контуры ИТП" =====
    has_heating_circuits = bool(getattr(project, "heating_circuits", {}))
    has_cooling_circuits = bool(getattr(project, "cooling_circuits", {}))
    if has_heating_circuits or has_cooling_circuits:
        ws = wb.create_sheet("Контуры ИТП")
        ws.append(["КОНТУРЫ ИНДИВИДУАЛЬНОГО ТЕПЛОВОГО ПУНКТА"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])

        # Heating circuits
        if has_heating_circuits:
            ws.append(["Контуры отопления"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            ws.append(["Контур", "ИТП", "Тип", "t_под, °C", "t_обр, °C",
                       "Смеситель", "Q, кВт", "Расход, м³/ч",
                       "Насос (модель)", "Напор, м", "Привязан AHU", "Примечание"])
            style_header(ws[ws.max_row])
            for cname, circ in sorted(project.heating_circuits.items()):
                net = project.pipe_networks.get(cname)
                q_kw = net.total_heat_load_w / 1000.0 if net else 0.0
                flow = getattr(net, "pump_flow_m3_h", 0.0) if net else 0.0
                pump = getattr(net, "pump_model", "") if net else (circ.pump_model or "—")
                head = getattr(net, "pump_head_m", 0.0) if net else circ.pump_head_m
                ws.append([
                    cname, circ.parent_system, circ.circuit_type,
                    circ.t_supply, circ.t_return,
                    "✓" if circ.has_mixing_node else "—",
                    round(q_kw, 1), round(flow, 2),
                    pump or "—", round(head, 1),
                    circ.serves_ahu or "—", circ.note,
                ])
            ws.append([])

        # Cooling circuits
        if has_cooling_circuits:
            ws.append(["Контуры холодоснабжения"])
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
            ws.append(["Контур", "Источник", "Тип", "t_под, °C", "t_обр, °C",
                       "Изоляция", "Q, кВт", "Расход, м³/ч",
                       "Насос (модель)", "Напор, м", "Привязан AHU", "Примечание"])
            style_header(ws[ws.max_row])
            for cname, circ in sorted(project.cooling_circuits.items()):
                net = project.cooling_pipe_networks.get(cname)
                q_kw = net.total_heat_load_w / 1000.0 if net else 0.0
                flow = getattr(net, "pump_flow_m3_h", 0.0) if net else 0.0
                pump = getattr(net, "pump_model", "") if net else (circ.pump_model or "—")
                head = getattr(net, "pump_head_m", 0.0) if net else circ.pump_head_m
                ws.append([
                    cname, circ.parent_system, circ.circuit_type,
                    circ.t_supply, circ.t_return,
                    "✓" if circ.insulated else "—",
                    round(q_kw, 1), round(flow, 2),
                    pump or "—", round(head, 1),
                    circ.serves_ahu or "—", circ.note,
                ])
            ws.append([])
        autofit(ws, 12)

    # ===== Лист "Циркуляц. насосы" — сводка по всем контурам =====
    all_nets = []
    for n in getattr(project, "pipe_networks", {}).values():
        all_nets.append(("Отопление", n))
    for n in getattr(project, "cooling_pipe_networks", {}).values():
        all_nets.append(("Холод", n))
    if all_nets:
        ws = wb.create_sheet("Циркуляц. насосы")
        ws.append(["ПОДБОР ЦИРКУЛЯЦИОННЫХ НАСОСОВ"])
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws.append([])
        ws.append(["Среда", "Контур", "ИТП/источник", "Тип контура",
                   "Q, кВт", "Расход, м³/ч", "Δp сети, кПа",
                   "Напор насоса (с зап. 1.3), м", "Подобранная модель"])
        style_header(ws[ws.max_row])
        for medium_label, net in all_nets:
            pump_head_m = getattr(net, "pump_head_m", 0.0) or (
                net.total_pressure_loss_pa / (977.7 * 9.81) * 1.3)
            pump_flow = (getattr(net, "pump_flow_m3_h", 0.0) or
                         net.total_flow_kg_h / 977.7)
            ws.append([
                medium_label,
                net.system_name,
                getattr(net, "parent_system", "") or "—",
                getattr(net, "circuit_type", "") or "—",
                round(net.total_heat_load_w / 1000.0, 1),
                round(pump_flow, 2),
                round(net.total_pressure_loss_pa / 1000.0, 2),
                round(pump_head_m, 1),
                getattr(net, "pump_model", "") or "—",
            ])
        autofit(ws, 9)
