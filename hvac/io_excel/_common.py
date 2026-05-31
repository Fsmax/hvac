# -*- coding: utf-8 -*-
"""Общие стили и помощники для Excel-экспорта."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

thin = Side(border_style="thin", color="888888")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
head_fill = PatternFill("solid", fgColor="1F4E78")
head_font = Font(bold=True, color="FFFFFF")
sum_fill = PatternFill("solid", fgColor="DCE6F1")
sum_font = Font(bold=True)


def style_header(row):
    for c in row:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = border


def autofit(ws, n_cols):
    for i in range(1, n_cols + 1):
        letter = get_column_letter(i)
        max_len = 8
        for cell in ws[letter]:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)) + 2, 40))
        ws.column_dimensions[letter].width = max_len
