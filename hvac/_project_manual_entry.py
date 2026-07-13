# -*- coding: utf-8 -*-
"""ManualEntryMixin — ручной ввод помещений и ограждений (v3.8).

Используется когда проект собирается без CSV из Revit: пользователь
сам добавляет Space и BoundaryElement через UI.
"""

from __future__ import annotations
from typing import List, Optional

from hvac.models import Space, BoundaryElement, Construction
from hvac.catalogs.constructions import (
    DEFAULT_U_BY_CATEGORY, construction_key, normalize_category,
)
from hvac.catalogs.room_types import apply_room_type_defaults


_ORIENT_DEG_MAP = {
    "N": 0.0, "NE": 45.0, "E": 90.0, "SE": 135.0,
    "S": 180.0, "SW": 225.0, "W": 270.0, "NW": 315.0,
}


class ManualEntryMixin:
    """Методы ручного создания/удаления/редактирования помещений и ограждений."""

    # ---------- Помещения ----------
    def add_space(self, number: str, name: str, level: str,
                  area_m2: float, height_m: float = 3.0,
                  room_type: str = "Прочее",
                  volume_m3: Optional[float] = None,
                  space_id: Optional[str] = None) -> Space:
        """Создаёт новое помещение и добавляет в проект.

        Используется для ручного ввода (без Revit).
        Если space_id не задан — генерируется автоматически.
        Возвращает созданный объект Space.
        """
        if space_id is None:
            # Простой автоинкремент: M-001, M-002, ... (M = manual)
            existing = [s.space_id for s in self.spaces]
            i = 1
            while f"M-{i:04d}" in existing:
                i += 1
            space_id = f"M-{i:04d}"
        elif space_id in self._space_by_id:
            raise ValueError(f"space_id '{space_id}' уже существует")

        if volume_m3 is None:
            volume_m3 = area_m2 * height_m

        sp = Space(
            space_id=space_id,
            number=number,
            name=name,
            level=level,
            area_m2=area_m2,
            volume_m3=volume_m3,
            height_m=height_m,
            room_type=room_type,
            manual_entry=True,
            user_modified=True,
        )
        # Применяем типовые параметры по типу помещения
        apply_room_type_defaults(sp)
        sp.user_modified = True   # apply_room_type_defaults снимает флаг
        sp.manual_entry = True

        self.spaces.append(sp)
        self._space_by_id[space_id] = sp
        self.emit("spaces_changed")
        return sp

    def remove_space(self, space_id: str) -> bool:
        """Удаляет помещение и ВСЕ его ограждения. Возвращает True если удалено."""
        sp = self._space_by_id.pop(space_id, None)
        if sp is None:
            return False
        self.spaces = [s for s in self.spaces if s.space_id != space_id]
        self.elements = [e for e in self.elements if e.space_id != space_id]
        self._invalidate_elements_index()
        self.emit("spaces_changed")
        return True

    def update_space(self, space_id: str, **fields) -> bool:
        """Изменяет произвольные поля помещения. Помечает user_modified=True."""
        sp = self._space_by_id.get(space_id)
        if sp is None:
            return False
        for k, v in fields.items():
            if hasattr(sp, k):
                setattr(sp, k, v)
        sp.user_modified = True
        # Если изменились area_m2 или height_m — пересчитываем volume_m3
        if "area_m2" in fields or "height_m" in fields:
            if sp.height_m > 0 and sp.area_m2 > 0 and "volume_m3" not in fields:
                sp.volume_m3 = sp.area_m2 * sp.height_m
        self.emit("spaces_changed")
        return True

    # ---------- Ограждения ----------
    def add_element(self, space_id: str, row_type: str, category: str,
                    family: str, type_name: str,
                    area_m2: float,
                    is_exterior: bool = True,
                    orientation: str = "",
                    thickness_mm: float = 0.0,
                    u_value: float = 0.0,
                    shgc: float = 0.6,
                    element_id: Optional[str] = None,
                    host_element_id: str = "",
                    boundary_length_m: float = 0.0) -> BoundaryElement:
        """Создаёт новый граничный элемент (стена / окно / дверь) вручную.

        row_type: "external_wall" (для стен, перекрытий, покрытий) или "opening"
                  (для окон, витражей, дверей).
        Возвращает созданный BoundaryElement.
        """
        if space_id not in self._space_by_id:
            raise ValueError(f"Помещение '{space_id}' не найдено")

        if element_id is None:
            existing = {e.element_id for e in self.elements}
            i = 1
            while f"E-{space_id}-{i:03d}" in existing:
                i += 1
            element_id = f"E-{space_id}-{i:03d}"

        orient_deg = _ORIENT_DEG_MAP.get(orientation, None)

        sp = self._space_by_id[space_id]
        height = sp.height_m if sp.height_m > 0 else 3.0

        elem = BoundaryElement(
            space_id=space_id,
            row_type=row_type,
            is_exterior=is_exterior,
            element_id=element_id,
            category=category,
            family=family,
            type_name=type_name,
            boundary_length_m=boundary_length_m or (area_m2 / height if height else 0),
            space_height_m=height,
            approx_area_m2=area_m2,
            element_area_m2=area_m2,
            thickness_mm=thickness_mm,
            function="Exterior" if is_exterior else "Interior",
            host_element_id=host_element_id,
            boundary_space_count=1,
            orientation_deg=orient_deg,
            orientation=orientation,
            u_value=u_value,
            net_area_m2=area_m2,
            manual_entry=True,
        )
        self.elements.append(elem)
        self._invalidate_elements_index()

        # Обновляем каталог конструкций (если новая конструкция)
        cat = normalize_category(category, family, type_name)
        key = construction_key(cat, family, type_name, thickness_mm)
        elem.construction_key = key
        if key not in self.constructions:
            self.constructions[key] = Construction(
                key=key, category=cat, family=family, type_name=type_name,
                thickness_mm=thickness_mm, u_value=u_value, shgc=shgc,
            )
        else:
            if u_value > 0:
                self.constructions[key].u_value = u_value
            if shgc > 0 and cat in ("Окна", "Витраж"):
                self.constructions[key].shgc = shgc

        self._recompute_net_areas()
        self.emit("elements_changed")
        return elem

    def remove_element(self, element_id: str) -> bool:
        """Удаляет граничный элемент по его element_id."""
        before = len(self.elements)
        self.elements = [e for e in self.elements if e.element_id != element_id]
        if len(self.elements) == before:
            return False
        self._invalidate_elements_index()
        self._recompute_net_areas()
        self.emit("elements_changed")
        return True

    def update_element(self, element_id: str, *,
                       in_space: Optional[str] = None, **fields) -> bool:
        """Изменяет поля граничного элемента.

        in_space: если задан space_id — правится элемент именно ЭТОГО
        помещения. Это важно для общих стен/витражей: один и тот же
        element_id присутствует у нескольких помещений (одна запись на
        помещение), и без скоупа обновился бы первый совпавший глобально —
        т.е. ограждение ЧУЖОГО помещения, а текущее осталось бы без правки
        (например, пометка «внутреннее» не влияла бы на расчёт).

        Без in_space — старое поведение (первый совпавший по element_id).
        Возвращает True, если хоть один элемент изменён.
        """
        found = False
        need_net = False
        moved = False
        for e in self.elements:
            if e.element_id != element_id:
                continue
            if in_space is not None and e.space_id != in_space:
                continue
            for k, v in fields.items():
                if hasattr(e, k):
                    setattr(e, k, v)
            e.user_modified = True   # для сохранения как element_overrides
            if "orientation" in fields:
                e.orientation_deg = _ORIENT_DEG_MAP.get(
                    e.orientation, e.orientation_deg)
            if "approx_area_m2" in fields or "element_area_m2" in fields:
                need_net = True
            if "space_id" in fields:
                moved = True
            found = True
            if in_space is None:
                break   # совместимость: только первый совпавший
        if need_net:
            self._recompute_net_areas()
        if moved:
            self._invalidate_elements_index()
        if found:
            self.emit("elements_changed")
        return found

    def set_rooms_exterior(self, space_ids, is_exterior: bool) -> int:
        """Массово помечает все наружные стены и проёмы указанных помещений
        как наружные (is_exterior=True) или внутренние (False).

        «Внутреннее» помещение не теряет тепло через ограждения — в расчёте
        теплопотерь остаётся только инфильтрация. Полезно, когда Revit
        ошибочно дал «наружные» стены помещению, полностью окружённому
        другими (коридор, санузел, кладовая в ядре здания).

        Затрагивает только стены/проёмы (row_type external_wall/opening);
        пол по грунту и покрытие (флаги помещения has_floor_to_ground/has_roof)
        не трогаются. Помещения помечаются user_modified. Пересчёт теплопотерь —
        на стороне вызывающего (project.recalculate()).

        Возвращает число фактически изменённых элементов.
        """
        ids = set(space_ids)
        changed = 0
        for e in self.elements:
            if (e.space_id in ids
                    and e.row_type in ("external_wall", "opening")
                    and e.is_exterior != is_exterior):
                e.is_exterior = is_exterior
                e.user_modified = True   # для сохранения как element_overrides
                changed += 1
        if changed:
            for sid in ids:
                sp = self._space_by_id.get(sid)
                if sp is not None:
                    sp.user_modified = True
            self.emit("elements_changed")
        return changed

    def set_elements_exterior(self, pairs, is_exterior: bool) -> int:
        """Массово помечает конкретные ограждения наружными/внутренними.

        pairs: итерируемое из (space_id, element_id) — точечный выбор
        элементов (для общепроектного редактора ограждений, где видно
        сразу все стены/проёмы со всех помещений). Затрагивает только
        стены/проёмы. Возвращает число изменённых элементов; пересчёт —
        на стороне вызывающего.
        """
        want = {(s, e) for s, e in pairs}
        changed = 0
        touched_spaces = set()
        for el in self.elements:
            if ((el.space_id, el.element_id) in want
                    and el.row_type in ("external_wall", "opening")
                    and el.is_exterior != is_exterior):
                el.is_exterior = is_exterior
                el.user_modified = True
                touched_spaces.add(el.space_id)
                changed += 1
        if changed:
            for sid in touched_spaces:
                sp = self._space_by_id.get(sid)
                if sp is not None:
                    sp.user_modified = True
            self.emit("elements_changed")
        return changed

    # ---------- Конструкции (каталог) ----------
    def create_construction(self, category: str, family: str = "",
                            type_name: str = "", thickness_mm: float = 0.0,
                            u_value: float = 0.0, shgc: float = 0.0,
                            note: str = "") -> Construction:
        """Создаёт новую конструкцию в каталоге вручную.

        Ключ формируется из категории/семейства/типа/толщины. Если такой
        ключ уже есть — поднимает ValueError. Категория обязательна, плюс
        нужно хотя бы семейство или тип (иначе ключ вырожденный).
        """
        category = (category or "").strip()
        family = (family or "").strip()
        type_name = (type_name or "").strip()
        if not category:
            raise ValueError("Категория обязательна")
        if not family and not type_name:
            raise ValueError("Укажите семейство или тип конструкции")
        cat = normalize_category(category, family, type_name)
        key = construction_key(cat, family, type_name, thickness_mm)
        if key in self.constructions:
            raise ValueError(f"Конструкция '{key}' уже существует")
        c = Construction(
            key=key, category=cat, family=family, type_name=type_name,
            thickness_mm=thickness_mm,
            u_value=u_value or DEFAULT_U_BY_CATEGORY.get(cat, 0.5),
            shgc=shgc, note=note or "Создано вручную",
        )
        self.constructions[key] = c
        self.emit("constructions_changed")
        return c

    def update_construction(self, old_key: str, *, category: str, family: str,
                            type_name: str, thickness_mm: float,
                            u_value: float, shgc: float,
                            note: str = "") -> Construction:
        """Изменяет конструкцию. При смене идентификации (категория/семейство/
        тип/толщина) пересобирает ключ и переносит ссылки у элементов.

        Поднимает ValueError, если ключ занят другой конструкцией.
        """
        c = self.constructions.get(old_key)
        if c is None:
            raise ValueError(f"Конструкция '{old_key}' не найдена")
        category = (category or "").strip()
        family = (family or "").strip()
        type_name = (type_name or "").strip()
        if not category:
            raise ValueError("Категория обязательна")
        if not family and not type_name:
            raise ValueError("Укажите семейство или тип конструкции")
        cat = normalize_category(category, family, type_name)
        new_key = construction_key(cat, family, type_name, thickness_mm)
        if new_key != old_key and new_key in self.constructions:
            raise ValueError(f"Конструкция '{new_key}' уже существует")

        c.category = cat
        c.family = family
        c.type_name = type_name
        c.thickness_mm = thickness_mm
        c.u_value = u_value
        c.shgc = shgc
        c.note = note
        if new_key != old_key:
            c.key = new_key
            del self.constructions[old_key]
            self.constructions[new_key] = c
            # Переносим ссылки элементов на новый ключ И синхронизируем их
            # идентификационные поля: apply_constructions() пересобирает ключ
            # из el.category/family/type_name/thickness_mm, и без синхронизации
            # первый же пересчёт «воскрешал» старую запись в каталоге, а
            # элементы перепривязывались обратно (переименование выглядело
            # как дублирование). Элементы находим и по construction_key, и по
            # идентификационным полям — у внутренних ограждений ключ может
            # быть ещё не присвоен.
            for el in self.elements:
                el_cat = normalize_category(el.category, el.family,
                                            el.type_name)
                el_key = construction_key(el_cat, el.family, el.type_name,
                                          el.thickness_mm)
                if el.construction_key == old_key or el_key == old_key:
                    el.construction_key = new_key
                    el.category = cat
                    el.family = family
                    el.type_name = type_name
                    el.thickness_mm = thickness_mm
                    # В CSV-режиме поля элемента восстанавливаются из CSV;
                    # без флага правка слетит при перезагрузке проекта
                    # (element_overrides сохраняются только для user_modified).
                    el.user_modified = True
        self.emit("constructions_changed")
        return c

    def delete_construction(self, key: str) -> int:
        """Удаляет конструкцию из каталога. Возвращает число элементов,
        которые на неё ссылались (становятся «осиротевшими» до пересчёта)."""
        if key not in self.constructions:
            return 0
        n = sum(1 for el in self.elements if el.construction_key == key)
        del self.constructions[key]
        self.emit("constructions_changed")
        return n

    def get_room_elements(self, space_id: str) -> list:
        """Возвращает все граничные элементы помещения."""
        return list(self.elements_for(space_id))

    def duplicate_space(self, space_id: str,
                        new_number: Optional[str] = None) -> Optional[Space]:
        """Создаёт копию помещения вместе с его ограждениями.
        Возвращает новое помещение или None если исходное не найдено."""
        from copy import deepcopy
        src = self._space_by_id.get(space_id)
        if src is None:
            return None
        new_sp = deepcopy(src)
        # сгенерировать уникальный space_id
        existing = {s.space_id for s in self.spaces}
        i = 1
        while f"{src.space_id}-copy{i}" in existing:
            i += 1
        new_sp.space_id = f"{src.space_id}-copy{i}"
        if new_number is not None:
            new_sp.number = new_number
        else:
            new_sp.number = f"{src.number} (копия)"
        new_sp.user_modified = True
        new_sp.manual_entry = True
        # Копии не несут результаты расчёта от оригинала
        new_sp.heat_loss_w = 0.0
        new_sp.heat_gain_w = 0.0
        new_sp.heat_loss_breakdown = {}
        new_sp.heat_gain_breakdown = {}
        self.spaces.append(new_sp)
        self._space_by_id[new_sp.space_id] = new_sp

        # Скопировать ограждения
        for el in [e for e in self.elements if e.space_id == space_id]:
            new_el = deepcopy(el)
            new_el.space_id = new_sp.space_id
            existing_eids = {e.element_id for e in self.elements}
            j = 1
            while f"{el.element_id}-c{j}" in existing_eids:
                j += 1
            new_el.element_id = f"{el.element_id}-c{j}"
            new_el.manual_entry = True
            self.elements.append(new_el)

        self._invalidate_elements_index()
        self.emit("spaces_changed")
        return new_sp

    # ---------- Массовое создание по шаблону ----------
    def add_spaces_from_template(self, template) -> List[Space]:
        """Создаёт помещения по шаблону «N этажей × M квартир × K комнат».

        template — объект с полями:
            n_floors: int
            first_floor_number: int
            apartments_per_floor: int
            rooms_per_apartment: list[(name, room_type, area_m2)] (объекты
                с этими атрибутами)
            height_m: float
            level_prefix: str

        Возвращает список созданных помещений.
        """
        created: List[Space] = []
        for floor_offset in range(template.n_floors):
            floor_num = template.first_floor_number + floor_offset
            level_name = f"{template.level_prefix}{floor_num}"
            for apt_idx in range(1, template.apartments_per_floor + 1):
                for room in template.rooms_per_apartment:
                    number = f"{floor_num:02d}-{apt_idx:02d}-{room.name[:3]}"
                    name = f"кв.{apt_idx} · {room.name}"
                    sp = self.add_space(
                        number=number,
                        name=name,
                        level=level_name,
                        area_m2=room.area_m2,
                        height_m=template.height_m,
                        room_type=room.room_type,
                    )
                    created.append(sp)
        return created

    # ---------- Импорт из табличного файла ----------
    def import_spaces_from_excel(self, path: str,
                                 sheet: Optional[str] = None) -> int:
        """Импорт списка помещений из Excel.

        Ожидаемые колонки (заголовки регистронезависимые, синонимы поддерживаются):
            №, номер, number       — номер помещения
            имя, название, name    — название
            этаж, level, floor     — этаж
            тип, type              — тип (по справочнику)
            площадь, area, s, м²   — площадь, м²
            высота, height, h      — высота, м (опционально)

        Возвращает количество успешно добавленных строк.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise RuntimeError("Для импорта из Excel нужен openpyxl") from e
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet] if sheet else wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return 0
        mapping = _resolve_header_mapping(headers)
        if "number" not in mapping or "area" not in mapping:
            raise ValueError(
                "В таблице должны быть колонки с номером и площадью.")
        n = self._import_rows(rows, mapping)
        self.emit("spaces_changed")
        return n

    def import_spaces_from_csv(self, path: str,
                               delimiter: Optional[str] = None) -> int:
        """Аналог import_spaces_from_excel но для CSV."""
        import csv
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            sample = f.read(4096)
            f.seek(0)
            if delimiter is None:
                try:
                    dialect = csv.Sniffer().sniff(sample,
                                                  delimiters=",;\t|")
                    delimiter = dialect.delimiter
                except csv.Error:
                    delimiter = ","
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, None)
            if not headers:
                return 0
            mapping = _resolve_header_mapping(headers)
            if "number" not in mapping or "area" not in mapping:
                raise ValueError(
                    "В таблице должны быть колонки с номером и площадью.")
            n = self._import_rows(reader, mapping)
        self.emit("spaces_changed")
        return n

    def _import_rows(self, rows, mapping: dict) -> int:
        n = 0
        for row in rows:
            if row is None:
                continue
            row = list(row)
            number = _cell(row, mapping.get("number"))
            area = _cell_float(row, mapping.get("area"))
            if not number or not area:
                continue
            name = _cell(row, mapping.get("name")) or number
            level = _cell(row, mapping.get("level")) or "1 этаж"
            rtype = _cell(row, mapping.get("type")) or ""
            height = _cell_float(row, mapping.get("height")) or 3.0
            if not rtype:
                from hvac.catalogs.room_types import auto_detect_room_type
                rtype = auto_detect_room_type(name)
            try:
                self.add_space(number=str(number), name=str(name),
                                level=str(level), area_m2=area,
                                height_m=height, room_type=rtype)
                n += 1
            except ValueError:
                # Пропустить дубликаты, продолжить
                continue
        return n

    # ---------- Оборудование помещения ----------
    def set_room_equipment(self, space_id: str, **kwargs) -> bool:
        """Назначает поля оборудования для помещения.
        Создаёт RoomEquipment если ещё нет."""
        sp = self._space_by_id.get(space_id)
        if sp is None:
            return False
        eq = sp.get_or_create_equipment()
        for k, v in kwargs.items():
            if hasattr(eq, k):
                setattr(eq, k, v)
        self.emit("equipment_changed")
        return True

    def apply_room_equipment(self, space_ids, fields: dict) -> int:
        """Применяет один набор полей оборудования к нескольким помещениям
        (групповое назначение / вставка из буфера). Возвращает число
        затронутых помещений. Событие equipment_changed эмитится один раз."""
        n = 0
        for sid in space_ids:
            sp = self._space_by_id.get(sid)
            if sp is None:
                continue
            eq = sp.get_or_create_equipment()
            for k, v in fields.items():
                if hasattr(eq, k):
                    setattr(eq, k, v)
            n += 1
        if n:
            self.emit("equipment_changed")
        return n

    def clear_room_equipment(self, space_ids) -> int:
        """Снимает оборудование с указанных помещений (room_equipment=None).
        Возвращает число очищенных (где оборудование было)."""
        n = 0
        for sid in space_ids:
            sp = self._space_by_id.get(sid)
            if sp is None or sp.room_equipment is None:
                continue
            sp.room_equipment = None
            n += 1
        if n:
            self.emit("equipment_changed")
        return n


# ===== Импорт-хелперы =====
_HEADER_SYNONYMS = {
    "number": ("№", "номер", "number", "no", "no."),
    "name":   ("имя", "название", "name", "title", "помещение"),
    "level":  ("этаж", "уровень", "level", "floor"),
    "type":   ("тип", "type", "room_type", "категория"),
    "area":   ("площадь", "area", "s", "м²", "m2", "sq m"),
    "height": ("высота", "height", "h", "h_m", "м"),
}


def _resolve_header_mapping(headers) -> dict:
    """Сопоставляет реальные заголовки таблицы с каноническими именами."""
    mapping: dict = {}
    for idx, raw in enumerate(headers):
        if raw is None:
            continue
        norm = str(raw).strip().lower()
        for canon, synonyms in _HEADER_SYNONYMS.items():
            if norm in synonyms or any(syn in norm for syn in synonyms):
                mapping.setdefault(canon, idx)
                break
    return mapping


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return str(v).strip() if v is not None else ""


def _cell_float(row, idx) -> float:
    if idx is None or idx >= len(row):
        return 0.0
    v = row[idx]
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0
