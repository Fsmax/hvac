# -*- coding: utf-8 -*-
"""Экспорт результатов расчёта в master-таблицу проекта (HLGC Design Table).

Структура таблицы:
- Лист "HLGC" с заголовками в строках 2-11, данными от строки ~13.
- Колонка 3 = номер комнаты (например "B00-043") — ключ для сопоставления.
- Многоуровневые заголовки описывают группы: ROOM / AREA / VOLUME /
  ROOM DESIGN TEMPERATURE / PEOPLE / ROOM DESIGN COOLING LOAD /
  ROOM HEATING LOADS / VENTILATION REQUIREMENT / ROOM AIR FLOW RATE /
  SYSTEM SELECTION / FAN-COIL LIST / RADIATOR LIST.

Два движка записи:
1. **COM (Excel)** — основной, для .xls с внешними ссылками. Excel сам
   корректно сохраняет файл со всеми формулами, ссылками, стилями.
   Требует pywin32 + установленный Excel.
2. **openpyxl** — fallback, если COM недоступен. Не поддерживает корректное
   сохранение внешних ссылок (.xls файл может потребовать восстановления
   при открытии).

Программа автоматически выбирает COM если он доступен.

Колонки-приёмники (1-индексация Excel) — см. HLGC_COLUMN_MAP.
"""

from __future__ import annotations
import os
import shutil
from typing import TYPE_CHECKING, Dict, List, Optional, Tuple

if TYPE_CHECKING:
    from hvac.project import HVACProject
    from hvac.models import Space


# Колонки для идентификации помещения (заполняются для новых строк)
ROOM_ID_COLUMNS = {
    3: "number",       # B00-043
    5: "name",         # Русское имя
}

# Максимальная колонка которую копируем при создании новых строк (формат)
MAX_COL_FOR_COPY = 62


# Колонки для записи (1-индексация Excel)
HLGC_COLUMN_MAP = {
    # Геометрия
    6:  ("area_m2",            "Площадь, м²"),
    7:  ("height_m",           "Высота, м"),
    8:  ("volume_m3",          "Объём, м³"),
    # Температуры
    9:  ("t_in_cool",          "tвн летом, °C"),
    11: ("t_in_heat",          "tвн зимой, °C"),
    # Люди
    14: ("occupancy_people",   "Чел."),
    # Теплопоступления (cooling load)
    21: ("heat_gain_sensible_w", "Q_охл явная, Вт"),
    22: ("heat_gain_latent_w",   "Q_охл скрытая, Вт"),
    23: ("heat_gain_w",          "Q_охл итого, Вт"),
    24: ("__q_cool_per_m2",      "Q_охл уд., Вт/м²"),
    # Теплопотери (heating load)
    25: ("heat_loss_w",          "Q_отопл, Вт"),
    26: ("__q_heat_per_m2",      "Q_отопл уд., Вт/м²"),
    # Вентиляция
    28: ("__vent_rate_per_occ",  "Норма свеж., м³/ч·чел"),
    29: ("ach_calculated",       "ACH, 1/ч"),
    30: ("__min_fresh_air",      "Мин. свеж. воздух, м³/ч"),
    31: ("supply_m3h",           "Приток, м³/ч"),
    32: ("exhaust_m3h",          "Вытяжка, м³/ч"),
    33: ("hood_m3h",             "Зонт, м³/ч"),
    # Системы
    35: ("system_ventilation",   "Система — приток"),
    36: ("system_ventilation",   "Система — вытяжка"),
    # Прочее
    38: ("room_type",            "Тип помещения"),
}

HLGC_SHEET_NAME = "HLGC"
ROOM_NUMBER_COLUMN = 3
DATA_START_ROW = 12

# Сколько знаков после запятой для каждой колонки
_ROUND_DIGITS = {
    6: 2, 7: 2, 8: 2,
    9: 1, 11: 1,
    21: 0, 22: 0, 23: 0,
    24: 1,
    25: 0, 26: 1,
    28: 1, 29: 2, 30: 1,
    31: 1, 32: 1, 33: 1,
}


def _get_value_for_field(sp: "Space", field: str, col: int = 0):
    """Возвращает значение из Space по имени поля."""
    val = None
    if field.startswith("__"):
        if field == "__q_cool_per_m2":
            val = sp.heat_gain_w / sp.area_m2 if sp.area_m2 else 0
        elif field == "__q_heat_per_m2":
            val = sp.heat_loss_w / sp.area_m2 if sp.area_m2 else 0
        elif field == "__vent_rate_per_occ":
            br = sp.ventilation_breakdown or {}
            val = br.get("fresh_air_per_person", 0)
        elif field == "__min_fresh_air":
            br = sp.ventilation_breakdown or {}
            val = br.get("fresh_air_per_person", 0) * sp.occupancy_people
    else:
        val = getattr(sp, field, None)

    if val is None:
        return None
    if isinstance(val, float):
        digits = _ROUND_DIGITS.get(col, 2)
        return round(val, digits)
    return val


def _is_com_available() -> bool:
    """Проверяет можно ли использовать Excel COM (pywin32 + Excel установлены)."""
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Quit()
        return True
    except Exception:
        return False


# ============================================================================
# Движок 1: Excel COM (предпочтительный — сохраняет всё)
# ============================================================================

def _collect_room_values(sp: "Space", write_id_columns: bool = False
                          ) -> Dict[int, "object"]:
    """Собирает {col: value} для одной комнаты — без вызовов COM."""
    result: Dict[int, object] = {}
    if write_id_columns:
        for col, field in ROOM_ID_COLUMNS.items():
            v = getattr(sp, field, "")
            if v:
                result[col] = v
    for col, (field, _label) in HLGC_COLUMN_MAP.items():
        v = _get_value_for_field(sp, field, col)
        if v is not None and v != "":
            result[col] = v
    return result


def _batch_write_rows_com(ws, row_changes: List[Tuple[int, Dict[int, object]]],
                           overwrite_only_empty: bool,
                           preserve_formulas: bool,
                           max_col: int = MAX_COL_FOR_COPY) -> int:
    """Пакетная запись изменений в несколько строк через массивы.

    Один Range.Value на чтение существующих значений и формул, один на
    запись. Это в 20-50 раз быстрее чем cell.Value = ... на каждую ячейку.

    row_changes: список (row_number, {col: value}).
    Возвращает общее число записанных ячеек.
    """
    if not row_changes:
        return 0

    # Объединяем смежные строки в один range для пакетной операции
    row_changes.sort(key=lambda t: t[0])
    total_written = 0

    # Группируем по непрерывным диапазонам строк
    groups: List[List[Tuple[int, Dict[int, object]]]] = []
    current_group: List[Tuple[int, Dict[int, object]]] = []
    prev_row = None
    for r, changes in row_changes:
        if prev_row is None or r == prev_row + 1:
            current_group.append((r, changes))
        else:
            if current_group:
                groups.append(current_group)
            current_group = [(r, changes)]
        prev_row = r
    if current_group:
        groups.append(current_group)

    # Обрабатываем каждую группу одним пакетом
    for group in groups:
        first_row = group[0][0]
        last_row = group[-1][0]
        # Читаем существующие значения и формулы (одним вызовом каждый)
        rng = ws.Range(ws.Cells(first_row, 1),
                        ws.Cells(last_row, max_col))
        existing_values = rng.Value
        existing_formulas = rng.Formula

        # COM возвращает tuple of tuples для многострочного диапазона,
        # а для одной строки — tuple. Нормализуем к 2D.
        if not isinstance(existing_values, tuple):
            # Скаляр (одна ячейка) — не наш случай, но защитимся
            continue
        if not isinstance(existing_values[0], tuple):
            existing_values = (existing_values,)
            existing_formulas = (existing_formulas,)

        # Готовим новые значения как 2D-список
        new_values = [list(row) for row in existing_values]

        for r, changes in group:
            local_idx = r - first_row
            for col, val in changes.items():
                col_idx = col - 1
                if col_idx >= max_col:
                    continue
                existing = existing_values[local_idx][col_idx]
                formula = existing_formulas[local_idx][col_idx]
                # Пропускаем формулы
                if preserve_formulas and isinstance(formula, str) \
                        and formula.startswith("="):
                    continue
                # Пропускаем непустые если режим overwrite_only_empty
                if overwrite_only_empty and existing not in (None, ""):
                    continue
                new_values[local_idx][col_idx] = val
                total_written += 1

        # Один вызов записи на весь диапазон
        try:
            rng.Value = tuple(tuple(row) for row in new_values)
        except Exception:
            # Fallback: построчная запись (если merged-ячейки мешают)
            for i, row_vals in enumerate(new_values):
                try:
                    target = ws.Range(
                        ws.Cells(first_row + i, 1),
                        ws.Cells(first_row + i, max_col))
                    target.Value = (tuple(row_vals),)
                except Exception:
                    pass

    return total_written


def _copy_row_format_com(ws, src_row: int, dst_row: int,
                          excel, max_col: int = MAX_COL_FOR_COPY) -> None:
    """Копирует формат + формулы из src_row в dst_row."""
    _copy_format_to_range_com(ws, src_row, dst_row, dst_row, excel, max_col)


def _copy_format_to_range_com(ws, src_row: int,
                                first_dst_row: int, last_dst_row: int,
                                excel, max_col: int = MAX_COL_FOR_COPY) -> None:
    """Копирует формат + формулы из src_row сразу в диапазон строк
    [first_dst_row..last_dst_row]. Excel сделает это пакетно (формулы
    автоматически адаптируются для каждой целевой строки).

    Это в N раз быстрее чем N отдельных вызовов _copy_row_format_com."""
    if first_dst_row > last_dst_row:
        return
    source = ws.Range(ws.Cells(src_row, 1), ws.Cells(src_row, max_col))
    target = ws.Range(ws.Cells(first_dst_row, 1),
                       ws.Cells(last_dst_row, max_col))
    source.Copy()
    try:
        # -4123 = xlPasteFormulas (с адаптацией ссылок)
        target.PasteSpecial(Paste=-4123)
        # -4122 = xlPasteFormats
        target.PasteSpecial(Paste=-4122)
    finally:
        excel.CutCopyMode = False


def _export_via_com(project: "HVACProject", source_path: str,
                     output_path: str,
                     overwrite_only_empty: bool,
                     preserve_formulas: bool,
                     mode: str = "match") -> Dict:
    """Запись через Excel COM. Excel сам открывает файл, изменяет ячейки
    и сохраняет — все внешние ссылки, формулы, стили и форматирование
    сохраняются 1-в-1.

    Режимы:
    - 'match'   : обновляются только строки таблицы, чей № совпадает с
                  Space.number из проекта. Все остальные строки не трогаются.
                  Помещения проекта без пары в таблице игнорируются.
    - 'append'  : как 'match', плюс помещения проекта, которых нет в таблице,
                  ДОБАВЛЯЮТСЯ в конец с копированием форматирования и
                  формул из первой строки данных шаблона.
    - 'rebuild' : вся область данных перезаписывается ВСЕМИ помещениями
                  проекта (по порядку sp.number), используя первую строку
                  данных как стилевой шаблон. Лишние строки очищаются.
    """
    import win32com.client

    src = os.path.abspath(source_path)
    out = os.path.abspath(output_path)
    out_ext = os.path.splitext(out)[1].lower()

    # Карта помещений проекта по номеру + полный отсортированный список
    spaces_by_num: Dict[str, "Space"] = {}
    for sp in project.spaces:
        key = sp.number.strip().upper()
        if key:
            spaces_by_num[key] = sp
    # Сортируем помещения: сначала по уровню, потом по номеру
    project_spaces_sorted = sorted(
        [sp for sp in project.spaces if sp.number.strip()],
        key=lambda s: (s.level or "", s.number),
    )

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    excel.AskToUpdateLinks = False
    try:
        excel.EnableEvents = False
    except Exception:
        pass

    rows_total = 0
    matched = 0
    appended = 0
    cleared = 0
    cells_written = 0
    unmatched: List[str] = []

    wb = None
    try:
        wb = excel.Workbooks.Open(src, UpdateLinks=0, ReadOnly=False,
                                   IgnoreReadOnlyRecommended=True)
        try:
            excel.Calculation = -4135   # xlCalculationManual
        except Exception:
            pass
        try:
            ws = wb.Worksheets(HLGC_SHEET_NAME)
        except Exception:
            sheet_names = [wb.Worksheets(i + 1).Name
                           for i in range(wb.Worksheets.Count)]
            raise ValueError(
                f"В файле нет листа '{HLGC_SHEET_NAME}'. "
                f"Найдены: {sheet_names}")

        # Сканируем существующие строки данных
        max_row = ws.UsedRange.Rows.Count
        existing_rows: List[Tuple[int, str]] = []   # [(row, room_num), ...]
        template_row: Optional[int] = None          # первая строка с данными
        for r in range(DATA_START_ROW, max_row + 1):
            room_num = ws.Cells(r, ROOM_NUMBER_COLUMN).Value
            if room_num is None:
                continue
            room_num = str(room_num).strip()
            if not room_num:
                continue
            existing_rows.append((r, room_num))
            if template_row is None:
                template_row = r

        rows_total = len(existing_rows)
        if template_row is None:
            # Нет ни одной строки данных в шаблоне — берём DATA_START_ROW + 1
            template_row = DATA_START_ROW + 1

        existing_numbers = {n.upper() for _, n in existing_rows}

        # =================================================================
        # Режим: 'match' / 'append'
        # =================================================================
        if mode in ("match", "append"):
            # Шаг 1: собираем все изменения для совпавших строк
            row_changes: List[Tuple[int, Dict[int, object]]] = []
            for r, room_num in existing_rows:
                sp = spaces_by_num.get(room_num.upper())
                if sp is None:
                    unmatched.append(room_num)
                    continue
                matched += 1
                changes = _collect_room_values(sp, write_id_columns=False)
                if changes:
                    row_changes.append((r, changes))
            # Пакетная запись
            cells_written += _batch_write_rows_com(
                ws, row_changes, overwrite_only_empty, preserve_formulas)

            # Шаг 2: 'append' — добавляем новые помещения в конец
            if mode == "append":
                new_spaces = [sp for sp in project_spaces_sorted
                              if sp.number.strip().upper() not in existing_numbers]
                if new_spaces:
                    last_row = existing_rows[-1][0] if existing_rows else (
                        template_row - 1)
                    first_new = last_row + 1
                    last_new = last_row + len(new_spaces)
                    # Один Copy/PasteSpecial на весь диапазон новых строк
                    _copy_format_to_range_com(
                        ws, template_row, first_new, last_new, excel)
                    # Собираем изменения для всех новых строк
                    append_changes: List[Tuple[int, Dict[int, object]]] = []
                    for i, sp in enumerate(new_spaces):
                        target_row = first_new + i
                        changes = _collect_room_values(
                            sp, write_id_columns=True)
                        if changes:
                            append_changes.append((target_row, changes))
                        appended += 1
                    cells_written += _batch_write_rows_com(
                        ws, append_changes,
                        overwrite_only_empty=False,
                        preserve_formulas=preserve_formulas)

        # =================================================================
        # Режим: 'rebuild' — все помещения проекта на месте существующих
        # =================================================================
        elif mode == "rebuild":
            last_existing_row = existing_rows[-1][0] if existing_rows else (
                template_row)
            n_rooms = len(project_spaces_sorted)
            if n_rooms > 0:
                first_target = template_row
                last_target = template_row + n_rooms - 1

                # 1. Если нужны строки за пределами существующих — копируем
                #    формат + формулы пакетно
                if last_target > last_existing_row:
                    first_extra_row = max(last_existing_row + 1, first_target)
                    _copy_format_to_range_com(
                        ws, template_row, first_extra_row, last_target, excel)

                # 2. Очищаем ВЕСЬ целевой диапазон одним пакетом
                #    (сохраняя формулы)
                _clear_rows_com(ws, first_target, last_target,
                                 MAX_COL_FOR_COPY, preserve_formulas)

                # 3. Собираем изменения для всех помещений
                rebuild_changes: List[Tuple[int, Dict[int, object]]] = []
                for i, sp in enumerate(project_spaces_sorted):
                    target_row = template_row + i
                    changes = _collect_room_values(sp, write_id_columns=True)
                    if changes:
                        rebuild_changes.append((target_row, changes))
                    matched += 1
                # 4. Пакетная запись
                cells_written += _batch_write_rows_com(
                    ws, rebuild_changes,
                    overwrite_only_empty=False,
                    preserve_formulas=preserve_formulas)

            # 5. Очищаем оставшиеся «лишние» строки пакетно
            first_extra = template_row + n_rooms
            if first_extra <= last_existing_row:
                _clear_rows_com(ws, first_extra, last_existing_row,
                                 MAX_COL_FOR_COPY, preserve_formulas)
                cleared = last_existing_row - first_extra + 1

        else:
            raise ValueError(
                f"Неизвестный режим: {mode!r}. "
                f"Допустимо: 'match', 'append', 'rebuild'.")

        # Включаем пересчёт обратно перед сохранением
        try:
            excel.Calculation = -4105   # xlCalculationAutomatic
        except Exception:
            pass
        fmt = {".xlsx": 51, ".xls": 56, ".xlsm": 52}.get(out_ext, 51)
        if os.path.exists(out):
            try:
                os.unlink(out)
            except Exception:
                pass
        wb.SaveAs(out, FileFormat=fmt)
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        try:
            excel.Quit()
        except Exception:
            pass

    return {
        "mode": mode,
        "rows_total": rows_total,
        "rows_matched": matched,
        "rows_appended": appended,
        "rows_cleared": cleared,
        "rows_unmatched": unmatched,
        "cells_written": cells_written,
        "output_path": out,
        "engine": "com",
    }


def _clear_rows_com(ws, first_row: int, last_row: int, max_col: int,
                     preserve_formulas: bool) -> None:
    """Пакетная очистка значений в диапазоне строк. Если preserve_formulas
    — формулы не трогает (читаем formulas, обнуляем только non-formula)."""
    if first_row > last_row:
        return
    rng = ws.Range(ws.Cells(first_row, 1), ws.Cells(last_row, max_col))
    if not preserve_formulas:
        # Просто очищаем всё (значения, не форматирование)
        try:
            rng.ClearContents()
        except Exception:
            pass
        return
    # Иначе — нужно сохранить формулы: читаем formulas, пишем None туда
    # где формул нет
    try:
        existing_formulas = rng.Formula
    except Exception:
        return
    # Нормализация в 2D
    if not isinstance(existing_formulas, tuple):
        return
    if not isinstance(existing_formulas[0], tuple):
        existing_formulas = (existing_formulas,)
    n_rows = len(existing_formulas)
    new_values = []
    for row_formulas in existing_formulas:
        new_row = []
        for f in row_formulas:
            if isinstance(f, str) and f.startswith("="):
                # Сохраняем формулу
                new_row.append(f)
            else:
                new_row.append(None)
        new_values.append(tuple(new_row))
    try:
        rng.Formula = tuple(new_values)
    except Exception:
        pass


def _clear_row_values_com(ws, row: int, max_col: int,
                           preserve_formulas: bool) -> None:
    """Backwards-compat обёртка над _clear_rows_com для одной строки."""
    _clear_rows_com(ws, row, row, max_col, preserve_formulas)


# ============================================================================
# Движок 2: openpyxl (fallback — без COM/Excel)
# ============================================================================

def _ensure_xlsx_via_com(source_path: str) -> Tuple[str, bool]:
    """Конвертирует .xls в .xlsx. Используется когда основной движок —
    openpyxl, но входной файл в старом формате .xls.

    Сначала пытается через Excel COM (сохраняет формулы и форматирование).
    Если pywin32 / Excel недоступны — fallback через xlrd: читает данные
    в новую .xlsx-книгу. В fallback-режиме теряется форматирование и
    формулы (они материализуются в значения).

    Если ни один путь не доступен — поднимает понятную ошибку с
    инструкцией пользователю.
    """
    ext = os.path.splitext(source_path)[1].lower()
    if ext == ".xlsx":
        return source_path, False
    if ext != ".xls":
        raise ValueError(f"Ожидался .xls или .xlsx, получено: {ext}")
    src = os.path.abspath(source_path)
    base, _ = os.path.splitext(src)
    out = base + "_temp.xlsx"
    if os.path.exists(out):
        try:
            os.unlink(out)
        except OSError:
            pass

    # --- Попытка 1: COM (предпочтительно, сохраняет формулы и стили) ---
    try:
        import win32com.client
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        try:
            wb = excel.Workbooks.Open(src, UpdateLinks=0, ReadOnly=True)
            wb.SaveAs(out, FileFormat=51)
            wb.Close(SaveChanges=False)
        finally:
            excel.Quit()
        return out, True
    except ImportError:
        com_error = "pywin32 не установлен"
    except Exception as e:
        com_error = f"COM/Excel: {e}"

    # --- Попытка 2: xlrd → openpyxl (без COM, без формул/форматирования) ---
    try:
        import xlrd
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError(
            f"Не удалось открыть .xls файл «{os.path.basename(src)}».\n\n"
            f"Причина: формат .xls (Excel 97-2003) не поддерживается "
            f"openpyxl напрямую. Для конвертации нужно одно из:\n\n"
            f"  1) Установить Microsoft Excel + pywin32:\n"
            f"     pip install pywin32\n\n"
            f"  2) Установить xlrd (без сохранения формул/форматирования):\n"
            f"     pip install xlrd==2.0.1\n\n"
            f"  3) Открыть файл в Excel вручную и сохранить как .xlsx, "
            f"затем выбрать в программе уже .xlsx-версию.\n\n"
            f"Подробности: {com_error}")

    # xlrd 2.x читает только .xls — это нам и нужно
    import logging
    logging.getLogger(__name__).warning(
        "Конвертация .xls через xlrd-fallback: формулы материализуются "
        "в значения, форматирование теряется. Для полной точности "
        "установите pywin32 (нужен также Microsoft Excel).")

    book = xlrd.open_workbook(src, formatting_info=False)
    wb = Workbook()
    # Удаляем дефолтный лист openpyxl
    wb.remove(wb.active)
    for sheet_idx in range(book.nsheets):
        sh = book.sheet_by_index(sheet_idx)
        ws = wb.create_sheet(title=sh.name[:31])  # 31 — лимит Excel
        for r in range(sh.nrows):
            for c in range(sh.ncols):
                ws.cell(row=r + 1, column=c + 1).value = sh.cell_value(r, c)
    wb.save(out)
    return out, True


def _copy_row_format_openpyxl(ws, src_row: int, dst_row: int,
                                max_col: int = MAX_COL_FOR_COPY) -> None:
    """Копирует стиль и формулы (с адаптацией ссылок) из src_row в dst_row."""
    from copy import copy
    import re
    for c in range(1, max_col + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        # Стиль
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)
        # Формула с адаптацией row reference (упрощённо: заменяем
        # все вхождения src_row на dst_row)
        if isinstance(src.value, str) and src.value.startswith("="):
            # Заменяем references вида A13 → A14 (только row part)
            adapted = re.sub(
                rf"(\$?[A-Z]+\$?){src_row}\b",
                lambda m: f"{m.group(1)}{dst_row}",
                src.value,
            )
            dst.value = adapted


def _clear_row_openpyxl(ws, row: int, max_col: int,
                         preserve_formulas: bool) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        if preserve_formulas and isinstance(cell.value, str) \
                and cell.value.startswith("="):
            continue
        cell.value = None


def _write_room_openpyxl(ws, row: int, sp: "Space",
                          overwrite_only_empty: bool,
                          preserve_formulas: bool,
                          write_id_columns: bool = False) -> int:
    n = 0
    if write_id_columns:
        for col, field in ROOM_ID_COLUMNS.items():
            val = getattr(sp, field, "")
            if val:
                ws.cell(row, col).value = val
                n += 1
    for col, (field, _label) in HLGC_COLUMN_MAP.items():
        val = _get_value_for_field(sp, field, col)
        if val is None or val == "":
            continue
        cell = ws.cell(row, col)
        existing = cell.value
        if overwrite_only_empty and existing not in (None, ""):
            continue
        if preserve_formulas and isinstance(existing, str) \
                and existing.startswith("="):
            continue
        cell.value = val
        n += 1
    return n


def _export_via_openpyxl(project: "HVACProject", source_path: str,
                          output_path: str,
                          overwrite_only_empty: bool,
                          preserve_formulas: bool,
                          mode: str = "match") -> Dict:
    """Запись через openpyxl. Поддерживает те же режимы что COM
    (match/append/rebuild). Выход — только .xlsx."""
    from openpyxl import load_workbook

    xlsx_in, converted = _ensure_xlsx_via_com(source_path)
    base, ext = os.path.splitext(output_path)
    if ext.lower() != ".xlsx":
        output_path = base + ".xlsx"

    shutil.copy(xlsx_in, output_path)
    wb = load_workbook(output_path, keep_links=False)
    if HLGC_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"В файле нет листа '{HLGC_SHEET_NAME}'. "
            f"Найдены: {wb.sheetnames}")
    ws = wb[HLGC_SHEET_NAME]

    spaces_by_num = {sp.number.strip().upper(): sp
                     for sp in project.spaces if sp.number.strip()}
    project_spaces_sorted = sorted(
        [sp for sp in project.spaces if sp.number.strip()],
        key=lambda s: (s.level or "", s.number),
    )

    # Сканируем существующие строки
    existing_rows: List[Tuple[int, str]] = []
    template_row: Optional[int] = None
    for r in range(DATA_START_ROW, ws.max_row + 1):
        room_num = ws.cell(r, ROOM_NUMBER_COLUMN).value
        if room_num is None:
            continue
        room_num = str(room_num).strip()
        if not room_num:
            continue
        existing_rows.append((r, room_num))
        if template_row is None:
            template_row = r

    rows_total = len(existing_rows)
    if template_row is None:
        template_row = DATA_START_ROW + 1
    existing_numbers = {n.upper() for _, n in existing_rows}

    matched = appended = cleared = cells_written = 0
    unmatched: List[str] = []

    if mode in ("match", "append"):
        for r, room_num in existing_rows:
            sp = spaces_by_num.get(room_num.upper())
            if sp is None:
                unmatched.append(room_num)
                continue
            matched += 1
            cells_written += _write_room_openpyxl(
                ws, r, sp, overwrite_only_empty, preserve_formulas)

        if mode == "append":
            new_spaces = [sp for sp in project_spaces_sorted
                          if sp.number.strip().upper() not in existing_numbers]
            if new_spaces:
                last_row = existing_rows[-1][0] if existing_rows else (
                    template_row - 1)
                for i, sp in enumerate(new_spaces, start=1):
                    target_row = last_row + i
                    _copy_row_format_openpyxl(ws, template_row, target_row)
                    cells_written += _write_room_openpyxl(
                        ws, target_row, sp,
                        overwrite_only_empty=False,
                        preserve_formulas=preserve_formulas,
                        write_id_columns=True)
                    appended += 1

    elif mode == "rebuild":
        last_existing_row = existing_rows[-1][0] if existing_rows else (
            template_row)
        for i, sp in enumerate(project_spaces_sorted):
            target_row = template_row + i
            if target_row > last_existing_row:
                _copy_row_format_openpyxl(ws, template_row, target_row)
            _clear_row_openpyxl(ws, target_row, MAX_COL_FOR_COPY,
                                 preserve_formulas)
            cells_written += _write_room_openpyxl(
                ws, target_row, sp,
                overwrite_only_empty=False,
                preserve_formulas=preserve_formulas,
                write_id_columns=True)
            matched += 1
        # Очищаем лишние строки
        first_extra = template_row + len(project_spaces_sorted)
        if first_extra <= last_existing_row:
            for r in range(first_extra, last_existing_row + 1):
                _clear_row_openpyxl(ws, r, MAX_COL_FOR_COPY,
                                     preserve_formulas)
                cleared += 1
    else:
        raise ValueError(
            f"Неизвестный режим: {mode!r}. "
            f"Допустимо: 'match', 'append', 'rebuild'.")

    wb.save(output_path)
    wb.close()

    if converted and os.path.exists(xlsx_in) and xlsx_in != output_path:
        try:
            os.unlink(xlsx_in)
        except Exception:
            pass

    return {
        "mode": mode,
        "rows_total": rows_total,
        "rows_matched": matched,
        "rows_appended": appended,
        "rows_cleared": cleared,
        "rows_unmatched": unmatched,
        "cells_written": cells_written,
        "output_path": output_path,
        "engine": "openpyxl",
    }


# ============================================================================
# Главный API
# ============================================================================

def export_to_hlgc(project: "HVACProject", source_path: str,
                    output_path: Optional[str] = None,
                    overwrite_only_empty: bool = False,
                    preserve_formulas: bool = True,
                    engine: str = "auto",
                    mode: str = "match",
                    ) -> Dict:
    """Записывает результаты расчёта в HLGC Design Table.

    Параметры
    ---------
    project : HVACProject с готовыми результатами
    source_path : путь к исходной таблице (.xls или .xlsx)
    output_path : куда сохранить (по умолчанию — рядом с исходником,
                  с суффиксом "_filled")
    overwrite_only_empty : True — записывать только в пустые ячейки
    preserve_formulas : True — не перезаписывать ячейки с формулами
    engine : "auto" (COM если доступен, иначе openpyxl) | "com" | "openpyxl"
    mode : 'match'   — обновить только совпадающие по № строки (как было)
           'append'  — обновить совпавшие + ДОБАВИТЬ новые помещения в конец
           'rebuild' — полностью перезаписать всю таблицу всеми помещениями
                       проекта (использовать первую строку как стилевой шаблон)

    Возвращает словарь со статистикой:
        - mode, engine, output_path
        - rows_total : сколько было строк в шаблоне
        - rows_matched : сколько обновлено по совпадению
        - rows_appended : сколько добавлено в конец (append)
        - rows_cleared : сколько лишних очищено (rebuild)
        - rows_unmatched : номера комнат таблицы без пары в проекте
        - cells_written : всего ячеек записано
    """
    if output_path is None:
        base, ext = os.path.splitext(source_path)
        output_path = base + "_filled" + ext
    out_ext = os.path.splitext(output_path)[1].lower()

    if engine == "auto":
        if _is_com_available():
            engine = "com"
        else:
            engine = "openpyxl"
            if out_ext == ".xls":
                output_path = os.path.splitext(output_path)[0] + ".xlsx"

    if engine == "com":
        return _export_via_com(project, source_path, output_path,
                                overwrite_only_empty, preserve_formulas,
                                mode=mode)
    elif engine == "openpyxl":
        return _export_via_openpyxl(project, source_path, output_path,
                                     overwrite_only_empty, preserve_formulas,
                                     mode=mode)
    else:
        raise ValueError(f"Неизвестный движок: {engine!r}. "
                         f"Допустимо: 'auto', 'com', 'openpyxl'.")
