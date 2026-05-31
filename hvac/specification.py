# -*- coding: utf-8 -*-
"""Спецификация оборудования по ГОСТ 21.110-2013.

Назначение
----------
Формирует сводную таблицу оборудования и материалов в формате
ГОСТ 21.110-2013 «Правила выполнения рабочей документации. Системы
противопожарной защиты и ОВиК». Используется в составе проектной
документации, проходящей экспертизу.

Структура позиции (таблица «Спецификация оборудования»)
-------------------------------------------------------
    Позиция     — порядковый номер
    Обозначение — артикул/каталог изготовителя
    Наименование, тип, обозначение документа на поставку — модель,
                  технические параметры
    Единица     — шт. / м / м² / комплект / кг
    Количество  — N
    Масса ед., кг — справочно
    Примечание

Группировка по разделам (ГОСТ 21.110 п. 5.6):
    1. Системы отопления
    2. Системы вентиляции
    3. Системы кондиционирования
    4. Системы дымоудаления
    5. Системы ГВС
    6. Узлы и трубопроводы

Источник позиций
----------------
    • project.heating_systems   — котлы / источники тепла
    • project.cooling_systems   — чиллеры / источники холода
    • project.ventilation_systems — AHU
    • project.heating_circuits + radiator_picks — приборы отопления
    • project.fancoil_picks     — фанкойлы
    • project.vrf_systems       — VRF внешние и внутренние блоки
    • project.smoke_systems     — СДУ/СПВ
    • project.dhw_systems       — бойлеры / ёмкости ГВС
    • project.heating_hydraulics_results — насосы, баки
    • project.acoustics_results — шумоглушители
    • project.underfloor_loops  — трубы тёплого пола (м)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Dict, List, TYPE_CHECKING

if TYPE_CHECKING:
    from hvac.project import HVACProject


# ============================================================================
# Структуры
# ============================================================================

@dataclass
class SpecificationItem:
    """Одна позиция спецификации по ГОСТ 21.110."""
    position: int = 0
    designation: str = ""        # «Carrier», артикул, ГОСТ — поставщик
    name: str = ""               # «Кассетный фанкойл 42GWC020»
    technical_data: str = ""     # «Q=2.3 кВт, ΔP=42 Па»
    unit: str = "шт."            # шт / м / м² / комплект / кг
    quantity: float = 1.0
    weight_kg: float = 0.0
    note: str = ""
    section: str = "Прочее"      # раздел спецификации


@dataclass
class Specification:
    """Полная спецификация по проекту."""
    project_name: str = ""
    items: List[SpecificationItem] = field(default_factory=list)

    def by_section(self) -> Dict[str, List[SpecificationItem]]:
        out: Dict[str, List[SpecificationItem]] = {}
        for it in self.items:
            out.setdefault(it.section, []).append(it)
        return out

    def renumber(self) -> None:
        """Перенумеровывает позиции после построения, по порядку секций."""
        # Порядок секций по ГОСТ 21.110
        order = [
            "Отопление", "Вентиляция", "Кондиционирование",
            "Дымоудаление и подпор", "ГВС", "Гидравлика и обвязка",
            "Шумоглушители", "Прочее",
        ]
        seen_order = []
        sec_map = self.by_section()
        for s in order:
            if s in sec_map:
                seen_order.append(s)
        for s in sec_map.keys():
            if s not in seen_order:
                seen_order.append(s)

        n = 0
        new_items: List[SpecificationItem] = []
        for s in seen_order:
            for it in sec_map[s]:
                n += 1
                it.position = n
                new_items.append(it)
        self.items = new_items


# ============================================================================
# Сборка позиций по проекту
# ============================================================================

def _add_heating_systems(spec: Specification, project: "HVACProject") -> None:
    for sys in project.heating_systems.values():
        spec.items.append(SpecificationItem(
            designation=f"Котёл {sys.system_type}",
            name=sys.name,
            technical_data=f"t_пд/об = {sys.t_supply:.0f}/{sys.t_return:.0f}°C; "
                            f"КПД {sys.efficiency * 100:.0f}%; "
                            f"топливо {sys.fuel}",
            section="Отопление",
            note=sys.note or "Источник тепла",
        ))


def _add_cooling_systems(spec: Specification, project: "HVACProject") -> None:
    for sys in project.cooling_systems.values():
        spec.items.append(SpecificationItem(
            designation=f"Чиллер {sys.system_type}",
            name=sys.name,
            technical_data=(f"t_пд/об = {sys.t_supply:.0f}/{sys.t_return:.0f}°C; "
                            f"COP {sys.cop:.2f}; хладагент {sys.refrigerant}"),
            section="Кондиционирование",
            note=sys.note or "Источник холода",
        ))


def _add_ventilation_systems(spec: Specification,
                              project: "HVACProject") -> None:
    for ahu in project.ventilation_systems.values():
        load_info = ""
        ahu_load = (project.ahu_loads or {}).get(ahu.name)
        if ahu_load:
            load_info = (f"L={ahu_load.get('supply_m3h', 0):.0f} м³/ч; "
                          f"Q_нагр={ahu_load.get('q_heater_w', 0)/1000:.1f} кВт; "
                          f"Q_охл={ahu_load.get('q_cooler_total_w', 0)/1000:.1f} кВт")
        rec = ""
        if ahu.has_recovery:
            rec = (f"; рекуператор η_з={ahu.recovery_efficiency_winter:.2f}, "
                    f"η_л={ahu.recovery_efficiency_summer:.2f}")
        spec.items.append(SpecificationItem(
            designation="AHU",
            name=ahu.name,
            technical_data=f"{load_info}{rec}",
            section="Вентиляция",
            note=ahu.note or "Приточная установка",
        ))


def _add_smoke_systems(spec: Specification, project: "HVACProject") -> None:
    for sm in project.smoke_systems.values():
        if sm.system_type == "air_supply":
            tech = (f"Подпор {sm.L_smoke_m3h:.0f} м³/ч; "
                    f"P_изб = {sm.pressure_pa:.0f} Па; "
                    f"огнестойкость {sm.fire_rating}")
        else:
            tech = (f"L_дым = {sm.L_smoke_m3h:.0f} м³/ч; "
                    f"компенсация {sm.L_makeup_m3h:.0f} м³/ч; "
                    f"огнестойкость {sm.fire_rating}; "
                    f"t_дыма {sm.t_smoke_C:.0f}°C")
        spec.items.append(SpecificationItem(
            designation=sm.calc_method,
            name=sm.name,
            technical_data=tech,
            section="Дымоудаление и подпор",
            note=sm.note,
        ))


def _add_dhw(spec: Specification, project: "HVACProject") -> None:
    for sys in project.dhw_systems.values():
        spec.items.append(SpecificationItem(
            designation="ГВС",
            name=sys.name,
            technical_data=(f"V_сут = {getattr(sys, 'v_daily_m3', 0):.2f} м³/сут; "
                            f"Q_пик = {getattr(sys, 'q_peak_w', 0)/1000:.1f} кВт; "
                            f"бак {getattr(sys, 'storage_tank_l', 0):.0f} л"),
            section="ГВС",
            note=getattr(sys, "note", ""),
        ))


def _add_radiators(spec: Specification, project: "HVACProject") -> None:
    """Группируем радиаторы по модели — выводим штучный итог."""
    by_model: Dict[str, Dict] = {}
    for sid, pick in (project.radiator_picks or {}).items():
        key = (pick.model.name, pick.sections if pick.model.is_sectional else 0)
        if key not in by_model:
            by_model[key] = {
                "model": pick.model,
                "sections": pick.sections,
                "count": 0,
                "weight": 0.0,
                "q_total": 0.0,
            }
        rec = by_model[key]
        rec["count"] += 1
        rec["q_total"] += pick.actual_power_w
        rec["weight"] += pick.model.weight_kg * (
            pick.sections if pick.model.is_sectional else 1)
    for (name, secs), rec in by_model.items():
        m = rec["model"]
        if m.is_sectional:
            tech = (f"{secs} секц. ×{m.height_mm} мм; "
                    f"Q_секц ном. = {m.q_nominal_w:.0f} Вт; "
                    f"Σ Q факт. = {rec['q_total']:.0f} Вт")
        else:
            tech = (f"{m.height_mm}×{m.length_mm} мм; "
                    f"Q ном. = {m.q_nominal_w:.0f} Вт; "
                    f"Σ Q факт. = {rec['q_total']:.0f} Вт")
        spec.items.append(SpecificationItem(
            designation="EN 442",
            name=name,
            technical_data=tech,
            unit="шт.",
            quantity=rec["count"],
            weight_kg=rec["weight"] / rec["count"] if rec["count"] else 0,
            section="Отопление",
        ))


def _add_fancoils(spec: Specification, project: "HVACProject") -> None:
    by_model: Dict[str, Dict] = {}
    for sid, pick in (project.fancoil_picks or {}).items():
        if pick.model.name not in by_model:
            by_model[pick.model.name] = {"model": pick.model, "count": 0}
        by_model[pick.model.name]["count"] += 1
    for name, rec in by_model.items():
        m = rec["model"]
        spec.items.append(SpecificationItem(
            designation="EN 1397",
            name=name,
            technical_data=(f"Q_х = {m.q_cool_nom_w:.0f} Вт; "
                            f"Q_т = {m.q_heat_nom_w:.0f} Вт; "
                            f"L = {m.air_flow_m3_h:.0f} м³/ч; "
                            f"{m.pipes}-трубный; шум {m.noise_db_a:.0f} дБА"),
            unit="шт.",
            quantity=rec["count"],
            section="Кондиционирование",
        ))


def _add_vrf(spec: Specification, project: "HVACProject") -> None:
    # Внешние блоки
    out_count: Dict[str, int] = {}
    in_count: Dict[str, Dict] = {}
    pipe_lengths: Dict[str, float] = {"6.35": 0, "9.52": 0, "12.7": 0,
                                       "15.88": 0, "19.05": 0,
                                       "22.22": 0, "25.4": 0,
                                       "28.58": 0, "31.75": 0, "34.92": 0}
    from hvac.vrf import pipe_diameters_by_index
    for sys in (project.vrf_systems or {}).values():
        if sys.outdoor:
            out_count[sys.outdoor.name] = out_count.get(
                sys.outdoor.name, 0) + 1
        for a in sys.indoors:
            if a.indoor.name not in in_count:
                in_count[a.indoor.name] = {"unit": a.indoor, "count": 0}
            in_count[a.indoor.name]["count"] += 1
            liq, gas = pipe_diameters_by_index(a.indoor.capacity_index)
            length = a.pipe_length_m or sys.max_pipe_length_to_indoor_m
            key_l = f"{liq:g}"
            key_g = f"{gas:g}"
            pipe_lengths[key_l] = pipe_lengths.get(key_l, 0) + length
            pipe_lengths[key_g] = pipe_lengths.get(key_g, 0) + length
        # Магистраль
        if sys.outdoor:
            liq = sys.outdoor.main_pipe_liquid_mm
            gas = sys.outdoor.main_pipe_gas_mm
            pipe_lengths[f"{liq:g}"] = pipe_lengths.get(
                f"{liq:g}", 0) + sys.main_pipe_length_m
            pipe_lengths[f"{gas:g}"] = pipe_lengths.get(
                f"{gas:g}", 0) + sys.main_pipe_length_m

    for name, n in out_count.items():
        spec.items.append(SpecificationItem(
            designation="VRF outdoor",
            name=name,
            unit="шт.",
            quantity=n,
            section="Кондиционирование",
            note="Внешний блок VRF",
        ))
    for name, rec in in_count.items():
        idu = rec["unit"]
        spec.items.append(SpecificationItem(
            designation="VRF indoor",
            name=name,
            technical_data=(f"{idu.family}; "
                            f"Q_х = {idu.q_cool_w:.0f} Вт; "
                            f"индекс {idu.capacity_index}"),
            unit="шт.",
            quantity=rec["count"],
            section="Кондиционирование",
        ))
    # Медные трубы
    for size, length in pipe_lengths.items():
        if length <= 0:
            continue
        spec.items.append(SpecificationItem(
            designation="Cu",
            name=f"Труба медная Ø{size} мм",
            unit="м",
            quantity=round(length, 1),
            section="Кондиционирование",
            note="Хладагент",
        ))


def _add_hydraulics(spec: Specification, project: "HVACProject") -> None:
    """Насосы, расширительные баки, узлы подпитки."""
    for name, r in (project.heating_hydraulics_results or {}).items():
        if r.pump.selected_model:
            spec.items.append(SpecificationItem(
                designation="Pump",
                name=r.pump.selected_model,
                technical_data=(f"Q = {r.pump.selected_flow_m3_h:.2f} м³/ч; "
                                f"H = {r.pump.selected_head_m:.1f} м; "
                                f"N = {r.pump.selected_power_w:.0f} Вт; "
                                f"контур {name}"),
                section="Гидравлика и обвязка",
            ))
        if r.expansion_tank.selected_model:
            spec.items.append(SpecificationItem(
                designation="ExpTank",
                name=r.expansion_tank.selected_model,
                technical_data=(f"P_max = {r.expansion_tank.p_max_bar:.1f} бар; "
                                f"P_init = {r.expansion_tank.p_init_bar:.1f} бар; "
                                f"V_сист = {r.expansion_tank.system_volume_l:.0f} л"),
                section="Гидравлика и обвязка",
            ))


def _add_silencers(spec: Specification, project: "HVACProject") -> None:
    sil_count: Dict[str, Dict] = {}
    for name, a in (project.acoustics_results or {}).items():
        if not a.silencer_selected:
            continue
        sil = a.silencer_selected
        key = sil.name
        if key not in sil_count:
            sil_count[key] = {"sil": sil, "count": 0}
        sil_count[key]["count"] += 1
    for name, rec in sil_count.items():
        sil = rec["sil"]
        spec.items.append(SpecificationItem(
            designation="Silencer",
            name=name,
            technical_data=f"L = {sil.length_mm} мм; ΔP = {sil.pressure_drop_pa:.0f} Па",
            unit="шт.",
            quantity=rec["count"],
            section="Шумоглушители",
        ))


def _add_underfloor_pipes(spec: Specification,
                           project: "HVACProject") -> None:
    """Метраж труб тёплого пола по типу."""
    by_pipe: Dict[str, float] = {}
    for sid, loop in (project.underfloor_loops or {}).items():
        if loop.pipe is None:
            continue
        name = loop.pipe.name
        by_pipe[name] = by_pipe.get(name, 0.0) + loop.pipe_length_m
    for name, length in by_pipe.items():
        if length <= 0:
            continue
        spec.items.append(SpecificationItem(
            designation="UFH pipe",
            name=f"Труба ТП {name}",
            unit="м",
            quantity=round(length, 0),
            section="Отопление",
            note="Греющий контур тёплого пола",
        ))


# ============================================================================
# Главный фасад
# ============================================================================

def build_specification(project: "HVACProject") -> Specification:
    """Собирает полную спецификацию оборудования и материалов."""
    spec = Specification(project_name=project.params.project_name)
    _add_heating_systems(spec, project)
    _add_cooling_systems(spec, project)
    _add_ventilation_systems(spec, project)
    _add_smoke_systems(spec, project)
    _add_dhw(spec, project)
    _add_radiators(spec, project)
    _add_fancoils(spec, project)
    _add_vrf(spec, project)
    _add_underfloor_pipes(spec, project)
    _add_hydraulics(spec, project)
    _add_silencers(spec, project)
    spec.renumber()
    return spec


# ============================================================================
# Экспорт в Excel
# ============================================================================

def export_specification_xlsx(spec: Specification, path: str) -> None:
    """Сохраняет спецификацию в xlsx по форме ГОСТ 21.110-2013."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side,
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError(
            "Не установлен openpyxl. Выполните: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Спецификация"

    # Заголовок проекта
    ws["A1"] = "СПЕЦИФИКАЦИЯ ОБОРУДОВАНИЯ И МАТЕРИАЛОВ"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Проект: {spec.project_name or '—'}"
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:G2")

    ws["A3"] = "(по ГОСТ 21.110-2013)"
    ws["A3"].font = Font(size=9, italic=True)
    ws.merge_cells("A3:G3")
    ws["A3"].alignment = Alignment(horizontal="center")

    # Заголовок таблицы
    headers = ["Поз.", "Обозначение", "Наименование, тех. данные",
               "Единица", "Кол-во", "Масса ед., кг", "Примечание"]
    ws.append([])   # пустая строка
    ws.append(headers)
    head_row = ws.max_row
    thin = Side(border_style="thin", color="666666")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    for c in ws[head_row]:
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        c.border = border

    # Группировка по разделам
    section_font = Font(bold=True, italic=True, color="333333")
    section_fill = PatternFill("solid", fgColor="DCE6F1")
    section_order = [
        "Отопление", "Вентиляция", "Кондиционирование",
        "Дымоудаление и подпор", "ГВС", "Гидравлика и обвязка",
        "Шумоглушители", "Прочее",
    ]
    by_sec = spec.by_section()
    seen = list(section_order) + [s for s in by_sec if s not in section_order]
    for sec in seen:
        items = by_sec.get(sec)
        if not items:
            continue
        ws.append([f"— Раздел: {sec} —", "", "", "", "", "", ""])
        sec_row = ws.max_row
        for c in ws[sec_row]:
            c.font = section_font
            c.fill = section_fill
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border
        ws.merge_cells(start_row=sec_row, start_column=1,
                        end_row=sec_row, end_column=7)

        for it in items:
            # Совмещаем «Наименование» и «Тех. данные» в одной ячейке
            # как принято в ГОСТ-таблицах с переносом строки.
            if it.technical_data:
                full_name = f"{it.name}\n{it.technical_data}"
            else:
                full_name = it.name
            ws.append([
                it.position,
                it.designation,
                full_name,
                it.unit,
                round(it.quantity, 2) if it.quantity != int(it.quantity)
                else int(it.quantity),
                round(it.weight_kg, 2) if it.weight_kg else "",
                it.note,
            ])
            r = ws.max_row
            for c in ws[r]:
                c.border = border
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # Ширины колонок
    widths = [6, 16, 56, 10, 10, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
